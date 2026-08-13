# -*- coding: utf-8 -*-
"""Generate admin review workbook v1.1 with improved UX."""

import json, hashlib
from datetime import datetime, timezone
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PACKAGE = Path("rules/imports/r01-v1.0")
OUT = Path("data/price-review-r01-v1.1.xlsx")
OUT_JSON = Path("data/price-review-r01-v1.1.json")

YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
RED = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
ORANGE = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
HEADER = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BLUE_FONT = Font(color="0000FF")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

def load():
    with open(PACKAGE / "pricing-rules-excel-r01-v1.0.json", encoding="utf-8") as f:
        return json.load(f)

def hash_file(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True)

def style_editable(ws, row, cols):
    for c in cols:
        ws.cell(row=row, column=c).fill = YELLOW
        ws.cell(row=row, column=c).font = BLUE_FONT

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

def make_dv(ws, formula, col_letter, rows):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "請從下拉選單選擇"
    dv.errorTitle = "無效選項"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{rows}")

# ============================================================
# Main
# ============================================================
data = load()
records = data.get("pricing_source_records", [])
now = datetime.now(timezone.utc).isoformat()
wb = openpyxl.Workbook()

# === Sheet 0: Instructions ===
ws0 = wb.active; ws0.title = "Instructions"
ws0.column_dimensions['A'].width = 80
instructions = [
    ["管理員價格審核工作簿 — 使用說明"],
    [""],
    ["可修改的工作表：Company Price Candidates、Process Rate Candidates、Surface Rate Candidates"],
    ["不可修改的工作表：Supplier Prices（來源資料）、Exceptions（系統生成異常）、Publication Summary（自動統計）"],
    [""],
    ["--- 欄位填寫說明 ---"],
    ["Publish?：輸入 TRUE 或 FALSE"],
    ["Selected Origin Record ID：從下拉選單選擇對應的候選 Record ID，禁止手填"],
    ["Company Price：輸入公司確認單價（必須 > 0）"],
    ["Price Basis：選擇 EXCLUDING_TAX（未稅）或 INCLUDING_TAX（含稅）"],
    ["Effective From：生效日期，格式 YYYY-MM-DD"],
    ["Effective To：失效日期（可留空）"],
    ["Reason：審核理由"],
    ["Approver：審批人姓名"],
    ["Note：備註"],
    [""],
    ["--- 完整材料審核範例 ---"],
    ["Material: A6061-T6 | Spec: (無) | Unit: kg"],
    ["候選1: Tongrui/28/PR-xxx/UNKNOWN | 候選2: Liangwei/35/PR-yyy/UNKNOWN | 候選3: Wendi/25/PR-zzz/UNKNOWN"],
    ["Publish?=TRUE | Selected Origin=PR-xxx | Company Price=28 | Basis=EXCLUDING_TAX | Eff From=2026-08-01"],
    [""],
    ["--- 加工費審核範例 ---"],
    ["Process: CNC | Candidate: 80/hour (WS2-Process)"],
    ["Publish?=TRUE | Company Price=80 | Basis=EXCLUDING_TAX | Eff From=2026-08-01"],
    [""],
    ["--- 表面處理審核範例 ---"],
    ["Surface: 陽極 | Candidate: 20/kg (WS2-Surface)"],
    ["Publish?=TRUE | Company Price=20 | Basis=EXCLUDING_TAX | Eff From=2026-08-01"],
]
for row in instructions:
    ws0.append(row)

# === Sheet 1: Company Price Candidates (one row per key) ===
ws1 = wb.create_sheet("Company Price Candidates")
ws1.freeze_panes = "A2"

# Group records by canonical + spec + unit
mat_groups = defaultdict(list)
for r in records:
    if r.get("target_type") not in ("MATERIAL", "PROFILE"):
        continue
    key = (r.get("canonical_material_code","?"), r.get("material_spec") or "",
           r.get("normalized_unit","kg"))
    mat_groups[key].append(r)

# Build headers
max_suppliers = max(len(v) for v in mat_groups.values()) if mat_groups else 3
admin_cols = ["Publish?", "Selected Origin Record ID", "Company Price", "Price Basis",
              "Effective From", "Effective To", "Reason", "Approver", "Note"]
supplier_cols = []
for i in range(max_suppliers):
    supplier_cols += [f"Supplier{i+1}", f"Price{i+1}", f"RecordID{i+1}", f"Tax{i+1}"]
headers = ["Material", "Spec", "Unit"] + supplier_cols + admin_cols
ws1.append(headers)
style_header(ws1, len(headers))

# Data rows
row_idx = 2
for (mat, spec, unit), recs in sorted(mat_groups.items()):
    row_data = [mat, spec or "", unit]
    # Fill supplier columns
    for r in recs:
        row_data += [r.get("supplier_name",""), r.get("unit_price"), r.get("record_id",""), r.get("tax_inclusion_status","")]
    # Pad remaining supplier columns
    for _ in range(max_suppliers - len(recs)):
        row_data += ["", "", "", ""]
    # Admin fields (empty)
    row_data += [""] * len(admin_cols)
    ws1.append(row_data)
    # Style admin columns as editable (yellow)
    admin_start = 3 + max_suppliers * 4 + 1
    style_editable(ws1, row_idx, range(admin_start, admin_start + len(admin_cols)))
    row_idx += 1

# Data validation for admin columns
last_row = row_idx - 1
# Publish dropdown
make_dv(ws1, '"TRUE,FALSE"', get_column_letter(admin_start), last_row)
# Basis dropdown
make_dv(ws1, '"EXCLUDING_TAX,INCLUDING_TAX"', get_column_letter(admin_start + 3), last_row)

# Selected Origin: per-row dropdown (approximate — in practice would be per-key)
for r in range(2, last_row + 1):
    # Collect record IDs for this row's material
    record_ids = []
    for c in range(5, admin_start, 4):  # RecordID columns
        rid = ws1.cell(row=r, column=c).value
        if rid:
            record_ids.append(str(rid))
    if record_ids:
        dv = DataValidation(type="list", formula1='"' + ','.join(record_ids) + '"', allow_blank=True)
        dv.error = "請從下拉選單選擇對應 Record ID"
        ws1.add_data_validation(dv)
        dv.add(f"{get_column_letter(admin_start + 1)}{r}")

auto_width(ws1)
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

# === Sheet 2: Process Rate Candidates ===
ws2 = wb.create_sheet("Process Rate Candidates")
ws2.freeze_panes = "A2"
proc_admin = ["Publish?", "Company Price", "Price Basis", "Effective From", "Effective To", "Reason", "Approver", "Note"]
proc_headers = ["Process", "Source Org", "Candidate Price", "Unit", "Record ID", "Source Cell"] + proc_admin
ws2.append(proc_headers)
style_header(ws2, len(proc_headers))

proc_recs = [r for r in records if r.get("target_type") in ("PROCESS",) and r.get("process_code")]
for r in proc_recs:
    ws2.append([r.get("process_code"), r.get("source_organization_id",""), r.get("unit_price"),
                 r.get("normalized_unit","hour"), r.get("record_id",""), r.get("source_cell","")] + [""] * 8)
    style_editable(ws2, ws2.max_row, range(7, 15))

make_dv(ws2, '"TRUE,FALSE"', "G", ws2.max_row)
make_dv(ws2, '"EXCLUDING_TAX,INCLUDING_TAX"', "I", ws2.max_row)
auto_width(ws2)
ws2.auto_filter.ref = f"A1:{get_column_letter(len(proc_headers))}{ws2.max_row}"

# === Sheet 3: Surface Rate Candidates ===
ws3 = wb.create_sheet("Surface Rate Candidates")
ws3.freeze_panes = "A2"
surf_admin = ["Publish?", "Company Price", "Price Basis", "Effective From", "Effective To", "Reason", "Approver", "Note"]
surf_headers = ["Surface", "Source Org", "Candidate Price", "Unit", "Record ID", "Source Cell"] + surf_admin
ws3.append(surf_headers)
style_header(ws3, len(surf_headers))

surf_recs = [r for r in records if r.get("target_type") in ("SURFACE",) and r.get("surface_code")]
for r in surf_recs:
    ws3.append([r.get("surface_code"), r.get("source_organization_id",""), r.get("unit_price"),
                 r.get("normalized_unit","kg"), r.get("record_id",""), r.get("source_cell","")] + [""] * 8)
    style_editable(ws3, ws3.max_row, range(7, 15))

make_dv(ws3, '"TRUE,FALSE"', "G", ws3.max_row)
make_dv(ws3, '"EXCLUDING_TAX,INCLUDING_TAX"', "I", ws3.max_row)
auto_width(ws3)
ws3.auto_filter.ref = f"A1:{get_column_letter(len(surf_headers))}{ws3.max_row}"

# === Sheet 4: Supplier Prices (read-only reference) ===
ws4 = wb.create_sheet("Supplier Prices")
ws4.freeze_panes = "A2"
sp_headers = ["Record ID", "Supplier", "Material", "Spec", "Price", "Unit", "Tax", "Status", "Source Cell"]
ws4.append(sp_headers)
style_header(ws4, len(sp_headers))
for r in records:
    if r.get("price_source") == "S":
        ws4.append([r["record_id"], r.get("supplier_name"), r.get("canonical_material_code"),
                     r.get("material_spec"), r.get("unit_price"), r.get("normalized_unit"),
                     r.get("tax_inclusion_status"), r.get("status"), r.get("source_cell")])
# Protect supplier data (read-only)
ws4.protection.sheet = True
ws4.protection.set_password("readonly")  # Simple protection
auto_width(ws4)

# === Sheet 5: Exceptions (deduped) ===
ws5 = wb.create_sheet("Exceptions")
ws5.freeze_panes = "A2"
exc_headers = ["Target", "Canonical Code", "Spec", "Supplier", "Status", "Issue",
                "Publish Allowed", "Admin Resolution", "Resolved?", "Resolution Note"]
ws5.append(exc_headers)
style_header(ws5, len(exc_headers))

BLOCKED = {"CONFLICT", "UNIT_CONFLICT", "UNKNOWN_PRICE", "AMBIGUOUS_MATERIAL_SPEC"}
seen = set()
for r in records:
    if r.get("status") not in BLOCKED:
        continue
    key = (r.get("target_type"), r.get("canonical_material_code") or r.get("process_code") or r.get("surface_code"),
           r.get("material_spec") or "", r.get("supplier_name") or "", r.get("status"))
    if key in seen: continue
    seen.add(key)
    issue = "; ".join(r.get("issues", [])) if isinstance(r.get("issues"), list) else str(r.get("issues", ""))
    ws5.append([r.get("target_type"), r.get("canonical_material_code") or r.get("process_code") or r.get("surface_code"),
                 r.get("material_spec") or "", r.get("supplier_name"), r.get("status"), issue,
                 "FALSE", "", "", ""])
    style_editable(ws5, ws5.max_row, [8, 9, 10])
    for c in range(1, 8):
        ws5.cell(row=ws5.max_row, column=c).fill = RED
make_dv(ws5, '"TRUE,FALSE"', "I", ws5.max_row)
auto_width(ws5)
ws5.auto_filter.ref = f"A1:{get_column_letter(len(exc_headers))}{ws5.max_row}"

# === Sheet 6: Publication Summary (auto-generated, read-only) ===
ws6 = wb.create_sheet("Publication Summary")
summary_data = [
    ["Price Version ID", "R01-COMPANY-PRICE-V1.0-DRAFT"],
    ["Status", "DRAFT"],
    ["Source Package SHA256", hash_file(PACKAGE / "pricing-rules-excel-r01-v1.0.json")],
    ["Generated At", now],
    ["", ""],
    ["Material Candidates", len(mat_groups)],
    ["Process Candidates", len(proc_recs)],
    ["Surface Candidates", len(surf_recs)],
    ["Supplier Price Records", sum(1 for r in records if r.get("price_source") == "S")],
    ["Exceptions (blocked)", len(seen)],
    ["", ""],
    ["Selected Material Count", "=COUNTIF('Company Price Candidates'!D2:D100,\"TRUE\")"],
    ["Selected Process Count", "=COUNTIF('Process Rate Candidates'!G2:G20,\"TRUE\")"],
    ["Selected Surface Count", "=COUNTIF('Surface Rate Candidates'!G2:G20,\"TRUE\")"],
    ["Unresolved Exception Count", f"={len(seen)}"],
    ["Draft Publishable", "AWAITING_ADMIN_REVIEW"],
]
for row in summary_data:
    ws6.append(row)
ws6.protection.sheet = True
auto_width(ws6)

# Save
wb.save(OUT)
print(f"Saved: {OUT}")

# Also save JSON v1.1
review_json = {
    "version": "1.1",
    "generated_at": now,
    "source_package_sha256": hash_file(PACKAGE / "pricing-rules-excel-r01-v1.0.json"),
    "material_candidates": len(mat_groups),
    "process_candidates": len(proc_recs),
    "surface_candidates": len(surf_recs),
    "supplier_records": sum(1 for r in records if r.get("price_source") == "S"),
    "exceptions": len(seen),
    "status": "DRAFT",
}
with open(OUT_JSON, "w", encoding="utf-8") as f:
    json.dump(review_json, f, ensure_ascii=False, indent=2)
print(f"Saved: {OUT_JSON}")
print(f"Materials: {len(mat_groups)} | Process: {len(proc_recs)} | Surface: {len(surf_recs)} | Exceptions: {len(seen)}")
