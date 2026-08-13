# -*- coding: utf-8 -*-
"""Full cell-level pricing audit — column-index based, no Chinese in code."""
import json, re
from pathlib import Path
import openpyxl

EXCEL = r"D:\claude\加工件报价\3.0報價表-R01（機構預估價格）.xlsx"
OUT = Path("data/pricing-import-preview-full.json")

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws1, ws2 = wb["工作表1"], wb["工作表2"]

records = []

def add(r):
    records.append(r)

# === Sheet2: Multi-supplier materials (K4:O33) ===
# Column map: K(11)=通瑞, L(12)=良伟, M(13)=富裕昌, N(14)=廣致誠, O(15)=穩迪
# Material name in J(10), unit in P(16), spec for certain rows
SUP_COLS = {11: "Tongrui", 12: "Liangwei", 13: "Fuyuchang", 14: "Guangzhicheng", 15: "Wendi"}
SPECS = {11: "2mm", 31: "20x30", 32: "30x30", 33: "40x40"}

for row in range(4, 34):
    mat_val = ws2.cell(row=row, column=10).value
    if not mat_val:
        continue
    material = str(mat_val).strip()
    spec = SPECS.get(row)
    unit_raw = str(ws2.cell(row=row, column=16).value or "kg").strip()
    unit = "m" if "米" in unit_raw or unit_raw == "m" else "kg"

    row_has = False
    for col, supp_alias in SUP_COLS.items():
        v = ws2.cell(row=row, column=col).value
        if v is None:
            continue
        row_has = True
        raw = str(v).strip()
        m = re.match(r"(\d+(?:\.\d+)?)", raw)
        price = float(m.group(1)) if m else None

        # Determine status
        status = "PARSED_PENDING_REVIEW"
        issues = []
        if price is None:
            status = "UNKNOWN_PRICE"
        if "/" in raw and ("kg" in unit_raw.lower() or "公斤" in unit_raw):
            status = "UNIT_CONFLICT"
            issues.append(f"price={raw} but unit={unit_raw}")
        if row == 20 and col == 11 and price == 25:
            status = "CONFLICT"
            issues.append("R22 also has SUJ2=15 from same supplier")
        if row == 22 and col == 11 and price == 15:
            status = "CONFLICT"
            issues.append("R20 also has SUJ2=25 from same supplier")

        add({
            "source_file": "3.0報價表-R01.xlsx",
            "source_sheet": "工作表2",
            "source_cell": f"{openpyxl.utils.get_column_letter(col)}{row}",
            "supplier": supp_alias,
            "material_code": material,
            "material_spec": spec,
            "original_value": raw,
            "parsed_value": price,
            "normalized_unit": unit,
            "currency": "CNY",
            "tax_included": False,
            "effective_from": None,
            "status": status,
            "issues": issues,
        })

    if not row_has:
        add({
            "source_file": "3.0報價表-R01.xlsx", "source_sheet": "工作表2",
            "source_cell": f"P{row}", "supplier": None,
            "material_code": material, "material_spec": spec,
            "original_value": None, "parsed_value": None,
            "normalized_unit": unit, "currency": "CNY",
            "tax_included": False, "effective_from": None,
            "status": "UNKNOWN_PRICE",
            "issues": [f"No supplier price for {material} {spec or ''}"],
        })

# === Sheet2: Process prices (T3:T12) ===
PROCS = {3: "車床", 4: "銑床", 5: "磨床", 6: "鉗工", 7: "其它", 8: "放電",
         9: "快絲", 10: "慢絲", 11: "夾頭", 12: "CNC"}
for row, proc in PROCS.items():
    v = ws2.cell(row=row, column=20).value
    st = "PARSED_PENDING_REVIEW" if v else "UNKNOWN_PRICE"
    add({
        "source_file": "3.0報價表-R01.xlsx", "source_sheet": "工作表2",
        "source_cell": f"T{row}", "supplier": "WS2-Process",
        "process_code": proc,
        "original_value": str(v) if v else None, "parsed_value": float(v) if v else None,
        "normalized_unit": "hour", "currency": "CNY",
        "tax_included": False, "effective_from": None,
        "status": st,
        "issues": [f"Empty price for {proc}"] if not v else [],
    })

# === Sheet2: Surface prices (T18:T21) ===
SURFS = {18: "鍍鉻", 19: "熱處理", 20: "陽極", 21: "發黑"}
for row, surf in SURFS.items():
    v = ws2.cell(row=row, column=20).value
    if v:
        add({
            "source_file": "3.0報價表-R01.xlsx", "source_sheet": "工作表2",
            "source_cell": f"T{row}", "supplier": "WS2-Surface",
            "surface_code": surf,
            "original_value": str(v), "parsed_value": float(v),
            "normalized_unit": "kg", "currency": "CNY",
            "tax_included": False, "effective_from": None,
            "status": "PARSED_PENDING_REVIEW", "issues": [],
        })

# === Sheet1: Internal price table (F216:F228) ===
INT_MAT = {
    216: ("電木", "A級", 22), 217: ("電木", "B級", 18), 218: ("電木", "C級", None),
    219: ("鋁", "普通", 25), 220: ("鋁", "A6061T6", 33),
    221: ("黃銅", None, 50), 222: ("鈹銅", None, 220),
    223: ("鋼", "H13", 40), 224: ("鋼", "S136H", 110),
    225: ("鋼", "S136", 65), 226: ("鐵", None, 7),
    227: ("4Cr13", None, 28), 228: ("亞克力", None, 28),
}
for row, (mat, sp, pr) in INT_MAT.items():
    add({
        "source_file": "3.0報價表-R01.xlsx", "source_sheet": "工作表1",
        "source_cell": f"F{row}", "supplier": "INTERNAL_PRICE_TABLE",
        "material_code": mat, "material_spec": sp,
        "original_value": str(pr) if pr else None, "parsed_value": pr,
        "normalized_unit": "kg", "currency": "CNY",
        "tax_included": False, "effective_from": None,
        "status": "PARSED_PENDING_REVIEW" if pr else "UNKNOWN_PRICE", "issues": [],
    })

# === Sheet1: JMD price table (F232:F245) ===
JMD_MAT = {
    232: ("電木", "A級", 16), 233: ("電木", "B級", None), 234: ("電木", "C級", None),
    235: ("鋁", "5052", 25), 236: ("鋁", "A6061T6", 25),
    237: ("黃銅", None, 40), 238: ("鈹銅", None, 170),
    239: ("鋼", "H13", 17), 240: ("鋼", "S136H", 135),
    241: ("鋼", "S136", 55), 242: ("鐵", "S50C/S45C", 5.5),
    243: ("4Cr13", None, 13), 244: ("POM", None, 28), 245: ("亞克力", None, 25),
}
for row, (mat, sp, pr) in JMD_MAT.items():
    add({
        "source_file": "3.0報價表-R01.xlsx", "source_sheet": "工作表1",
        "source_cell": f"F{row}", "supplier": "JMD",
        "material_code": mat, "material_spec": sp,
        "original_value": str(pr) if pr else None, "parsed_value": pr,
        "normalized_unit": "kg", "currency": "CNY",
        "tax_included": False, "effective_from": None,
        "status": "PARSED_PENDING_REVIEW" if pr else "UNKNOWN_PRICE", "issues": [],
    })

# === Summary ===
OUT.parent.mkdir(exist_ok=True)
by_status = {}
by_supplier = {}
for r in records:
    by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    s = r.get("supplier") or "None"
    by_supplier[s] = by_supplier.get(s, 0) + 1

output = {
    "total_records": len(records),
    "pending_review": by_status.get("PARSED_PENDING_REVIEW", 0),
    "conflicts": by_status.get("CONFLICT", 0),
    "unit_conflicts": by_status.get("UNIT_CONFLICT", 0),
    "unknown_prices": by_status.get("UNKNOWN_PRICE", 0),
    "by_supplier": by_supplier,
    "by_status": by_status,
    "records": records,
}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Records: {len(records)}")
for s, c in sorted(by_status.items()): print(f"  {s}: {c}")
for s, c in sorted(by_supplier.items()): print(f"  {s}: {c}")
print(f"Saved: {OUT}")
