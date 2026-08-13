"""Unit tests for BOM Excel Reader."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
import openpyxl

from quotation.infrastructure.excel.bom_reader import BomReader, ColumnMapping


# ============================================================================
# Helpers — create test Excel files in memory
# ============================================================================

def _write_xlsx(wb: openpyxl.Workbook, tmp_path: Path) -> Path:
    """Write workbook to a temp file and return the path."""
    path = tmp_path / "test_bom.xlsx"
    wb.save(str(path))
    return path


def _make_bom_sheet(ws, headers: list[str], data: list[list], header_row: int = 1):
    """Populate a worksheet with headers and data."""
    for ci, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=ci, value=h)
    for ri, row in enumerate(data, header_row + 1):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)


# ============================================================================
# ColumnMapping tests
# ============================================================================

class TestColumnMapping:
    def test_english_headers(self):
        cm = ColumnMapping()
        cells = [(0, "Level"), (1, "Item"), (2, "Description"), (4, "Type"), (6, "Quantity"), (7, "Unit Cost")]
        matches = cm.build_from_row(cells)
        assert matches >= 4
        assert cm.get_col("level") == 0
        assert cm.get_col("item") == 1
        assert cm.get_col("description") == 2
        assert cm.is_valid()

    def test_chinese_headers(self):
        cm = ColumnMapping()
        cells = [(0, "級別"), (1, "料號"), (2, "品名"), (4, "類型"), (6, "數量"), (7, "單價")]
        matches = cm.build_from_row(cells)
        assert cm.get_col("item") == 1
        assert cm.get_col("description") == 2
        assert cm.is_valid()

    def test_mixed_headers_with_spaces(self):
        cm = ColumnMapping()
        # "Item " with trailing space (like the real BOM)
        cells = [(1, "Item "), (2, "Description"), (5, "UOM")]
        matches = cm.build_from_row(cells)
        assert cm.get_col("item") == 1
        assert cm.get_col("uom") == 5

    def test_missing_required_fields(self):
        cm = ColumnMapping()
        cells = [(0, "Level"), (3, "Notes")]
        cm.build_from_row(cells)
        assert not cm.is_valid()  # Missing item + description

    def test_case_insensitive(self):
        cm = ColumnMapping()
        cells = [(0, "LEVEL"), (1, "ITEM"), (2, "description")]
        cm.build_from_row(cells)
        assert cm.get_col("level") == 0
        assert cm.get_col("item") == 1


# ============================================================================
# BomReader — unit tests with generated Excel files
# ============================================================================

class TestBomReaderBasic:
    """Basic reading with standard English headers."""

    @pytest.fixture
    def sample_xlsx(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        _make_bom_sheet(
            ws,
            ["Level", "Item", "Description", "Type", "Uom", "Quantity", "Unit Cost", "Extended Cost", "Notes"],
            [
                [0, "UA001", "Finished product", "Finished good", "ST", 1, 0, 0, ""],
                [1, "UB001", "Sub-assembly A", "Subassembly", "ST", 2, 500, 1000, ""],
                [2, "UC001", "Machined part X;S50C;J003;928*796*15", "Purchased item", "EA", 1, 1425, 1425, "熱處理"],
                [2, "UC002", "Purchased sensor", "Purchased item", "EA", 3, 370.8, 1112.4, "品牌:台達"],
            ],
        )
        return _write_xlsx(wb, tmp_path)

    def test_read_entries(self, sample_xlsx):
        reader = BomReader()
        sheet = reader.read(sample_xlsx)
        assert sheet.total_rows == 4
        assert len(sheet.entries) == 4

    def test_source_tracking(self, sample_xlsx):
        reader = BomReader()
        sheet = reader.read(sample_xlsx)
        entry = sheet.entries[2]  # UC001
        assert entry.source_file == str(sample_xlsx)
        assert entry.source_sheet == "BOM"
        assert entry.source_row >= 3

    def test_item_values(self, sample_xlsx):
        reader = BomReader()
        sheet = reader.read(sample_xlsx)
        items = [e.item for e in sheet.entries]
        assert "UC001" in items
        assert "UC002" in items

    def test_cost_parsing(self, sample_xlsx):
        reader = BomReader()
        sheet = reader.read(sample_xlsx)
        uc001 = [e for e in sheet.entries if e.item == "UC001"][0]
        assert uc001.unit_cost == 1425.0
        assert uc001.extended_cost == 1425.0


class TestBomReaderChineseHeaders:
    """Reading with Chinese column headers."""

    @pytest.fixture
    def chinese_xlsx(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "物料清單"
        _make_bom_sheet(
            ws,
            ["級別", "料號", "品名", "類型", "單位", "數量", "單價", "總價", "備註"],
            [
                [2, "UC001", "CNC加工件", "加工件", "EA", 1, 500, 500, ""],
                [2, "UC002", "車削件", "加工件", "EA", 2, 300, 600, ""],
            ],
        )
        return _write_xlsx(wb, tmp_path)

    def test_chinese_headers(self, chinese_xlsx):
        reader = BomReader()
        sheet = reader.read(chinese_xlsx)
        assert sheet.total_rows == 2


class TestBomReaderBlankRows:
    """Handling of blank rows and invalid data."""

    @pytest.fixture
    def blank_row_xlsx(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        _make_bom_sheet(
            ws,
            ["Level", "Item", "Description"],
            [
                [1, "UB001", "Assembly"],
                [1, "", ""],           # blank row → skip
                [1, "None", "None"],   # None values → skip
                [2, "UC001", "Part A"],
                [2, "", ""],           # blank → skip
                [2, "UC002", "Part B"],
            ],
        )
        return _write_xlsx(wb, tmp_path)

    def test_blank_rows_skipped(self, blank_row_xlsx):
        reader = BomReader()
        sheet = reader.read(blank_row_xlsx)
        # Should have 3 entries: UB001, UC001, UC002
        assert sheet.total_rows == 3
        items = [e.item for e in sheet.entries]
        assert items == ["UB001", "UC001", "UC002"]


class TestBomReaderMissingColumns:
    """Graceful handling of missing optional columns."""

    @pytest.fixture
    def minimal_xlsx(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Only Item and Description — no Level, no cost columns
        _make_bom_sheet(
            ws,
            ["Item", "Description"],
            [
                ["UC001", "A simple part"],
                ["UC002", "Another part"],
            ],
        )
        return _write_xlsx(wb, tmp_path)

    def test_minimal_columns(self, minimal_xlsx):
        reader = BomReader()
        sheet = reader.read(minimal_xlsx)
        # With only 2 known columns, header detection may use a lower threshold
        # The reader falls back to column-index defaults when header match < 3
        # In that case the header row itself may be read as data
        # Verify we at least get data rows (not just the header)
        assert sheet.total_rows >= 2
        # All entries should have valid item values (not header text like "Item")
        for entry in sheet.entries:
            assert entry.item not in ("Item", "Description", "Level")
            assert entry.item != ""


class TestBomReaderMultiSheet:
    """Multiple sheets — auto-detect BOM sheet."""

    @pytest.fixture
    def multisheet_xlsx(self, tmp_path):
        wb = openpyxl.Workbook()
        # Sheet 1: not a BOM
        ws1 = wb.active
        ws1.title = "Cover"
        ws1.cell(row=1, column=1, value="Project Info")

        # Sheet 2: BOM
        ws2 = wb.create_sheet("BOM")
        _make_bom_sheet(
            ws2,
            ["Level", "Item", "Description", "Unit Cost"],
            [[2, "UC001", "Part", 100.0]],
        )

        # Sheet 3: another BOM
        ws3 = wb.create_sheet("BOM2")
        _make_bom_sheet(
            ws3,
            ["Level", "Item", "Description", "Unit Cost"],
            [[2, "UC002", "Part2", 200.0]],
        )
        return _write_xlsx(wb, tmp_path)

    def test_auto_detect_bom_sheet(self, multisheet_xlsx):
        reader = BomReader()
        sheet = reader.read(multisheet_xlsx)  # No sheet specified
        assert sheet.total_rows >= 1

    def test_read_all_sheets(self, multisheet_xlsx):
        reader = BomReader()
        sheets = reader.read_all_sheets(multisheet_xlsx)
        assert len(sheets) == 2  # Should find BOM and BOM2


class TestBomReaderHeaderDetection:
    """Header row auto-detection with company info rows above."""

    @pytest.fixture
    def header_offset_xlsx(self, tmp_path):
        """Simulate real BOM: company info rows 1-3, header at row 4, data from row 5."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Rows 1-3: company info
        ws.cell(row=1, column=1, value="MPTZ")
        ws.cell(row=1, column=3, value="公司名稱")
        ws.cell(row=2, column=3, value="產品描述")
        ws.cell(row=4, column=1, value="Level")
        ws.cell(row=4, column=2, value="Item")
        ws.cell(row=4, column=3, value="Description")
        ws.cell(row=4, column=5, value="Type")
        ws.cell(row=4, column=6, value="Uom")
        ws.cell(row=4, column=7, value="Quantity")
        ws.cell(row=4, column=8, value="Unit Cost")
        ws.cell(row=4, column=9, value="Extended Cost")
        # Data
        ws.cell(row=5, column=1, value=0)
        ws.cell(row=5, column=2, value="UA001")
        ws.cell(row=5, column=3, value="Finished Product")
        ws.cell(row=5, column=5, value="Finished good")
        ws.cell(row=5, column=6, value="ST")
        ws.cell(row=6, column=2, value="UC001")
        ws.cell(row=6, column=3, value="Machined Part")
        return _write_xlsx(wb, tmp_path)

    def test_header_detected_after_info_rows(self, header_offset_xlsx):
        reader = BomReader()
        sheet = reader.read(header_offset_xlsx)
        assert sheet.total_rows >= 2

    def test_source_row_correct(self, header_offset_xlsx):
        reader = BomReader()
        sheet = reader.read(header_offset_xlsx)
        # UC001 should be at row 6
        uc001 = [e for e in sheet.entries if e.item == "UC001"][0]
        assert uc001.source_row == 6


class TestBomReaderFromYamlConfig:
    """Reader created from YAML config file."""

    @pytest.fixture
    def custom_config_yaml(self, tmp_path):
        import yaml
        config = {
            "fields": {
                "item": ["PartNumber", "PN"],
                "description": ["PartDesc", "Description"],
                "unit_cost": ["Price", "Cost"],
            },
            "header_scan": {"min_row": 1, "max_row": 5},
            "header_min_matches": 2,
        }
        path = tmp_path / "custom-mapping.yaml"
        with open(path, "w", encoding="utf-8") as f:
            yaml.dump(config, f)
        return path

    @pytest.fixture
    def custom_header_xlsx(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        _make_bom_sheet(
            ws,
            ["PN", "PartDesc", "Price"],
            [
                ["UC001", "Custom part", 99.99],
            ],
        )
        return _write_xlsx(wb, tmp_path)

    def test_custom_mapping(self, custom_config_yaml, custom_header_xlsx):
        reader = BomReader.from_yaml(custom_config_yaml)
        sheet = reader.read(custom_header_xlsx)
        assert sheet.total_rows == 1
        assert sheet.entries[0].unit_cost == 99.99


class TestBomReaderMergedCells:
    """Merged cells should be handled gracefully."""

    @pytest.fixture
    def merged_xlsx(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Merge cells A1:B1 for a title
        ws.merge_cells("A1:B1")
        ws.cell(row=1, column=1, value="BOM Title")

        # Header row
        _make_bom_sheet(
            ws,
            ["Level", "Item", "Description"],
            [[2, "UC001", "Part with merged cell above"]],
            header_row=2,
        )
        return _write_xlsx(wb, tmp_path)

    def test_merged_cells_handled(self, merged_xlsx):
        reader = BomReader()
        sheet = reader.read(merged_xlsx)
        assert sheet.total_rows >= 1
