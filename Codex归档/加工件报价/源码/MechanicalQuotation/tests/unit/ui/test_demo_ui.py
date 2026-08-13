"""Smoke tests for the quotation demo UI module."""

from __future__ import annotations

import json
import tempfile
from decimal import Decimal
from pathlib import Path

import pytest

from quotation.domain.quote import PriceSource, QuoteConfidence, QuoteItem
from quotation.ui.viewmodels import (
    QuoteItemViewModel,
    QuoteViewModel,
    TaxResult,
)


# ============================================================================
# Test 1: UI module imports
# ============================================================================

class TestUIModuleImports:
    def test_viewmodels_imports(self):
        """UI module can be imported."""
        from quotation.ui import viewmodels
        assert viewmodels is not None

    def test_widgets_imports(self):
        """Widgets module can be imported (tkinter required)."""
        import tkinter
        from quotation.ui import widgets
        assert widgets is not None

    def test_demo_app_imports(self):
        """Demo app module can be imported."""
        import tkinter
        from quotation.ui import demo_app
        assert demo_app is not None


# ============================================================================
# Test 2 & 3: Tax calculation
# ============================================================================

class TestTaxCalculation:
    def test_tax_17_percent_correct(self):
        """TaxCalculator gives 17% rate."""
        items = [
            QuoteItem(line_id="M1", category="material", name="Steel", amount=1000,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
        ]
        tax = TaxResult.calculate(items, Decimal("0.17"))
        assert tax.tax_rate == Decimal("0.17")

    def test_total_including_tax_equals_subtotal_times_1_17(self):
        """Tax-inclusive total = subtotal × 1.17."""
        items = [
            QuoteItem(line_id="M1", category="material", name="Steel", amount=1000,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
            QuoteItem(line_id="P1", category="process", name="CNC", amount=500,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
        ]
        tax = TaxResult.calculate(items, Decimal("0.17"))
        expected_subtotal = Decimal("1500.00")
        assert tax.subtotal_excluding_tax == expected_subtotal
        expected_total = Decimal("1755.00")  # 1500 * 1.17
        assert tax.total_including_tax == expected_total

    def test_unknown_items_excluded_from_tax(self):
        """Unknown (source=U) items should not contribute to tax base."""
        items = [
            QuoteItem(line_id="M1", category="material", name="Steel", amount=1000,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
            QuoteItem(line_id="U1", category="process", name="Magic", amount=0,
                      source=PriceSource.U, confidence=QuoteConfidence.UNCERTAIN),
        ]
        tax = TaxResult.calculate(items, Decimal("0.17"))
        assert tax.subtotal_excluding_tax == Decimal("1000.00")
        assert tax.total_including_tax == Decimal("1170.00")


# ============================================================================
# Test 4: Unknown amount display
# ============================================================================

class TestUnknownDisplay:
    def test_unknown_amount_displays_as_pending(self):
        """Unknown items show '—' not '¥0.00'."""
        item = QuoteItem(
            line_id="U1", category="process", name="TAP", amount=0,
            source=PriceSource.U, confidence=QuoteConfidence.UNCERTAIN,
        )
        vm = QuoteItemViewModel(item=item, index=1)
        assert vm.is_unknown
        assert vm.display_amount == "—"
        assert vm.display_unit_price == "—"
        assert vm.status_label == "待確認"
        assert "unknown" in vm.row_tags

    def test_known_zero_amount_displays_normally(self):
        """Known source with zero amount displays ¥0.00."""
        item = QuoteItem(
            line_id="Z1", category="purchased", name="FreePart", amount=0,
            source=PriceSource.M, confidence=QuoteConfidence.MEDIUM,
        )
        vm = QuoteItemViewModel(item=item, index=1)
        assert not vm.is_unknown
        assert vm.display_amount == "¥0.00"
        assert vm.status_label == "已確認"


# ============================================================================
# Test 5 & 6: J003 / W001 pipeline → ViewModel
# ============================================================================

class TestDemoPipeline:
    def test_j003_viewmodel_generates(self):
        """Full J003 pipeline → QuoteViewModel with tax."""
        from quotation.ui.demo_app import run_quotation_pipeline

        quote, feature_summary, error = run_quotation_pipeline("J003")
        assert error is None, f"Pipeline error: {error}"
        assert quote is not None
        assert len(quote.items) > 0
        assert quote.part_number == "UC1000005854"
        assert quote.quotation_status == "COMPLETE"
        assert quote.cost_completion == 100.0

        # Build ViewModel
        tax = TaxResult.calculate(quote.items, Decimal("0.17"))
        vm = QuoteViewModel(quote=quote, tax=tax)
        assert vm.status_color == "green"
        assert vm.status_text == "報價完整"
        assert len(vm.items_vm) > 0

    def test_w001_viewmodel_generates(self):
        """Full W001 pipeline → QuoteViewModel with tax."""
        from quotation.ui.demo_app import run_quotation_pipeline

        quote, feature_summary, error = run_quotation_pipeline("W001")
        assert error is None, f"Pipeline error: {error}"
        assert quote is not None
        assert len(quote.items) > 0
        assert quote.part_number == "UC2020083221"
        assert quote.quotation_status == "INCOMPLETE"
        assert quote.cost_completion == pytest.approx(85.7, rel=0.01)

        # Build ViewModel
        tax = TaxResult.calculate(quote.items, Decimal("0.17"))
        vm = QuoteViewModel(quote=quote, tax=tax)
        assert vm.status_color == "orange"
        assert "待確認" in vm.status_text
        assert len(vm.unknown_items) >= 1
        assert len(vm.items_vm) > 0


# ============================================================================
# Test 7: Excel export
# ============================================================================

class TestExcelExport:
    def test_excel_export_succeeds(self):
        """Export creates valid .xlsx file."""
        import openpyxl

        from quotation.ui.demo_app import run_quotation_pipeline

        quote, _, error = run_quotation_pipeline("J003")
        assert error is None
        assert quote is not None

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmppath = Path(f.name)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "報價單"

            # Write basic data
            ws.cell(row=1, column=1, value="機械加工件智能報價系統 — 報價單")
            ws.cell(row=3, column=1, value="圖號")
            ws.cell(row=3, column=2, value=quote.part_number or "")

            # Write items
            row = 8
            for i, item in enumerate(quote.items, 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=item.name)
                ws.cell(row=row, column=3, value=item.source.value)
                if item.source == PriceSource.U:
                    ws.cell(row=row, column=7, value="待確認")
                else:
                    ws.cell(row=row, column=7, value=item.amount)
                row += 1

            # Tax
            tax = TaxResult.calculate(quote.items, Decimal("0.17"))
            ws.cell(row=row + 1, column=1, value="未稅小計")
            ws.cell(row=row + 1, column=2, value=float(tax.subtotal_excluding_tax))
            ws.cell(row=row + 2, column=1, value="含稅總價")
            ws.cell(row=row + 2, column=2, value=float(tax.total_including_tax))

            wb.save(str(tmppath))
            assert tmppath.exists()
            assert tmppath.stat().st_size > 0

            # Verify it's a valid xlsx
            wb2 = openpyxl.load_workbook(str(tmppath))
            assert "報價單" in wb2.sheetnames
        finally:
            tmppath.unlink(missing_ok=True)


# ============================================================================
# UI Startup Smoke Test
# ============================================================================

class TestUIStartup:
    def test_demo_app_starts_and_destroys(self):
        """DemoApp can be created, renders, and destroyed without error."""
        import tkinter as tk

        # Skip in headless environments
        try:
            root = tk.Tk()
            root.destroy()
        except tk.TclError:
            pytest.skip("No display available")

        from quotation.ui.demo_app import DemoApp
        app = DemoApp()
        app.update_idletasks()

        # Verify content area exists
        assert app._content_area is not None
        # Verify default page is "新建報價"
        assert app._content is not None
        # NewQuotePage should have toolbar
        assert hasattr(app._content, '_toolbar')

        app.destroy()

    def test_chinese_source_labels(self):
        """Source labels use Chinese display names, not internal codes."""
        from quotation.ui.viewmodels import SOURCE_LABELS, SOURCE_SHORT, STATUS_DISPLAY

        # Source labels should NOT contain raw codes
        assert "公司核准價格" in SOURCE_LABELS["C"]
        assert "歷史成交價格" in SOURCE_LABELS["H"]
        assert "系統估算價格" in SOURCE_LABELS["E"]
        assert "價格待確認" in SOURCE_LABELS["U"]
        assert "人工確認價格" in SOURCE_LABELS["M"]

        # Status should use Chinese
        assert STATUS_DISPLAY["COMPLETE"] == "報價完整"
        assert STATUS_DISPLAY["INCOMPLETE"] == "部分價格待確認"
        assert STATUS_DISPLAY["REVIEW_REQUIRED"] == "需要人工審核"
