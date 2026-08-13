"""Batch Excel export for quotation results.

Produces multi-sheet workbooks with Summary, Quote Details,
Review Required, Source Files, and Trace information.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from quotation.domain.quote import PriceSource, QuoteItem
from quotation.application.quotation_service import JobStatus, QuoteJobResult, TaxResult

# Font family for Excel
_FONT = "Microsoft YaHei UI"


def export_batch_excel(
    results: list[QuoteJobResult],
    output_path: str | Path,
    scan_directory: str = "",
) -> Path:
    """Export batch quotation results to a multi-sheet Excel workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Styles
    title_font = Font(name=_FONT, size=14, bold=True)
    header_font = Font(name=_FONT, size=10, bold=True, color="ffffff")
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    normal_font = Font(name=_FONT, size=9)
    warn_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_headers(ws, headers: list[str], row: int) -> int:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        return row + 1

    # ==================================================================
    # Sheet 1: Summary
    # ==================================================================
    ws = wb.active
    ws.title = "Summary"
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1, value="機械加工件智能報價系統 — 批量報價彙總").font = title_font
    row += 2

    ws.cell(row=row, column=1, value=f"掃描目錄: {scan_directory}").font = normal_font
    row += 1
    ws.cell(row=row, column=1, value=f"生成時間: {datetime.now():%Y-%m-%d %H:%M:%S}").font = normal_font
    row += 1
    ws.cell(row=row, column=1, value=f"總文件數: {len(results)}").font = normal_font
    row += 2

    headers = [
        "圖號", "文件名", "原始路徑", "配對文件", "解析狀態", "報價狀態",
        "cost_completion", "未知項數", "未稅小計", "稅率", "稅額", "含稅總價",
        "規則版本", "價格版本", "AI使用", "生成時間",
    ]
    row = _write_headers(ws, headers, row)

    for jr in results:
        q = jr.quote
        matched = ", ".join(
            f.file_name for f in jr.bundle.files if f != jr.bundle.geometry_source
        ) or "-"
        geom = jr.bundle.geometry_source
        geom_name = geom.file_name if geom else "-"
        values = [
            jr.drawing_number,
            geom_name,
            str(geom.full_path) if geom else "-",
            matched,
            jr.status,
            "COMPLETE" if jr.is_complete else jr.status,
            f"{jr.cost_completion:.1f}%",
            jr.unknown_item_count,
            float(jr.subtotal_excluding_tax),
            "17%",
            float(jr.tax.tax_amount) if jr.tax else 0,
            float(jr.total_including_tax),
            q.price_version if q else "-",
            q.price_version if q else "-",
            "Y" if jr.ai_used else "N",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = normal_font
            cell.border = thin_border
            if jr.status in (JobStatus.COMPLETE,):
                cell.fill = green_fill
            elif jr.status in (JobStatus.PARSE_FAILED, JobStatus.QUOTE_FAILED):
                cell.fill = red_fill
        row += 1

    # Column widths
    for col, w in enumerate([16, 28, 50, 28, 14, 16, 12, 10, 14, 8, 14, 14, 12, 12, 8, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 2: Quote Details
    # ==================================================================
    ws2 = wb.create_sheet("Quote Details")
    row = 1
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws2.cell(row=row, column=1, value="報價明細").font = title_font
    row += 2

    detail_headers = ["圖號", "報價項目", "來源", "數量", "單位", "單價", "未稅金額", "狀態"]
    row = _write_headers(ws2, detail_headers, row)

    for jr in results:
        if jr.quote is None:
            continue
        for item in jr.quote.items:
            is_u = item.source == PriceSource.U
            values = [
                jr.drawing_number,
                item.name,
                item.source.value,
                "-" if is_u else item.quantity,
                "-" if is_u else item.unit,
                "-" if is_u else item.unit_price,
                "待確認" if is_u else item.amount,
                "待確認" if is_u else "已確認",
            ]
            for col, v in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=v)
                cell.font = normal_font
                cell.border = thin_border
                if is_u:
                    cell.fill = warn_fill
            row += 1

    for col, w in enumerate([16, 30, 8, 8, 8, 10, 14, 10], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 3: Review Required
    # ==================================================================
    ws3 = wb.create_sheet("Review Required")
    row = 1
    ws3.cell(row=row, column=1, value="待審查項目").font = title_font
    row += 2

    review_headers = ["圖號", "缺失資訊", "未知成本項", "解析警告", "AI建議", "建議處理"]
    row = _write_headers(ws3, review_headers, row)

    for jr in results:
        if jr.status in (JobStatus.COMPLETE,):
            continue
        unknown_items = []
        if jr.quote:
            unknown_items = [
                i.name for i in jr.quote.items if i.source == PriceSource.U
            ]
        ai_str = ""
        if jr.ai_suggestions:
            ai_str = str(jr.ai_suggestions.get("missing_fields", []))
        values = [
            jr.drawing_number,
            ", ".join(jr.warnings) if jr.warnings else "-",
            ", ".join(unknown_items) if unknown_items else "-",
            ", ".join(jr.warnings) if jr.warnings else "-",
            ai_str if ai_str else "-",
            "人工確認材料、厚度、表面處理及加工方式" if unknown_items else "-",
        ]
        for col, v in enumerate(values, 1):
            cell = ws3.cell(row=row, column=col, value=v)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = warn_fill
        row += 1

    for col, w in enumerate([16, 25, 30, 25, 30, 30], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 4: Source Files
    # ==================================================================
    ws4 = wb.create_sheet("Source Files")
    row = 1
    ws4.cell(row=row, column=1, value="原始文件清單").font = title_font
    row += 2

    src_headers = ["原始文件名", "文件類型", "完整路徑", "配對關係", "幾何來源", "輔助來源"]
    row = _write_headers(ws4, src_headers, row)

    for jr in results:
        geom = jr.bundle.geometry_source
        for f in jr.bundle.files:
            values = [
                f.file_name,
                f.extension.upper().lstrip('.'),
                str(f.full_path),
                "MATCHED" if jr.bundle.match_status == MatchStatus.MATCHED else "UNMATCHED",
                "Y" if f == geom else "N",
                "Y" if f.is_pdf else "N",
            ]
            for col, v in enumerate(values, 1):
                ws4.cell(row=row, column=col, value=v).font = normal_font
                ws4.cell(row=row, column=col).border = thin_border
            row += 1

    for col, w in enumerate([30, 10, 55, 12, 10, 10], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 5: Trace
    # ==================================================================
    ws5 = wb.create_sheet("Trace")
    row = 1
    ws5.cell(row=row, column=1, value="價格來源追蹤").font = title_font
    row += 2

    trace_headers = [
        "圖號", "項目", "quote_price_source", "resolution_source",
        "price_version_id", "company_price_id", "origin_price_record_id",
        "origin_supplier_id", "price_basis", "fallback_warning",
    ]
    row = _write_headers(ws5, trace_headers, row)

    for jr in results:
        if jr.quote is None:
            continue
        for item in jr.quote.items:
            values = [
                jr.drawing_number,
                item.name,
                item.quote_price_source or "-",
                item.resolution_source or "-",
                item.price_version_id or "-",
                item.company_price_id or "-",
                item.origin_price_record_id or "-",
                item.origin_supplier_id or "-",
                item.price_basis or "-",
                "Y" if item.fallback_warning else "N",
            ]
            for col, v in enumerate(values, 1):
                ws5.cell(row=row, column=col, value=v).font = normal_font
                ws5.cell(row=row, column=col).border = thin_border
            row += 1

    for col, w in enumerate([16, 28, 16, 24, 16, 22, 22, 18, 14, 12], 1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(str(output_path))
    return output_path


def export_single_excel(
    result: QuoteJobResult,
    output_path: str | Path,
) -> Path:
    """Export a single quotation to Excel (same format as batch but single-item)."""
    return export_batch_excel([result], output_path)


# Re-export for convenience
MatchStatus = __import__('quotation.application.file_scanner', fromlist=['MatchStatus']).MatchStatus
