#!/usr/bin/env python3
"""Generate an LFAF four-category machining quotation workbook.

Each drawing is classified as one of:
  加工件, 钣金件, 焊接件, 型材组装件

The script writes one row to the matching calculation table and preserves the
user-approved workbook formulas, rates, tables, formatting, summary, and base
parameters. It also creates one combined 图纸附件 sheet for traceability and
manual-review reminders.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.properties import CalcProperties


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
DEFAULT_TEMPLATE = SKILL_DIR / "assets" / "LFAF_四类零件统一报价计算模板_20260807.xlsx"

ATTACH_SHEET = "图纸附件"
BASE_PARAMETER_SHEET = "基础参数"
SUMMARY_SHEET = "统一汇总"
CATEGORY_SHEETS = ("加工件", "钣金件", "焊接件", "型材组装件")
DATA_START_ROW = 7
MATERIAL_LIST_NAME = "LFAF_MaterialList"

THIN_SIDE = Side(style="thin", color="000000")
FULL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
SEVERITY_LABELS = {"high": "高", "medium": "中", "low": "低"}
SEVERITY_FILLS = {
    "high": PatternFill(fill_type="solid", fgColor="F8CBAD"),
    "medium": PatternFill(fill_type="solid", fgColor="FFE699"),
    "low": PatternFill(fill_type="solid", fgColor="C6E0B4"),
}

CATEGORY_ALIASES = {
    "加工件": "加工件", "机加工件": "加工件", "機加工件": "加工件", "机械加工": "加工件",
    "machining": "加工件", "machinedpart": "加工件", "cnc": "加工件",
    "钣金件": "钣金件", "鈑金件": "钣金件", "钣金": "钣金件", "鈑金": "钣金件",
    "sheetmetal": "钣金件", "sheetmetalpart": "钣金件",
    "焊接件": "焊接件", "焊接": "焊接件", "焊件": "焊接件",
    "welding": "焊接件", "weldment": "焊接件", "weldedpart": "焊接件",
    "型材组装件": "型材组装件", "型材組裝件": "型材组装件", "型材件": "型材组装件",
    "型材组装": "型材组装件", "型材組裝": "型材组装件", "profileassembly": "型材组装件",
    "extrusionassembly": "型材组装件", "profile": "型材组装件",
}

MACHINING_PROCESS_HEADERS = {
    "車床": "车床h", "车床": "车床h", "lathe": "车床h",
    "銑床": "铣床h", "铣床": "铣床h", "milling": "铣床h",
    "磨床": "磨床h", "grinding": "磨床h",
    "鉗工": "钳工h", "钳工": "钳工h", "fitter": "钳工h",
    "其它": "其它h", "其他": "其它h", "other": "其它h",
    "放電": "放电h", "放电": "放电h", "edm": "放电h",
    "快絲": "快丝h", "快丝": "快丝h", "wirecutfast": "快丝h",
    "慢絲": "慢丝h", "慢丝": "慢丝h", "wirecutslow": "慢丝h",
    "夾頭": "夹头/治具费", "夹头": "夹头/治具费", "治具": "夹头/治具费",
    "fixture": "夹头/治具费", "fixturecost": "夹头/治具费",
    "精雕機": "精雕机h", "精雕机": "精雕机h", "engraving": "精雕机h",
    "cnc": "CNC h",
    "鍍鉻": "镀铬kg", "镀铬": "镀铬kg", "chrome": "镀铬kg",
    "熱處理": "热处理kg", "热处理": "热处理kg", "heattreatment": "热处理kg",
    "陽極": "阳极kg", "阳极": "阳极kg", "anodize": "阳极kg", "anodizing": "阳极kg",
    "發黑": "发黑kg", "发黑": "发黑kg", "blackening": "发黑kg",
}


def as_num(value, default=None):
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace("mm", "").replace("MM", "").strip())
    except Exception:
        return default


def round_up_to_multiple(value: float, multiple: int = 5) -> float:
    import math
    if value is None:
        return None
    try:
        number = float(value)
    except Exception:
        return None
    if number <= 0:
        return number
    return math.ceil(number / multiple) * multiple


def calc_stock_dimensions(part: dict) -> tuple[object, object, object]:
    raw_l = as_num(part.get("length"))
    raw_w = as_num(part.get("width"))
    raw_h = as_num(part.get("height"))
    stock_l = lookup_part_value(part, "stock_length", "stockLength", "备料长mm")
    stock_w = lookup_part_value(part, "stock_width", "stockWidth", "备料宽mm")
    stock_h = lookup_part_value(part, "stock_height", "stockHeight", "备料高mm")
    stock_l = as_num(stock_l)
    stock_w = as_num(stock_w)
    stock_h = as_num(stock_h)
    allowance = as_num(lookup_part_value(part, "stock_allowance_lw", "stockAllowanceLW"), 5)
    if stock_l is None and raw_l is not None:
        stock_l = round_up_to_multiple(raw_l + allowance, 5)
    if stock_w is None and raw_w is not None:
        stock_w = round_up_to_multiple(raw_w + allowance, 5)
    if stock_h is None and raw_h is not None:
        stock_h = round_up_to_multiple(raw_h, 5)
    return stock_l, stock_w, stock_h


def normalize_key(value) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch not in " _-/\\()（）")


def part_containers(part: dict) -> list[dict]:
    containers = [part]
    for key in (
        "calculation_inputs", "type_inputs", "machining_inputs", "sheet_metal_inputs",
        "welding_inputs", "profile_assembly_inputs", "加工件参数", "钣金件参数",
        "焊接件参数", "型材组装件参数",
    ):
        value = part.get(key)
        if isinstance(value, dict):
            containers.append(value)
    return containers


def lookup_part_value(part: dict, *aliases, default=None):
    containers = part_containers(part)
    for container in containers:
        for alias in aliases:
            if alias in container and container[alias] not in (None, ""):
                return container[alias]
    alias_keys = {normalize_key(alias) for alias in aliases}
    for container in containers:
        normalized = {normalize_key(key): value for key, value in container.items()}
        for alias_key in alias_keys:
            value = normalized.get(alias_key)
            if value not in (None, ""):
                return value
    return default


def normalize_material(material: str | None) -> str | None:
    if not material:
        return None
    text = str(material).strip().replace(" ", "")
    mapping = {
        "A6061": "A6061T6", "6061": "A6061T6", "AL6061": "A6061T6",
        "鋁6061": "A6061T6", "铝6061": "A6061T6", "AL": "鋁", "铝": "鋁",
        "亚克力": "亚克力", "亞克力": "亚克力", "电木": "电木A級", "電木": "电木A級",
    }
    return mapping.get(text.upper(), mapping.get(text, text))


def add_generator_review(part: dict, severity: str, category: str, issue: str,
                         assumption: str, impact: str, action: str) -> None:
    items = part.setdefault("_generator_review_items", [])
    if any(item.get("category") == category and item.get("issue") == issue for item in items):
        return
    items.append({
        "severity": severity,
        "category": category,
        "issue": issue,
        "ai_assumption": assumption,
        "cost_impact": impact,
        "manual_action": action,
    })


def normalize_review_items(part: dict) -> list[dict]:
    aliases = {"high": "high", "高": "high", "medium": "medium", "中": "medium", "low": "low", "低": "low"}
    raw_items = list(part.get("review_items") or []) + list(part.get("_generator_review_items") or [])
    items = []
    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        item = {
            "severity": aliases.get(str(raw.get("severity") or "").strip().lower(), "medium"),
            "category": str(raw.get("category") or "其他异常").strip(),
            "issue": str(raw.get("issue") or "").strip(),
            "ai_assumption": str(raw.get("ai_assumption") or "").strip(),
            "cost_impact": str(raw.get("cost_impact") or "").strip(),
            "manual_action": str(raw.get("manual_action") or "").strip(),
        }
        if item["issue"] or item["manual_action"]:
            items.append(item)
    fields = [str(field).strip() for field in part.get("needs_review_fields") or [] if str(field).strip()]
    if not items and fields:
        joined = "、".join(fields)
        items.append({
            "severity": "medium", "category": "信息不确定",
            "issue": f"以下字段无法从图纸中可靠确认：{joined}",
            "ai_assumption": "已按当前可见信息进行临时报价估算。",
            "cost_impact": "相关材料用量、加工工时或报价金额可能需要调整。",
            "manual_action": f"请人工核对原图并确认：{joined}",
        })
    return items


def format_review_comment(part: dict) -> str:
    lines = ["⚠ AI人工复核提醒"]
    for index, item in enumerate(normalize_review_items(part), 1):
        severity = SEVERITY_LABELS.get(item["severity"], "中")
        lines.extend([
            f"{index}. [{severity}][{item['category']}] {item['issue']}",
            f"AI假设：{item['ai_assumption'] or '未采用额外假设'}",
            f"可能影响：{item['cost_impact'] or '需由工程人员评估'}",
            f"请确认：{item['manual_action'] or '请核对原始图纸与工艺要求'}",
        ])
    return "\n".join(lines) if len(lines) > 1 else ""


def classify_part(part: dict) -> str:
    explicit = lookup_part_value(part, "part_type", "part_category", "category", "零件类型", "零件类别")
    if explicit not in (None, ""):
        normalized = normalize_key(explicit)
        category = CATEGORY_ALIASES.get(str(explicit).strip()) or CATEGORY_ALIASES.get(normalized)
        if category:
            return category
        add_generator_review(
            part, "high", "零件分类", f"无法识别零件类型“{explicit}”。",
            "暂按加工件计算。", "使用错误类别会显著改变材料和工序费用。",
            "请在JSON中将 part_type 明确设为加工件、钣金件、焊接件或型材组装件。",
        )
        return "加工件"

    hints = {
        "钣金件": ("unfolded_area_m2", "developed_area_m2", "cut_length_m", "bend_count", "展开面积m²"),
        "焊接件": ("bom_net_weight_kg", "weld_length_m", "weld_feature_mm", "BOM净重kg/套"),
        "型材组装件": ("net_length_m_per_set", "weight_kg_per_m", "standard_bar_length_m", "净长度m/套"),
    }
    for category, keys in hints.items():
        if lookup_part_value(part, *keys) not in (None, ""):
            add_generator_review(
                part, "medium", "零件分类", f"未提供 part_type，已根据专用字段推断为{category}。",
                f"暂按{category}工作表计算。", "若分类不正确，材料与工序计算方式会变化。",
                "请人工确认零件类别，并在后续识别数据中明确填写 part_type。",
            )
            return category

    add_generator_review(
        part, "high", "零件分类", "未提供可确认的零件类型。",
        "为兼容旧数据，暂按加工件计算。", "错误分类可能造成较大的报价偏差并降低CPK。",
        "请确认属于加工件、钣金件、焊接件或型材组装件，并补充 part_type。",
    )
    return "加工件"


def row_review_status(part: dict) -> str:
    severities = {item["severity"] for item in normalize_review_items(part)}
    if "high" in severities:
        return "高风险"
    if severities:
        return "待复核"
    return "无需复核"


def row_snapshot(ws, row: int) -> dict:
    cells = []
    for column in range(1, ws.max_column + 1):
        cell = ws.cell(row, column)
        cells.append({
            "value": cell.value,
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "font": copy(cell.font),
            "protection": copy(cell.protection),
        })
    return {"origin_row": row, "height": ws.row_dimensions[row].height, "cells": cells}


def restore_blank_formula_row(ws, snapshot: dict, target_row: int) -> None:
    origin_row = snapshot["origin_row"]
    ws.row_dimensions[target_row].height = snapshot["height"]
    for column, meta in enumerate(snapshot["cells"], 1):
        cell = ws.cell(target_row, column)
        cell._style = copy(meta["style"])
        cell.number_format = meta["number_format"]
        cell.alignment = copy(meta["alignment"])
        cell.border = copy(meta["border"])
        cell.fill = copy(meta["fill"])
        cell.font = copy(meta["font"])
        cell.protection = copy(meta["protection"])
        cell.comment = None
        value = meta["value"]
        if isinstance(value, str) and value.startswith("="):
            origin = f"{get_column_letter(column)}{origin_row}"
            cell.value = Translator(value, origin=origin).translate_formula(cell.coordinate)
        else:
            cell.value = None


def configure_defined_names(wb) -> None:
    if MATERIAL_LIST_NAME in wb.defined_names:
        del wb.defined_names[MATERIAL_LIST_NAME]
    wb.defined_names.add(DefinedName(MATERIAL_LIST_NAME, attr_text="'基础参数'!$A$6:$A$22"))


def add_validation(ws, cell_range: str, validation_type: str, formula1: str, formula2: str | None = None) -> None:
    validation = DataValidation(type=validation_type, formula1=formula1, formula2=formula2, allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def configure_sheet_validations(ws, category: str, last_row: int) -> None:
    ws.data_validations.dataValidation = []
    add_validation(ws, f"D{DATA_START_ROW}:D{last_row}", "list", f"={MATERIAL_LIST_NAME}")
    quantity_columns = {"加工件": "AD", "钣金件": "AD", "焊接件": "AG", "型材组装件": "AH"}
    status_columns = {"加工件": "AH", "钣金件": "AH", "焊接件": "AK", "型材组装件": "AL"}
    quantity_col = quantity_columns[category]
    status_col = status_columns[category]
    add_validation(ws, f"{quantity_col}{DATA_START_ROW}:{quantity_col}{last_row}", "whole", "1")
    add_validation(ws, f"{status_col}{DATA_START_ROW}:{status_col}{last_row}", "list", '"无需复核,待复核,高风险"')
    if category == "钣金件":
        add_validation(ws, f"I{DATA_START_ROW}:I{last_row}", "decimal", "0.1", "1")
    elif category == "焊接件":
        for column in ("H", "T", "X"):
            add_validation(ws, f"{column}{DATA_START_ROW}:{column}{last_row}", "decimal", "0.05", "1")
    elif category == "型材组装件":
        add_validation(ws, f"J{DATA_START_ROW}:J{last_row}", "decimal", "0.1", "1")


def prepare_category_sheet(ws, category: str, part_count: int) -> tuple[dict[str, int], int]:
    if not ws.tables:
        raise ValueError(f"工作表 {category} 缺少Excel表格")
    table = ws.tables[next(iter(ws.tables))]
    min_col, header_row, max_col, template_last_row = range_boundaries(table.ref)
    if header_row != DATA_START_ROW - 1:
        raise ValueError(f"工作表 {category} 表头不在第6行: {table.ref}")
    snapshot = row_snapshot(ws, DATA_START_ROW)
    template_data_rows = max(1, template_last_row - header_row)
    target_data_rows = max(template_data_rows, part_count)
    target_last_row = header_row + target_data_rows
    if ws.max_row < target_last_row:
        ws.insert_rows(ws.max_row + 1, amount=target_last_row - ws.max_row)
    for row in range(DATA_START_ROW, target_last_row + 1):
        restore_blank_formula_row(ws, snapshot, row)
    table.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{target_last_row}"
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref
    headers = {str(ws.cell(header_row, col).value): col for col in range(min_col, max_col + 1)}
    configure_sheet_validations(ws, category, target_last_row)
    return headers, target_last_row


def require_or_default(part: dict, field_label: str, aliases: tuple[str, ...], default=None,
                       severity: str = "medium", required: bool = False):
    value = lookup_part_value(part, *aliases)
    if value not in (None, ""):
        return value
    if required:
        add_generator_review(
            part, severity, "计算输入缺失", f"{part['_category']}缺少“{field_label}”。",
            f"暂按{default if default is not None else '空值/零值'}计算。",
            "该输入会直接影响材料费或工序费。", f"请依据图纸或工艺资料确认“{field_label}”。",
        )
    return default


def write_header_value(ws, headers: dict[str, int], row: int, header: str, value) -> None:
    column = headers.get(header)
    if column is None:
        raise ValueError(f"工作表 {ws.title} 缺少列: {header}")
    ws.cell(row, column).value = value


def fill_common_fields(ws, headers: dict[str, int], part: dict, row: int) -> None:
    write_header_value(ws, headers, row, "专案代码", part.get("project_code") or part.get("part_no") or "")
    write_header_value(ws, headers, row, "产品图号", part.get("drawing_no") or part.get("part_no") or "")
    write_header_value(ws, headers, row, "材质", normalize_material(part.get("material")) or "需人工确认")
    quantity = max(1, int(as_num(part.get("quantity"), 1) or 1))
    write_header_value(ws, headers, row, "数量", quantity)
    vendor_quote = lookup_part_value(part, "vendor_quote", "vendor_quote_untaxed", "厂商报价(未税)")
    if vendor_quote not in (None, ""):
        write_header_value(ws, headers, row, "厂商报价(未税)", as_num(vendor_quote, vendor_quote))


def fill_machining(ws, headers: dict[str, int], part: dict, row: int) -> None:
    stock_l, stock_w, stock_h = calc_stock_dimensions(part)
    for label, value in (("备料长mm", stock_l), ("备料宽mm", stock_w), ("备料高mm", stock_h)):
        if value is None:
            add_generator_review(
                part, "high", "备料尺寸缺失", f"加工件缺少“{label}”。", "对应材料尺寸暂留空。",
                "材料重量与材料费可能被低估。", f"请确认{label}后重新计算。",
            )
        write_header_value(ws, headers, row, label, value)
    write_header_value(ws, headers, row, "单位", lookup_part_value(part, "unit", "单位", default="KG"))
    process_headers = set(MACHINING_PROCESS_HEADERS.values())
    for header in process_headers:
        write_header_value(ws, headers, row, header, 0)
    process_values = part.get("process_hours") or lookup_part_value(part, "process_inputs", default={}) or {}
    if isinstance(process_values, dict):
        for name, value in process_values.items():
            key = str(name).strip()
            header = MACHINING_PROCESS_HEADERS.get(key) or MACHINING_PROCESS_HEADERS.get(normalize_key(key))
            if header:
                write_header_value(ws, headers, row, header, as_num(value, value))


def fill_sheet_metal(ws, headers: dict[str, int], part: dict, row: int) -> None:
    specs = [
        ("展开面积m²", ("unfolded_area_m2", "developed_area_m2", "展开面积m²"), None, True, "high"),
        ("板厚mm", ("sheet_thickness_mm", "thickness_mm", "板厚mm"), None, True, "high"),
        ("材料利用率", ("material_utilization", "material_yield", "材料利用率"), 0.80, True, "medium"),
        ("切割长度m", ("cut_length_m", "cutting_length_m", "切割长度m"), 0, True, "medium"),
        ("穿孔数", ("pierce_count", "穿孔数"), 0, False, "medium"),
        ("切割速度m/min", ("cutting_speed_m_per_min", "cut_speed_m_min", "切割速度m/min"), 3.0, True, "medium"),
        ("穿孔秒/次", ("pierce_seconds", "穿孔秒/次"), 1.5, True, "medium"),
        ("切割准备h/批", ("cutting_setup_hours_per_batch", "切割准备h/批"), 0.25, True, "medium"),
        ("折弯数", ("bend_count", "折弯数"), 0, False, "medium"),
        ("单刀分钟", ("minutes_per_bend", "单刀分钟"), 0.35, True, "medium"),
        ("折弯准备h/批", ("bend_setup_hours_per_batch", "折弯准备h/批"), 0.25, True, "medium"),
        ("五金/件", ("hardware_cost_each", "五金/件"), 0, False, "low"),
        ("手工分钟/件", ("manual_minutes_each", "手工分钟/件"), 0, False, "medium"),
        ("表处费/件", ("surface_treatment_cost_each", "表处费/件"), 0, False, "medium"),
        ("其它费/件", ("other_cost_each", "其它费/件"), 0, False, "low"),
    ]
    for label, aliases, default, required, severity in specs:
        value = require_or_default(part, label, aliases, default, severity, required)
        write_header_value(ws, headers, row, label, as_num(value, value))
    write_header_value(ws, headers, row, "单位", lookup_part_value(part, "unit", "单位", default="件"))


def fill_welding(ws, headers: dict[str, int], part: dict, row: int) -> None:
    specs = [
        ("BOM净重kg/套", ("bom_net_weight_kg", "bom_weight_kg", "BOM净重kg/套"), None, True, "high"),
        ("材料利用率", ("material_utilization", "材料利用率"), 0.95, True, "medium"),
        ("外购件/套", ("purchased_parts_cost_per_set", "外购件/套"), 0, False, "low"),
        ("备料h/批", ("prep_hours_per_batch", "备料h/批"), 0, True, "medium"),
        ("组立h/套", ("fitup_hours_per_set", "组立h/套"), 0, True, "medium"),
        ("焊缝长度m/套", ("weld_length_m", "weld_length_m_per_set", "焊缝长度m/套"), 0, True, "high"),
        ("焊脚/特征mm", ("weld_feature_mm", "weld_leg_mm", "焊脚/特征mm"), 0, True, "high"),
        ("截面积系数", ("cross_section_factor", "截面积系数"), 0.5, True, "medium"),
        ("焊速m/min", ("weld_speed_m_per_min", "焊速m/min"), 0.25, True, "medium"),
        ("电弧作业率", ("arc_duty_cycle", "电弧作业率"), 0.35, True, "medium"),
        ("焊材单价", ("filler_material_unit_price", "焊材单价"), 14, True, "medium"),
        ("熔敷效率", ("deposition_efficiency", "熔敷效率"), 0.85, True, "medium"),
        ("气体费率元/h", ("gas_rate_per_hour", "气体费率元/h"), 8, True, "medium"),
        ("打磨校形h/套", ("grinding_straightening_hours", "打磨校形h/套"), 0, False, "medium"),
        ("表处/外协/套", ("surface_outsource_cost_per_set", "表处/外协/套"), 0, False, "medium"),
        ("检验包装/套", ("inspection_packaging_cost_per_set", "检验包装/套"), 0, False, "low"),
    ]
    for label, aliases, default, required, severity in specs:
        value = require_or_default(part, label, aliases, default, severity, required)
        write_header_value(ws, headers, row, label, as_num(value, value))


def fill_profile_assembly(ws, headers: dict[str, int], part: dict, row: int) -> None:
    specs = [
        ("净长度m/套", ("net_length_m_per_set", "net_length_m", "净长度m/套"), None, True, "high"),
        ("理论重量kg/m", ("weight_kg_per_m", "theoretical_weight_kg_per_m", "理论重量kg/m"), None, True, "high"),
        ("标准料长m/支", ("standard_bar_length_m", "标准料长m/支"), 6, True, "medium"),
        ("排料利用率", ("nesting_utilization", "material_utilization", "排料利用率"), 0.88, True, "medium"),
        ("切断数/套", ("cuts_per_set", "切断数/套"), 0, True, "medium"),
        ("单刀分钟", ("minutes_per_cut", "单刀分钟"), 0.45, True, "medium"),
        ("锯切准备h/批", ("saw_setup_hours_per_batch", "锯切准备h/批"), 0.25, True, "medium"),
        ("钻孔数/套", ("holes_per_set", "drill_count_per_set", "钻孔数/套"), 0, False, "medium"),
        ("钻孔单孔分钟", ("minutes_per_drilled_hole", "钻孔单孔分钟"), 0.35, True, "medium"),
        ("攻牙数/套", ("taps_per_set", "tap_count_per_set", "攻牙数/套"), 0, False, "medium"),
        ("攻牙单孔分钟", ("minutes_per_tapped_hole", "攻牙单孔分钟"), 0.4, True, "medium"),
        ("连接件/套", ("connector_cost_per_set", "连接件/套"), 0, False, "low"),
        ("紧固件/套", ("fastener_cost_per_set", "紧固件/套"), 0, False, "low"),
        ("装配分钟/套", ("assembly_minutes_per_set", "装配分钟/套"), 0, True, "medium"),
        ("补充焊接分钟/套", ("supplemental_weld_minutes_per_set", "补充焊接分钟/套"), 0, False, "medium"),
        ("表处/外协/套", ("surface_outsource_cost_per_set", "表处/外协/套"), 0, False, "medium"),
        ("检验包装/套", ("inspection_packaging_cost_per_set", "检验包装/套"), 0, False, "low"),
    ]
    for label, aliases, default, required, severity in specs:
        value = require_or_default(part, label, aliases, default, severity, required)
        write_header_value(ws, headers, row, label, as_num(value, value))


def finalize_part_row(ws, headers: dict[str, int], part: dict, row: int) -> None:
    status = row_review_status(part)
    write_header_value(ws, headers, row, "复核状态", status)
    notes = []
    if part.get("notes"):
        notes.append(str(part["notes"]))
    if part.get("_generator_review_items"):
        notes.append("AI分类/默认参数需确认；详见图纸附件")
    write_header_value(ws, headers, row, "备注", "；".join(notes))
    review_comment = format_review_comment(part)
    if review_comment:
        write_header_value(ws, headers, row, "产品图号", part.get("drawing_no") or part.get("part_no") or "")
        ws.cell(row, headers["产品图号"]).comment = Comment(review_comment, "LFAF AI")


def fill_category_row(ws, headers: dict[str, int], part: dict, row: int) -> None:
    fill_common_fields(ws, headers, part, row)
    category = part["_category"]
    if category == "加工件":
        fill_machining(ws, headers, part, row)
    elif category == "钣金件":
        fill_sheet_metal(ws, headers, part, row)
    elif category == "焊接件":
        fill_welding(ws, headers, part, row)
    elif category == "型材组装件":
        fill_profile_assembly(ws, headers, part, row)
    finalize_part_row(ws, headers, part, row)


def build_attachment_sheet(wb, parts: list[dict]):
    if ATTACH_SHEET in wb.sheetnames:
        del wb[ATTACH_SHEET]
    ws = wb.create_sheet(ATTACH_SHEET)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "ED7D31"

    ws.merge_cells("A1:K1")
    ws["A1"] = "图纸附件与 AI 人工复核提醒"
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        "每张图纸仅占一行；零件类别决定对应计算工作表。多条异常、AI假设、可能影响和"
        "人工确认事项按编号合并。正式报价或生产前请完成人工复核。"
    )
    ws["A2"].font = Font(name="Microsoft YaHei", size=10, color="7F6000")
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34

    headers = [
        "序号", "零件类别", "产品图号", "最高等级", "异常/特殊情况", "AI计算采用假设",
        "可能影响", "人工确认事项", "来源文件", "预览图", "识别/估算备注",
    ]
    for column, value in enumerate(headers, 1):
        cell = ws.cell(3, column, value)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = FULL_BORDER
    ws.row_dimensions[3].height = 30

    widths = [8, 14, 20, 10, 38, 38, 34, 40, 48, 45, 50]
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width

    severity_rank = {"high": 3, "medium": 2, "low": 1}
    row = 4
    for index, part in enumerate(parts, 1):
        review_items = normalize_review_items(part)
        highest = max(review_items, key=lambda item: severity_rank.get(item["severity"], 2), default=None)
        severity = highest["severity"] if highest else "low"

        def numbered(field: str, fallback: str) -> str:
            if not review_items:
                return fallback
            return "\n\n".join(f"{number}. {item.get(field) or fallback}" for number, item in enumerate(review_items, 1))

        raw_dims = (part.get("length"), part.get("width"), part.get("height"))
        notes = []
        if part["_category"] == "加工件" and any(value not in (None, "") for value in raw_dims):
            stock_dims = calc_stock_dimensions(part)
            notes.append(
                f"成品尺寸: {raw_dims[0]} x {raw_dims[1]} x {raw_dims[2]}；"
                f"成本按备料尺寸: {stock_dims[0]} x {stock_dims[1]} x {stock_dims[2]}"
            )
        if part.get("notes"):
            notes.append(str(part["notes"]))

        values = [
            index, part["_category"], part.get("drawing_no") or part.get("part_no") or "",
            SEVERITY_LABELS.get(severity, "低") if review_items else "无",
            numbered("issue", "无特殊复核事项"), numbered("ai_assumption", "未采用额外假设"),
            numbered("cost_impact", "无已识别的特殊影响"), numbered("manual_action", "按公司流程复核"),
            part.get("source_file") or "", "", "\n".join(notes),
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row, column, value)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = FULL_BORDER
        ws.cell(row, 4).fill = SEVERITY_FILLS.get(severity, SEVERITY_FILLS["low"])
        ws.cell(row, 4).font = Font(name="Microsoft YaHei", size=10, bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="center")

        preview = part.get("preview_image") or part.get("source_file")
        if preview and str(preview).lower().endswith((".png", ".jpg", ".jpeg")) and os.path.exists(preview):
            try:
                image = XLImage(preview)
                ratio = min(260 / image.width, 180 / image.height, 1)
                image.width = int(image.width * ratio)
                image.height = int(image.height * ratio)
                ws.add_image(image, f"J{row}")
                ws.row_dimensions[row].height = max(150, image.height * 0.75)
            except Exception as exc:
                ws.cell(row, 10).value = f"图片嵌入失败：{exc}"
        else:
            ws.cell(row, 10).value = "非图片格式，请查看来源文件"
        row += 1

    if parts:
        ws.auto_filter.ref = f"A3:K{row - 1}"
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def infer_output_directory(data: dict, input_json_path: str) -> Path:
    input_path = Path(input_json_path).expanduser().resolve()
    configured = data.get("output_dir")
    if configured:
        configured_path = Path(str(configured)).expanduser()
        if not configured_path.is_absolute():
            configured_path = input_path.parent / configured_path
        return configured_path.resolve()
    for part in data.get("parts") or []:
        source = part.get("source_file")
        if not source:
            continue
        source_text = str(source).strip()
        if source_text.lower().startswith(("http://", "https://")):
            continue
        source_path = Path(source_text).expanduser()
        if not source_path.is_absolute():
            source_path = input_path.parent / source_path
        return source_path.resolve().parent
    return input_path.parent


def next_available_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    for version in range(2, 10000):
        candidate = path.with_name(f"{path.stem}_v{version:03d}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"unable to allocate a versioned output filename for {path}")


def normalize_output_path(output_arg: str | None, quote_date: str | None, data: dict, input_json_path: str) -> Path:
    from datetime import date
    qdate = (quote_date or str(date.today())).replace("/", "-")
    ymd = "".join(ch for ch in qdate if ch.isdigit())[:8] or date.today().strftime("%Y%m%d")
    filename = f"LFAF_四类零件预估报价明细_{ymd}.xlsx"
    if not output_arg:
        return next_available_output_path(infer_output_directory(data, input_json_path) / filename)
    out = Path(output_arg).expanduser()
    if output_arg.endswith(os.sep) or (out.exists() and out.is_dir()) or out.suffix.lower() != ".xlsx":
        out = out / filename
    return next_available_output_path(out.resolve())


def update_summary_limit(wb, category_last_rows: dict[str, int]) -> None:
    ws = wb[SUMMARY_SHEET]
    row_map = {"加工件": 7, "钣金件": 8, "焊接件": 9, "型材组装件": 10}
    for category, summary_row in row_map.items():
        last_row = max(506, category_last_rows[category])
        # C:K includes the original totals plus the vendor-comparison summary.
        for column in range(3, 12):
            cell = ws.cell(summary_row, column)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = cell.value.replace("$506", f"${last_row}")


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-json", required=True, help="Path to extracted part JSON")
    parser.add_argument("--output", required=False, help="Output xlsx path or directory")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Template xlsx path")
    args = parser.parse_args()

    with open(args.input_json, "r", encoding="utf-8") as source:
        data = json.load(source)
    raw_parts = data.get("parts") or []
    if not raw_parts:
        raise SystemExit("input JSON must contain at least one part in parts[]")

    parts = []
    for raw in raw_parts:
        part = dict(raw)
        part["_category"] = classify_part(part)
        parts.append(part)

    output = normalize_output_path(args.output, data.get("quote_date"), data, args.input_json)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(args.template)
    missing_sheets = [name for name in (*CATEGORY_SHEETS, BASE_PARAMETER_SHEET, SUMMARY_SHEET) if name not in wb.sheetnames]
    if missing_sheets:
        raise SystemExit(f"template missing sheets: {', '.join(missing_sheets)}")

    configure_defined_names(wb)
    grouped = {category: [] for category in CATEGORY_SHEETS}
    for part in parts:
        grouped[part["_category"]].append(part)

    category_last_rows = {}
    for category in CATEGORY_SHEETS:
        ws = wb[category]
        headers, last_row = prepare_category_sheet(ws, category, len(grouped[category]))
        category_last_rows[category] = last_row
        for offset, part in enumerate(grouped[category]):
            fill_category_row(ws, headers, part, DATA_START_ROW + offset)

    update_summary_limit(wb, category_last_rows)
    build_attachment_sheet(wb, parts)
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output)
    print(f"FINAL: {output}")
    print(output)


if __name__ == "__main__":
    main()
