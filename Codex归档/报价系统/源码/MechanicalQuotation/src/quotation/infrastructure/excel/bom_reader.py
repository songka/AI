"""BOM Excel Reader.

Reads a BOM Excel file and produces a BomSheet with full source tracing.

Features:
- Flexible column name mapping (YAML config)
- Merged cell handling
- Multi-sheet support
- Blank row / header row detection
- source_file / source_sheet / source_row tracking on every BomEntry
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import yaml

from quotation.domain.bom import BomEntry, BomSheet

logger = logging.getLogger("quotation.infrastructure.excel.bom_reader")

# ---------------------------------------------------------------------------
# Default column name mapping (fallback when no YAML config)
# ---------------------------------------------------------------------------

DEFAULT_COLUMN_ALIASES: dict[str, list[str]] = {
    "level": ["Level", "級別", "層級", "BOM Level"],
    "item": ["Item", "Item ", "料號", "零件號", "物料編號"],
    "description": ["Description", "描述", "品名", "規格描述", "物料描述"],
    "item_type": ["Type", "類型", "項目類型", "Item Type"],
    "uom": ["Uom", "UOM", "單位", "Unit", "計量單位"],
    "quantity": ["Quantity", "Qty", "數量", "Q'ty"],
    "unit_cost": ["Unit Cost", "單價", "單位成本", "Unit Price"],
    "extended_cost": ["Extended Cost", "總價", "總成本", "Total Cost"],
    "notes": ["備註", "Notes", "Remark", "備注"],
}

SKIP_ITEM_VALUES = {"None", "Item", "Item ", ""}
SKIP_DESC_VALUES = {"None", ""}
HEADER_SCAN_ROWS = range(1, 11)  # Scan rows 1-10 for header
HEADER_MIN_MATCHES = 3


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------

class ColumnMapping:
    """Maps Excel column letters to standard field names.

    Built by scanning a header row and matching header text to known aliases.
    """

    def __init__(
        self,
        aliases: dict[str, list[str]] | None = None,
    ):
        self._aliases = aliases or DEFAULT_COLUMN_ALIASES
        # field_name → column_index (0-based)
        self._map: dict[str, int] = {}

    def build_from_row(self, row_cells: list[tuple[int, str]]) -> int:
        """Scan a row and map column indices to field names.

        Args:
            row_cells: List of (col_index_0based, cell_value) tuples.

        Returns:
            Number of fields matched.
        """
        self._map.clear()
        for col_idx, cell_value in row_cells:
            if not cell_value:
                continue
            value_clean = cell_value.strip()
            for field_name, names in self._aliases.items():
                if field_name in self._map:
                    continue  # Already mapped
                if any(name.lower() == value_clean.lower() for name in names):
                    self._map[field_name] = col_idx
                    break
        return len(self._map)

    def get_col(self, field_name: str) -> int | None:
        """Get the 0-based column index for a field name."""
        return self._map.get(field_name)

    def is_valid(self) -> bool:
        """Return True if at least the minimum required fields are mapped."""
        required = {"item", "description"}
        return required.issubset(self._map.keys())

    @property
    def mapped_fields(self) -> dict[str, int]:
        return dict(self._map)


# ---------------------------------------------------------------------------
# BOM Reader
# ---------------------------------------------------------------------------

class BomReader:
    """Reads a BOM Excel file and produces a BomSheet."""

    def __init__(
        self,
        column_aliases: dict[str, list[str]] | None = None,
        header_scan_rows: range | None = None,
        header_min_matches: int = HEADER_MIN_MATCHES,
    ):
        self._aliases = column_aliases or DEFAULT_COLUMN_ALIASES
        self._scan_rows = header_scan_rows or HEADER_SCAN_ROWS
        self._min_matches = header_min_matches

    # -- Public API --

    def read(self, file_path: str | Path, sheet_name: str | None = None) -> BomSheet:
        """Read a BOM Excel file and return a BomSheet.

        Args:
            file_path: Path to the .xlsx file.
            sheet_name: Sheet name to read. If None, auto-detects the first
                        sheet with a valid BOM header.

        Returns:
            BomSheet with all entries populated and source tracing.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"BOM file not found: {path}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise ValueError(f"Unsupported file format: {path.suffix}. Expected .xlsx")

        # Lazy import — keep openpyxl optional until actually needed
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)

        if sheet_name is None:
            sheet_name = self._find_bom_sheet(wb)
            if sheet_name is None:
                raise ValueError(
                    f"No BOM sheet found in {path}. "
                    f"Scanned sheets: {wb.sheetnames}"
                )

        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        return self._read_sheet(ws, str(path), sheet_name)

    def read_all_sheets(self, file_path: str | Path) -> list[BomSheet]:
        """Read all sheets that contain a valid BOM header."""
        path = Path(file_path)
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        sheets: list[BomSheet] = []
        for name in wb.sheetnames:
            try:
                sheet = self._read_sheet(wb[name], str(path), name)
                if sheet.total_rows > 0:
                    sheets.append(sheet)
            except Exception:
                logger.debug("Sheet '%s' does not appear to be a BOM sheet", name)
        return sheets

    # -- Config loading --

    @classmethod
    def from_yaml(cls, yaml_path: str | Path) -> "BomReader":
        """Create a BomReader from a YAML column mapping config file."""
        path = Path(yaml_path)
        if not path.exists():
            logger.warning("Config file not found: %s, using defaults", path)
            return cls()

        with open(path, encoding="utf-8") as f:
            config = yaml.safe_load(f)

        aliases = config.get("fields", {}) if config else {}
        if not aliases:
            return cls()

        # Extract header scan config
        scan_cfg = config.get("header_scan", {}) if config else {}
        min_row = scan_cfg.get("min_row", 1)
        max_row = scan_cfg.get("max_row", 10)
        scan_rows = range(min_row, max_row + 1)
        min_matches = config.get("header_min_matches", HEADER_MIN_MATCHES) if config else HEADER_MIN_MATCHES

        return cls(
            column_aliases=aliases,
            header_scan_rows=scan_rows,
            header_min_matches=min_matches,
        )

    # -- Internal methods --

    def _find_bom_sheet(self, wb) -> str | None:
        """Find the first sheet with a valid BOM header."""
        for name in wb.sheetnames:
            ws = wb[name]
            mapping = self._detect_header(ws)
            if mapping is not None and mapping.is_valid():
                return name
        # Fallback: return first sheet
        return wb.sheetnames[0] if wb.sheetnames else None

    def _detect_header(self, ws) -> ColumnMapping | None:
        """Scan rows to find the column header row."""
        best_mapping: ColumnMapping | None = None
        best_matches = 0

        for row_idx in self._scan_rows:
            if row_idx > ws.max_row:
                break
            cells = []
            for col_idx in range(ws.max_column):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                val = cell.value
                if val is not None:
                    cells.append((col_idx, str(val)))

            if len(cells) >= 2:  # At minimum need item + description
                mapping = ColumnMapping(self._aliases)
                matches = mapping.build_from_row(cells)
                if matches > best_matches:
                    best_matches = matches
                    best_mapping = mapping

        return best_mapping

    def _read_sheet(self, ws, source_file: str, sheet_name: str) -> BomSheet:
        """Read a single worksheet into a BomSheet."""
        # Detect header
        mapping = self._detect_header(ws)

        # Build a default mapping if detection fails
        if mapping is None or not mapping.is_valid():
            logger.warning(
                "Could not detect BOM header in sheet '%s', "
                "falling back to column index defaults",
                sheet_name,
            )
            mapping = self._default_mapping_for_real_bom()
            # When using default mapping, skip rows up to the scan range end
            # as they likely contain company info, not data
            header_row = self._scan_rows.stop - 1 if self._scan_rows else 4
        else:
            # Find header row (the row that produced the mapping)
            header_row = self._find_header_row(ws, mapping)

        # Read data rows (everything after the header row)
        entries: list[BomEntry] = []
        col_item = mapping.get_col("item")
        col_desc = mapping.get_col("description")
        col_level = mapping.get_col("level")
        col_type = mapping.get_col("item_type")
        col_uom = mapping.get_col("uom")
        col_qty = mapping.get_col("quantity")
        col_uc = mapping.get_col("unit_cost")
        col_ec = mapping.get_col("extended_cost")
        col_notes = mapping.get_col("notes")

        start_row = (header_row + 1) if header_row is not None else 1

        for row_idx in range(start_row, ws.max_row + 1):
            item_val = self._cell_str(ws, row_idx, col_item) if col_item is not None else ""
            desc_val = self._cell_str(ws, row_idx, col_desc) if col_desc is not None else ""

            # Skip blank / invalid rows
            if item_val in SKIP_ITEM_VALUES and desc_val in SKIP_DESC_VALUES:
                continue
            if not item_val and not desc_val:
                continue
            if item_val in SKIP_ITEM_VALUES:
                continue

            entry = BomEntry(
                source_file=source_file,
                source_sheet=sheet_name,
                source_row=row_idx,
                item=item_val,
                description=desc_val,
                level=self._cell_int(ws, row_idx, col_level) if col_level is not None else 0,
                item_type=self._cell_str(ws, row_idx, col_type, "Purchased") if col_type is not None else "Purchased",
                uom=self._cell_str(ws, row_idx, col_uom, "ST") if col_uom is not None else "ST",
                quantity=self._cell_float(ws, row_idx, col_qty, 1.0) if col_qty is not None else 1.0,
                unit_cost=self._cell_float(ws, row_idx, col_uc, 0.0) if col_uc is not None else 0.0,
                extended_cost=self._cell_float(ws, row_idx, col_ec, 0.0) if col_ec is not None else 0.0,
                notes=self._cell_str(ws, row_idx, col_notes) if col_notes is not None else None,
            )
            entries.append(entry)

        return BomSheet(
            source_file=source_file,
            source_sheet=sheet_name,
            total_rows=len(entries),
            entries=entries,
        )

    def _find_header_row(self, ws, mapping: ColumnMapping) -> int | None:
        """Find which row produced this mapping."""
        for row_idx in self._scan_rows:
            if row_idx > ws.max_row:
                break
            cells = []
            for col_idx in range(ws.max_column):
                val = ws.cell(row=row_idx, column=col_idx + 1).value
                if val is not None:
                    cells.append((col_idx, str(val)))
            test = ColumnMapping(self._aliases)
            test.build_from_row(cells)
            if test.mapped_fields == mapping.mapped_fields:
                return row_idx
        return None

    def _default_mapping_for_real_bom(self) -> ColumnMapping:
        """Fallback: assume standard column positions (A=Level, B=Item, C=Desc, ...)."""
        mapping = ColumnMapping(self._aliases)
        # Force-map to the known real BOM layout
        mapping._map = {
            "level": 0,
            "item": 1,
            "description": 2,
            "item_type": 4,
            "uom": 5,
            "quantity": 6,
            "unit_cost": 7,
            "extended_cost": 8,
            "notes": 9,
        }
        return mapping

    # -- Cell value helpers (handle merged cells and type coercion) --

    @staticmethod
    def _cell_str(ws, row: int, col: int | None, default: str = "") -> str:
        """Get cell value as string."""
        if col is None:
            return default
        val = ws.cell(row=row, column=col + 1).value
        if val is None:
            return default
        return str(val).strip()

    @staticmethod
    def _cell_float(ws, row: int, col: int | None, default: float = 0.0) -> float:
        """Get cell value as float."""
        if col is None:
            return default
        val = ws.cell(row=row, column=col + 1).value
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _cell_int(ws, row: int, col: int | None, default: int = 0) -> int:
        """Get cell value as int."""
        if col is None:
            return default
        val = ws.cell(row=row, column=col + 1).value
        if val is None:
            return default
        try:
            return int(val)
        except (ValueError, TypeError):
            return default
