"""Permission-aware supplier master and append-only price maintenance."""

from __future__ import annotations

import re
import uuid
from datetime import datetime, timezone
from typing import Any

import openpyxl

from quotation.application.auth_service import AuthService
from quotation.domain.supplier import Supplier, SupplierStatus
from quotation.domain.supplier_price import (
    PriceStatus,
    SupplierPriceRecord,
    TargetType,
)
from quotation.domain.user import User
from quotation.infrastructure.supplier.repository import (
    SupplierPriceRepository,
    SupplierRepository,
)


class SupplierMaintenanceService:
    EDITABLE_FIELDS = {
        "supplier_name",
        "contact_person",
        "phone",
        "email",
        "address",
        "payment_terms",
        "currency",
        "default_tax_included",
        "lead_time_days",
        "quality_rating",
        "notes",
    }

    def __init__(
        self,
        suppliers: SupplierRepository,
        prices: SupplierPriceRepository,
        auth: AuthService,
    ) -> None:
        self.suppliers = suppliers
        self.prices = prices
        self.auth = auth

    def create_supplier(
        self,
        actor: User,
        *,
        supplier_id: str,
        supplier_name: str,
        **fields: Any,
    ) -> Supplier:
        self.auth.require_permission(actor, "price.modify")
        supplier_id = supplier_id.strip().upper()
        if not re.fullmatch(r"SUP-[A-Z0-9-]{3,40}", supplier_id):
            raise ValueError("供应商编号必须以 SUP- 开头，只能包含字母、数字和横线")
        if not supplier_name.strip():
            raise ValueError("供应商名称不能为空")
        now = self._now()
        supplier = Supplier(
            supplier_id=supplier_id,
            supplier_name=supplier_name.strip(),
            created_at=now,
            updated_at=now,
            **{key: value for key, value in fields.items() if key in self.EDITABLE_FIELDS},
        )
        return self.suppliers.create(supplier)

    def get_supplier(self, actor: User, supplier_id: str) -> Supplier:
        self.auth.require_permission(actor, "price.view_cost")
        supplier = self.suppliers.get(supplier_id)
        if supplier is None:
            raise KeyError(supplier_id)
        return supplier

    def list_suppliers(
        self, actor: User, *, query: str | None = None
    ) -> list[Supplier]:
        self.auth.require_permission(actor, "price.view_cost")
        suppliers = self.suppliers.list()
        if query:
            needle = query.casefold()
            suppliers = [
                item
                for item in suppliers
                if needle in item.supplier_id.casefold()
                or needle in item.supplier_name.casefold()
                or needle in str(item.contact_person or "").casefold()
            ]
        return sorted(suppliers, key=lambda item: item.supplier_name)

    def update_supplier(
        self, actor: User, supplier_id: str, changes: dict[str, Any]
    ) -> Supplier:
        self.auth.require_permission(actor, "price.modify")
        supplier = self.get_supplier(actor, supplier_id)
        unknown = set(changes) - self.EDITABLE_FIELDS
        if unknown:
            raise ValueError(f"不允许修改字段：{', '.join(sorted(unknown))}")
        updated = supplier.model_copy(update={**changes, "updated_at": self._now()})
        if not updated.supplier_name.strip():
            raise ValueError("供应商名称不能为空")
        return self.suppliers.update(updated)

    def set_supplier_status(
        self, actor: User, supplier_id: str, status: SupplierStatus
    ) -> Supplier:
        self.auth.require_permission(actor, "price.modify")
        supplier = self.get_supplier(actor, supplier_id)
        return self.suppliers.update(
            supplier.model_copy(update={"status": status, "updated_at": self._now()})
        )

    def delete_supplier(self, actor: User, supplier_id: str) -> None:
        self.auth.require_permission(actor, "price.modify")
        self.get_supplier(actor, supplier_id)
        if self.prices.has_records(supplier_id):
            raise ValueError("供应商存在历史报价，只能停用，不能删除")
        self.suppliers.delete(supplier_id)

    def create_price_record(
        self,
        actor: User,
        *,
        supplier_id: str,
        target_type: TargetType,
        unit_price: float | None,
        unit: str,
        material_code: str | None = None,
        material_spec: str | None = None,
        process_code: str | None = None,
        surface_code: str | None = None,
        effective_from: str | None = None,
        effective_to: str | None = None,
        quote_number: str | None = None,
        source_file: str | None = None,
        source_sheet: str | None = None,
        source_cell: str | None = None,
        currency: str = "CNY",
        tax_included: bool = False,
        tax_rate: float | None = None,
    ) -> SupplierPriceRecord:
        self.auth.require_permission(actor, "price.modify")
        supplier = self.get_supplier(actor, supplier_id)
        if supplier.status != SupplierStatus.ACTIVE:
            raise ValueError("停用或黑名单供应商不能新增报价")
        if unit_price == 0:
            raise ValueError("未知价格不能填写为 0；请留空并标记为未知价格")
        if not unit.strip():
            raise ValueError("计价单位不能为空")
        identifiers = {
            TargetType.MATERIAL: material_code,
            TargetType.PROFILE: material_code,
            TargetType.PROCESS: process_code,
            TargetType.SURFACE: surface_code,
            TargetType.OTHER: material_code or process_code or surface_code,
        }
        if not str(identifiers[target_type] or "").strip():
            raise ValueError("价格对象代码不能为空")
        status = PriceStatus.PENDING_REVIEW if unit_price is not None else PriceStatus.UNKNOWN_PRICE
        record = SupplierPriceRecord(
            price_record_id=f"PR-{datetime.now():%Y%m%d}-{uuid.uuid4().hex[:8].upper()}",
            supplier_id=supplier.supplier_id,
            supplier_name=supplier.supplier_name,
            target_type=target_type,
            material_code=material_code,
            material_spec=material_spec,
            process_code=process_code,
            surface_code=surface_code,
            unit_price=unit_price,
            unit=unit.strip(),
            currency=currency,
            tax_included=tax_included,
            tax_rate=tax_rate,
            effective_from=effective_from,
            effective_to=effective_to,
            quote_number=quote_number,
            source_file=source_file,
            source_sheet=source_sheet,
            source_cell=source_cell,
            status=status,
            created_by=actor.user_id,
            created_at=self._now(),
        )
        return self.prices.append(record)

    def list_price_records(
        self,
        actor: User,
        *,
        supplier_id: str | None = None,
        query: str | None = None,
    ) -> list[SupplierPriceRecord]:
        self.auth.require_permission(actor, "price.view_cost")
        return self.prices.list(supplier_id=supplier_id, query=query)

    def import_price_excel(self, actor: User, source_path: str) -> dict[str, Any]:
        """Import a fixed-column Excel sheet into append-only pending S records."""

        self.auth.require_permission(actor, "price.modify")
        workbook = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
        sheet = workbook.active
        headers = {
            str(cell.value or "").strip().casefold(): index
            for index, cell in enumerate(next(sheet.iter_rows()), start=1)
        }
        aliases = {
            "supplier_id": ("supplier_id", "供应商编号"),
            "target_type": ("target_type", "价格类型"),
            "material_code": ("material_code", "材料代码"),
            "material_spec": ("material_spec", "规格"),
            "process_code": ("process_code", "工序代码"),
            "surface_code": ("surface_code", "表面处理代码"),
            "unit_price": ("unit_price", "未税单价"),
            "unit": ("unit", "单位"),
            "effective_from": ("effective_from", "生效日期"),
            "quote_number": ("quote_number", "供应商报价单号"),
            "tax_included": ("tax_included", "是否含税"),
            "tax_rate": ("tax_rate", "税率"),
        }
        columns = {
            key: next(
                (headers[name.casefold()] for name in names if name.casefold() in headers),
                None,
            )
            for key, names in aliases.items()
        }
        required = ("supplier_id", "target_type", "unit_price", "unit")
        missing = [key for key in required if columns[key] is None]
        if missing:
            workbook.close()
            raise ValueError(f"价格导入模板缺少字段：{', '.join(missing)}")
        imported: list[str] = []
        errors: list[dict[str, Any]] = []
        source_name = str(source_path).replace("\\", "/").rsplit("/", 1)[-1]
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue

            def value(key: str, current_row=row):
                column = columns[key]
                return (
                    current_row[column - 1]
                    if column is not None and column <= len(current_row)
                    else None
                )

            try:
                raw_price = value("unit_price")
                raw_date = value("effective_from")
                if hasattr(raw_date, "date"):
                    raw_date = raw_date.date().isoformat()
                raw_tax = value("tax_included")
                record = self.create_price_record(
                    actor,
                    supplier_id=str(value("supplier_id") or "").strip(),
                    target_type=TargetType(str(value("target_type") or "").strip().upper()),
                    material_code=str(value("material_code") or "").strip() or None,
                    material_spec=str(value("material_spec") or "").strip() or None,
                    process_code=str(value("process_code") or "").strip() or None,
                    surface_code=str(value("surface_code") or "").strip() or None,
                    unit_price=float(raw_price) if raw_price not in (None, "") else None,
                    unit=str(value("unit") or "").strip(),
                    effective_from=str(raw_date).strip() if raw_date else None,
                    quote_number=str(value("quote_number") or "").strip() or None,
                    source_file=source_name,
                    source_sheet=sheet.title,
                    source_cell=(
                        f"{sheet.cell(row_number, columns['unit_price']).coordinate}"
                        if columns["unit_price"] is not None
                        else None
                    ),
                    tax_included=str(raw_tax or "").strip().casefold()
                    in {"1", "true", "yes", "是", "含税"},
                    tax_rate=(
                        float(value("tax_rate"))
                        if value("tax_rate") not in (None, "")
                        else None
                    ),
                )
                imported.append(record.price_record_id)
            except (KeyError, TypeError, ValueError) as exc:
                errors.append({"行号": row_number, "错误": str(exc)})
        workbook.close()
        return {
            "导入成功": len(imported),
            "导入失败": len(errors),
            "报价记录编号": imported,
            "错误明细": errors,
        }

    @staticmethod
    def _now() -> str:
        return datetime.now(timezone.utc).isoformat()
