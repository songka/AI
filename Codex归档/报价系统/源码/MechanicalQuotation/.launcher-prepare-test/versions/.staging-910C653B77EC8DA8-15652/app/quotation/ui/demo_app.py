"""Mechanical Quotation System — Demo UI Application.

Launch: .venv/Scripts/python -m quotation.ui.demo_app
"""

from __future__ import annotations

import sys
import traceback
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Any

# Ensure the project root is on the path
_PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from quotation.domain.quote import PriceSource, Quote, QuoteItem
from quotation.domain.supplier import SupplierStatus
from quotation.domain.supplier_price import TargetType
from quotation.domain.user import UserRole, UserSession, UserStatus
from quotation.application.auth_service import AuthService
from quotation.application.external_skill_settings import (
    build_external_skill_settings_service,
)
from quotation.application.history_service import QuotationHistory
from quotation.application.management_service import ManagementQueryService, export_history_quote
from quotation.application.price_approval_service import PriceApprovalService
from quotation.application.supplier_maintenance_service import SupplierMaintenanceService
from quotation.infrastructure.dxf.reader import DxfReader
from quotation.infrastructure.smb.client import SmbStorageClient
from quotation.infrastructure.feature.geometric import GeometricExtractor
from quotation.infrastructure.feature.manufacturing import ManufacturingExtractor
from quotation.infrastructure.feature.quotation_mapper import QuotationMapper
from quotation.infrastructure.rules.pricing_resolver import PricingResolver
from quotation.infrastructure.rules.quote_builder import QuoteBuilder
from quotation.infrastructure.supplier.repository import SupplierPriceRepository, SupplierRepository
from quotation.infrastructure.supplier.price_review_repository import PriceReviewRepository
from quotation.ui.viewmodels import QuoteViewModel, TaxResult
from quotation.ui.widgets import (
    CONTENT_BG, FONT_FAMILY,
    BatchQuotePage, ManagementPage, NavPanel, NewQuotePage, StructuredDetailWindow,
    SystemSettingsPage, quote_detail_sections, record_detail_sections,
)
from quotation.ui.localization import STATUS_LABELS, TYPE_LABELS, UNIT_LABELS, display_value

# ---------------------------------------------------------------------------
# Demo part definitions (mirrors cli/main.py DEMO_PARTS)
# ---------------------------------------------------------------------------

DEMO_PARTS = {
    "J003": {
        "part_number": "UC1000005854",
        "part_name": "J003",
        "material": "S50C",
        "historical_price": 1425.0,
        "size": (928, 796),
        "circles": [(200, 398, 3), (350, 398, 3), (500, 398, 3), (650, 398, 3)],
        "texts": [
            ("S50C", 10, 810, 8),
            ("6-M6", 200, 400, 5),
            ("表面鍍鉻", 10, 820, 5),
        ],
    },
    "W001": {
        "part_number": "UC2020083221",
        "part_name": "W001",
        "material": "鋁型材",
        "size": (1300, 1300),
        "circles": [],
        "texts": [
            ("鋁型材 40x40", 10, 1320, 6),
            ("防護圍欄", 10, 1340, 6),
            ("門組件", 10, 1360, 5),
            ("白色透明亞克力", 10, 1380, 4),
            ("合頁", 10, 1400, 4),
            ("磁吸", 10, 1420, 4),
            ("把手", 10, 1440, 4),
            ("角碼", 10, 1460, 4),
            ("加強筋焊接", 10, 1480, 4),
        ],
    },
}


def quote_feature_display_fields(summary: dict[str, Any]) -> list[tuple[str, str]]:
    """Build the Chinese feature card, including formal and reference totals."""

    weight_names = {
        "BBOX_ESTIMATE": "按外形尺寸估算",
        "PROFILE_GEOMETRY": "按型材几何计算",
        "UNRESOLVED_WELDMENT_STRUCTURE": "焊接结构待人工确认",
        "UNKNOWN": "未知",
    }
    dimensions = str(summary.get("bounding_box", "—")).replace("x", "×")
    return [
        ("外形尺寸", dimensions),
        ("孔数量", str(summary.get("mfg_holes", 0))),
        ("螺纹数量", str(summary.get("mfg_threads", 0))),
        ("框架数量", str(summary.get("frames", 0))),
        ("装配数量", str(summary.get("assemblies", 0))),
        ("附件数量", str(summary.get("accessories", 0))),
        ("焊接数量", str(summary.get("welds", 0))),
        ("估算重量", str(summary.get("weight", "—")).replace("kg", "千克")),
        ("重量依据", weight_names.get(summary.get("weight_resolution"), "待确认")),
        ("分项未税合计", str(summary.get("itemized_subtotal", "—"))),
        ("整件模型参考价（不计入）", str(summary.get("feature_calibration_reference", "—"))),
        ("多智能体审核结论", str(summary.get("agent_review_verdict", "—"))),
        ("多智能体审核摘要", str(summary.get("agent_review_summary", "—"))),
    ]


# ---------------------------------------------------------------------------
# Pipeline runner (reuses existing infrastructure — no formula duplication)
# ---------------------------------------------------------------------------

def run_quotation_pipeline(part_name: str) -> tuple[Quote | None, dict[str, Any], str | None]:
    """Run the full 6-layer quotation pipeline for a demo part.

    Returns:
        (Quote, feature_summary_dict, error_message_or_None)
    """
    import ezdxf

    part = DEMO_PARTS.get(part_name)
    if not part:
        return None, {}, f"Unknown part: {part_name}"

    cwd = Path.cwd()

    try:
        # 1. Generate DXF
        doc = ezdxf.new()
        doc.header["$INSUNITS"] = 4
        msp = doc.modelspace()
        w, h = part["size"]
        msp.add_line((0, 0), (w, 0))
        msp.add_line((w, 0), (w, h))
        msp.add_line((w, h), (0, h))
        msp.add_line((0, h), (0, 0))
        for cx, cy, r in part["circles"]:
            msp.add_circle((cx, cy), radius=r)
        for content, x, y, height in part["texts"]:
            msp.add_text(content, height=height).set_placement((x, y))
        dxf_path = cwd / f"demo_{part_name}.dxf"
        doc.saveas(str(dxf_path))

        # 2. CAD Import -> Drawing
        reader = DxfReader()
        import_result = reader.read(dxf_path)
        drawing = import_result.drawing

        # 3. Feature Extraction
        geo_ext = GeometricExtractor()
        geo = geo_ext.extract(drawing.raw_entities)

        mfg_ext = ManufacturingExtractor()
        mfg = mfg_ext.extract(geo)

        # 4. Quotation Mapping
        mapper = QuotationMapper()
        qf = mapper.map(mfg, geo)

        # 5. Pricing
        resolver = PricingResolver()
        items: list[QuoteItem] = []
        for mq in qf.machining:
            items.extend(resolver.resolve_machining(mq))
        for fq in qf.frames:
            items.extend(resolver.resolve_frame(fq))
        for aq in qf.assemblies:
            items.extend(resolver.resolve_assembly(aq))

        # 6. Quote Builder
        builder = QuoteBuilder()
        feat_conf = mfg.material.confidence if mfg.material else None
        quote = builder.build(
            quote_id=f"Q-DEMO-{part_name}",
            drawing_id=f"DEMO-{part_name}",
            part_number=part["part_number"],
            part_name=part["part_name"],
            material=part["material"],
            items=items,
            feature_confidence=feat_conf,
            price_version=resolver.price_version,
            rule_version="1.0",
        )

        # Cleanup temp DXF
        dxf_path.unlink(missing_ok=True)

        # Feature summary
        bbox = geo.bounding_box
        weight_kg = None
        for item in quote.items:
            if item.category == "material" and item.evidence:
                import re
                m = re.search(r"weight_kg=([\d.]+)", item.evidence)
                if m:
                    weight_kg = float(m.group(1))
                    break

        feature_summary = {
            "bounding_box": f"{bbox.length:.0f}×{bbox.width:.0f} mm" if bbox else "—",
            "hole_candidates": geo.candidate_count,
            "mfg_holes": mfg.total_holes,
            "mfg_threads": mfg.total_threads,
            "frames": len(mfg.frames),
            "assemblies": len(mfg.structure_assemblies),
            "accessories": len(mfg.structure_accessories),
            "welds": len(mfg.welds),
            "weight": f"{weight_kg:.1f} kg" if weight_kg else "—",
            "part_type": "加工件" if len(mfg.frames) == 0 else "結構件",
            "material_raw": part["material"],
        }

        return quote, feature_summary, None

    except FileNotFoundError as e:
        return None, {}, f"規則文件缺失：{e}"
    except Exception as e:
        return None, {}, f"解析失敗：{e}\n{traceback.format_exc()}"


# ---------------------------------------------------------------------------
# Main Demo Application
# ---------------------------------------------------------------------------

class DemoApp(tk.Tk):
    """Main Tkinter application window."""

    def __init__(
        self,
        session: UserSession | None = None,
        auth_service: AuthService | None = None,
    ):
        super().__init__()
        self.title("机械加工件智能报价系统")
        self.geometry("1280x720")
        self.minsize(1024, 600)
        self.configure(bg=CONTENT_BG)

        # Center on screen
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - 1280) // 2
        y = (sh - 720) // 2
        self.geometry(f"+{x}+{y}")

        # Font defaults
        self.option_add("*Font", (FONT_FAMILY[0], 10))

        # State
        self._current_quote: Quote | None = None
        self._current_feature_summary: dict[str, Any] = {}
        self._current_input_file: Path | None = None
        self._current_demo_name: str | None = None
        self._content: tk.Frame | None = None
        self._history = QuotationHistory()
        self._management = ManagementQueryService()
        self._session = session
        self._auth_service = auth_service
        self._supplier_maintenance: SupplierMaintenanceService | None = None
        self._price_approval: PriceApprovalService | None = None
        self._maintenance_actor = None
        self._external_skill_settings = None
        self._configure_authenticated_services()

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _configure_authenticated_services(self) -> None:
        self._supplier_maintenance = None
        self._price_approval = None
        self._maintenance_actor = None
        self._external_skill_settings = None
        session = self._session
        auth_service = self._auth_service
        if session is not None and auth_service is not None:
            from quotation.application.settings_service import UserSettingsService

            settings = UserSettingsService().load()
            root = Path(settings["smb_root"]) / "suppliers"
            self._supplier_maintenance = SupplierMaintenanceService(
                SupplierRepository(root),
                SupplierPriceRepository(root / "prices"),
                auth_service,
            )
            self._maintenance_actor = auth_service.get_user_by_username(session.username)
            if (
                self._maintenance_actor is not None
                and "system.config" in session.permissions
            ):
                self._external_skill_settings = build_external_skill_settings_service(
                    settings, auth_service
                )
            if "rule.approve" in session.permissions:
                from quotation.application.cache_sync_service import CacheSyncService

                storage = SmbStorageClient(settings["smb_root"])
                self._price_approval = PriceApprovalService(
                    SupplierPriceRepository(root / "prices"),
                    PriceReviewRepository(
                        Path(settings["smb_root"]) / "change-requests" / "price-reviews"
                    ),
                    storage,
                    auth_service,
                    CacheSyncService(storage, settings["smb_cache_dir"]),
                )

    def _activate_authentication(self) -> bool:
        """Authenticate on demand without restarting the desktop application."""

        from quotation.ui.auth_dialog import authenticate_desktop

        try:
            # Use the actual top-level owner so Windows keeps every login form
            # in front of the already visible desktop application.
            self.update_idletasks()
            context = authenticate_desktop(parent=self.winfo_toplevel())
        except Exception as exc:
            # GUI executables do not have a console.  Never leave an
            # authentication startup error invisible to the operator.
            messagebox.showerror(
                "管理员登录无法打开",
                f"登录组件初始化失败：{exc}\n\n"
                "请确认交付文件夹完整，并检查公司公共槽连接后重试。",
                parent=self,
            )
            return False
        if context is None:
            return False
        self._session = context.session
        self._auth_service = context.service
        self._configure_authenticated_services()
        self.title(
            f"机械加工件智能报价系统 — {context.session.display_name}"
            f"（{context.session.role.value}）"
        )
        self._main.destroy()
        self._content = None
        self._build_ui()
        return True

    def _logout_to_guest(self) -> None:
        """Drop the privileged desktop session and return to guest quotation mode."""

        self._session = None
        self._auth_service = None
        self._configure_authenticated_services()
        self.title("机械加工件智能报价系统")
        self._main.destroy()
        self._content = None
        self._build_ui()

    # ------------------------------------------------------------------
    # UI Setup
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        # Main container
        self._main = tk.Frame(self, bg=CONTENT_BG)
        self._main.pack(fill=tk.BOTH, expand=True)

        # Left nav
        allowed_items = self._allowed_nav_items()
        self._nav = NavPanel(
            self._main,
            on_nav_change=self._switch_page,
            allowed_items=allowed_items,
        )
        self._nav.pack(side=tk.LEFT, fill=tk.Y)

        # Right content area
        self._content_area = tk.Frame(self._main, bg=CONTENT_BG)
        self._content_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Show default page
        preferred = ["新建报价", "报价记录", "价格管理", "系统设置"]
        default_page = next(
            (name for name in preferred if allowed_items is None or name in allowed_items),
            "报价记录",
        )
        self._switch_page(default_page)

    def _allowed_nav_items(self) -> set[str] | None:
        if self._session is None:
            return {
                "新建报价",
                "批量报价",
                "报价记录",
                "价格管理",
                "管理员登录",
            }
        permissions = set(self._session.permissions)
        allowed: set[str] = set()
        if "quotation.create" in permissions:
            allowed.update({"新建报价", "批量报价"})
        if "quotation.view" in permissions:
            allowed.add("报价记录")
        if "price.view_cost" in permissions:
            allowed.update({"价格管理", "供应商管理"})
        if "rule.approve" in permissions:
            allowed.add("价格审核")
        if "user.view" in permissions:
            allowed.add("用户管理")
        if "system.config" in permissions:
            allowed.update({"外接Skill设置", "系统设置"})
        allowed.add("退出登录")
        return allowed

    def _switch_page(self, name: str) -> None:
        """Destroy current content and build the requested page."""
        if name == "管理员登录":
            self._activate_authentication()
            return
        if name == "退出登录":
            if messagebox.askyesno("退出登录", "退出当前账户并返回免登录模式？", parent=self):
                self._logout_to_guest()
            return
        if self._content is not None:
            self._content.destroy()

        permissions = set(self._session.permissions) if self._session is not None else None
        can_export = permissions is None or "quotation.export" in permissions
        can_review = permissions is not None and "price.modify" in permissions
        can_delete = permissions is not None and "quotation.delete" in permissions
        can_view_skill_debug = (
            permissions is not None and "skill.debug.view" in permissions
        )

        if name == "新建报价":
            self._content = NewQuotePage(
                self._content_area,
                on_load_j003=lambda: self._load_demo("J003"),
                on_load_w001=lambda: self._load_demo("W001"),
                on_select_file=self._select_quote_file,
                on_run_file=self._run_selected_quote,
                on_recalculate=self._recalculate_current_quote,
                on_export=self._export_excel if can_export else None,
                can_view_skill_debug=can_view_skill_debug,
            )
        elif name == "批量报价":
            self._content = BatchQuotePage(
                self._content_area,
                on_scan_dir=self._batch_scan_dir,
                on_scan_files=self._batch_scan_files,
                on_run_batch=self._batch_run,
                on_export_selected=self._batch_export_selected if can_export else None,
                on_export_all=self._batch_export_all if can_export else None,
                on_open_dir=self._batch_open_dir,
            )
        elif name == "报价记录":
            self._content = ManagementPage(
                self._content_area,
                title="报价记录与人工审核",
                columns=[
                    ("quote_id", "报价编号", 210), ("drawing_number", "图号", 130),
                    ("status_display", "状态", 130), ("cost_completion", "完成度%", 90),
                    ("total_incl_tax", "含稅總額", 110), ("quote_version", "版本", 70),
                    ("quoted_by", "报价人", 110), ("pc_username", "电脑登录者", 120),
                    ("pc_name", "电脑名称", 120), ("pc_ip", "IP地址", 120),
                    ("updated_at", "更新時間", 190),
                ],
                loader=self._load_history_rows,
                on_detail=self._show_history_detail,
                on_export=self._reexport_history if can_export else None,
                on_review=self._open_manual_review if can_review else None,
                actions=[("删除报价", self._delete_history_quote, True)] if can_delete else None,
                filter_values=["", "报价完整", "部分价格待确认", "需要人工审核", "图纸解析失败"],
            )
        elif name == "价格管理":
            self._content = ManagementPage(
                self._content_area,
                title="已发布公司价格表（只读）",
                columns=[
                    ("target_type_display", "类型", 100), ("canonical_code_display", "材料/工艺名称", 200),
                    ("specification", "規格", 130), ("unit_price", "未稅單價", 100),
                    ("unit_display", "单位", 70), ("origin_supplier_name", "来源供应商", 150),
                    ("price_version_id", "價格版本", 210),
                ],
                loader=self._load_pricebook_rows,
                on_detail=self._show_json_row,
                filter_values=["", "材料价格", "加工价格", "表面处理价格"],
            )
        elif name == "供应商管理":
            if self._supplier_maintenance is not None and self._maintenance_actor is not None:
                can_edit = bool(self._session and "price.modify" in self._session.permissions)
                actions = [("查看报价记录", self._view_supplier_prices, True)]
                if can_edit:
                    actions = [
                        ("新增供应商", self._add_supplier, False),
                        ("编辑供应商", self._edit_supplier, True),
                        ("新增报价", self._add_supplier_price, True),
                        ("导入报价Excel", self._import_supplier_prices, False),
                        ("停用", self._deactivate_supplier, True),
                        ("删除", self._delete_supplier, True),
                        *actions,
                    ]
                self._content = ManagementPage(
                    self._content_area,
                    title="供应商与原始报价维护（待审核报价不参与正式计算）",
                    columns=[
                        ("supplier_id", "供应商编号", 150),
                        ("supplier_name", "供应商名称", 180),
                        ("contact_person", "联系人", 100),
                        ("phone", "电话", 130),
                        ("status_display", "状态", 100),
                        ("quality_rating", "质量等级", 90),
                        ("updated_at", "更新时间", 190),
                    ],
                    loader=self._load_maintained_suppliers,
                    on_detail=self._show_json_row,
                    actions=actions,
                )
            else:
                self._content = ManagementPage(
                    self._content_area,
                    title="供应商报价（只读，待审核价格不参与正式计算）",
                    columns=[
                        ("record_id", "来源记录编号", 180), ("supplier_name", "供应商", 120),
                        ("material_display", "材料", 170), ("material_spec", "規格", 120),
                        ("parsed_value", "报价", 90), ("unit", "单位", 70),
                        ("status_display", "状态", 140),
                    ],
                    loader=self._load_supplier_rows,
                    on_detail=self._show_json_row,
                )
        elif name == "价格审核":
            self._content = ManagementPage(
                self._content_area,
                title="供应商价格审核与正式版本发布",
                columns=[
                    ("price_record_id", "来源记录编号", 190),
                    ("supplier_name", "供应商", 140),
                    ("target_name", "价格对象", 170),
                    ("unit_price", "原始报价", 90),
                    ("unit", "单位", 70),
                    ("tax_display", "税价口径", 120),
                    ("effective_from", "生效日期", 100),
                    ("review_status_display", "审核状态", 150),
                    ("published_price_version_id", "发布版本", 210),
                ],
                loader=self._load_price_approvals,
                on_detail=self._show_json_row,
                actions=[
                    ("批准并发布", self._approve_supplier_price, True),
                    ("驳回", self._reject_supplier_price, True),
                ],
                filter_values=["", "待审核", "已批准并发布", "已驳回"],
            )
        elif name == "用户管理":
            can_manage = bool(self._session and "user.manage" in self._session.permissions)
            actions = []
            if can_manage:
                actions = [
                    ("新增用户", self._add_user, False),
                    ("分配角色与权限", self._edit_user_access, True),
                    ("重置密码", self._reset_user_password, True),
                    ("启用账号", self._enable_user, True),
                    ("停用账号", self._disable_user, True),
                ]
            self._content = ManagementPage(
                self._content_area,
                title="用户、角色与功能权限管理",
                columns=[
                    ("username", "用户名", 140),
                    ("display_name", "姓名", 140),
                    ("role_display", "角色", 100),
                    ("status_display", "状态", 100),
                    ("permissions_display", "已授权功能", 520),
                    ("last_login_display", "最后登录", 180),
                ],
                loader=self._load_users,
                on_detail=self._show_json_row,
                actions=actions,
                filter_values=["", "启用", "停用", "锁定"],
            )
        elif name == "外接Skill设置":
            if self._external_skill_settings is None or self._maintenance_actor is None:
                raise PermissionError("当前用户没有外接 Skill 设置权限")
            from quotation.ui.external_skill_settings_page import (
                ExternalSkillSettingsPage,
            )

            self._content = ExternalSkillSettingsPage(
                self._content_area,
                self._external_skill_settings,
                self._maintenance_actor,
            )
        elif name == "系统设置":
            from quotation.application.settings_service import UserSettingsService
            self._content = SystemSettingsPage(
                self._content_area,
                UserSettingsService(),
                on_auth_required=self._activate_authentication,
            )
        else:
            raise ValueError(f"未知页面：{name}")

        self._content.pack(fill=tk.BOTH, expand=True)

    def _user_context(self):
        if self._session is None or self._auth_service is None:
            raise RuntimeError("用户管理需要先登录")
        actor = self._auth_service.get_user_by_username(self._session.username)
        if actor is None:
            raise RuntimeError("当前登录用户不存在")
        return self._auth_service, actor

    def _load_users(self, query: str, status: str) -> list[dict[str, Any]]:
        service, actor = self._user_context()
        status_codes = {"启用": "active", "停用": "disabled", "锁定": "locked"}
        role_names = {"admin": "管理员", "engineer": "工程师", "sales": "业务", "viewer": "查看者"}
        status_names = {"active": "启用", "disabled": "停用", "locked": "锁定"}
        catalog = service.permission_catalog(actor)
        rows = []
        for user in service.list_users(actor):
            permissions = service.get_user_permissions(user)
            if query and query.casefold() not in f"{user.username} {user.display_name}".casefold():
                continue
            if status_codes.get(status) and user.status.value != status_codes[status]:
                continue
            rows.append({
                "user_id": user.user_id,
                "username": user.username,
                "display_name": user.display_name,
                "role": user.role.value,
                "role_display": role_names[user.role.value],
                "status": user.status.value,
                "status_display": status_names[user.status.value],
                "permissions": permissions,
                "permissions_display": "、".join(catalog[p]["name"] for p in permissions),
                "permission_mode": "按角色默认" if user.assigned_permissions is None else "管理员单独分配",
                "last_login_display": user.last_login_time or "尚未登录",
            })
        return rows

    def _permission_dialog(
        self, title: str, *, role: UserRole, permissions: list[str]
    ) -> tuple[UserRole, list[str]] | None:
        service, actor = self._user_context()
        roles = service.role_catalog(actor)
        catalog = service.permission_catalog(actor)
        window = tk.Toplevel(self)
        window.title(title)
        window.transient(self)
        window.grab_set()
        role_names = {key: value["name"] for key, value in roles.items()}
        role_by_name = {value: key for key, value in role_names.items()}
        role_var = tk.StringVar(value=role_names[role.value])
        tk.Label(window, text="角色").pack(anchor=tk.W, padx=16, pady=(12, 3))
        role_box = ttk.Combobox(window, values=list(role_by_name), textvariable=role_var, state="readonly")
        role_box.pack(fill=tk.X, padx=16)
        tk.Label(window, text="功能权限（勾选后才显示或允许使用）").pack(anchor=tk.W, padx=16, pady=(12, 3))
        permission_vars: dict[str, tk.BooleanVar] = {}
        body = tk.Frame(window)
        body.pack(fill=tk.BOTH, expand=True, padx=16)
        for code, definition in catalog.items():
            variable = tk.BooleanVar(value=code in permissions)
            permission_vars[code] = variable
            ttk.Checkbutton(
                body,
                text=f"{definition['name']} — {definition['description']}",
                variable=variable,
            ).pack(anchor=tk.W, pady=2)

        def apply_role_defaults(_event=None):
            selected = role_by_name[role_var.get()]
            defaults = set(roles[selected].get("permissions", []))
            for code, variable in permission_vars.items():
                variable.set(code in defaults)

        role_box.bind("<<ComboboxSelected>>", apply_role_defaults)
        result: list[tuple[UserRole, list[str]]] = []

        def confirm():
            selected = [code for code, variable in permission_vars.items() if variable.get()]
            result.append((UserRole(role_by_name[role_var.get()]), selected))
            window.destroy()

        buttons = tk.Frame(window)
        buttons.pack(fill=tk.X, padx=16, pady=14)
        tk.Button(buttons, text="保存", command=confirm).pack(side=tk.RIGHT, padx=4)
        tk.Button(buttons, text="取消", command=window.destroy).pack(side=tk.RIGHT, padx=4)
        window.wait_window()
        return result[0] if result else None

    def _add_user(self, _row: dict[str, Any] | None) -> None:
        values = self._maintenance_form(
            "新增用户",
            [("username", "用户名"), ("display_name", "姓名"), ("password", "临时密码")],
        )
        if values is None:
            return
        service, actor = self._user_context()
        defaults = service.role_catalog(actor)[UserRole.VIEWER.value]["permissions"]
        access = self._permission_dialog(
            "分配新用户角色与权限",
            role=UserRole.VIEWER,
            permissions=list(defaults),
        )
        if access is None:
            return
        role, permissions = access
        try:
            service.create_user(
                actor,
                values["username"],
                values["password"],
                values["display_name"],
                role,
                assigned_permissions=permissions,
            )
            messagebox.showinfo("新增完成", "用户已建立，首次登录必须修改临时密码", parent=self)
        except (PermissionError, ValueError) as exc:
            messagebox.showerror("新增失败", str(exc), parent=self)

    def _edit_user_access(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        access = self._permission_dialog(
            f"分配权限 - {row['display_name']}",
            role=UserRole(row["role"]),
            permissions=list(row["permissions"]),
        )
        if access is None:
            return
        service, actor = self._user_context()
        role, permissions = access
        try:
            service.set_user_access(
                actor, row["user_id"], role=role, permissions=permissions
            )
        except (KeyError, PermissionError, ValueError) as exc:
            messagebox.showerror("保存失败", str(exc), parent=self)

    def _reset_user_password(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        values = self._maintenance_form(
            f"重置密码 - {row['display_name']}", [("password", "新临时密码")]
        )
        if values is None:
            return
        service, actor = self._user_context()
        try:
            service.reset_password(actor, row["user_id"], values["password"])
            messagebox.showinfo("重置完成", "用户下次登录必须修改临时密码", parent=self)
        except (KeyError, PermissionError, ValueError) as exc:
            messagebox.showerror("重置失败", str(exc), parent=self)

    def _enable_user(self, row: dict[str, Any] | None) -> None:
        self._set_user_status(row, UserStatus.ACTIVE)

    def _disable_user(self, row: dict[str, Any] | None) -> None:
        if row is None or not messagebox.askyesno(
            "确认停用", "停用后该用户将不能登录，是否继续？", parent=self
        ):
            return
        self._set_user_status(row, UserStatus.DISABLED)

    def _set_user_status(
        self, row: dict[str, Any] | None, status: UserStatus
    ) -> None:
        if row is None:
            return
        service, actor = self._user_context()
        try:
            service.set_user_status(actor, row["user_id"], status)
        except (KeyError, PermissionError, ValueError) as exc:
            messagebox.showerror("状态修改失败", str(exc), parent=self)

    def _load_history_rows(self, query: str, status: str) -> list[dict[str, Any]]:
        status_codes = {value: key for key, value in STATUS_LABELS.items()}
        return self._history.search(
            drawing_number=query or None, status=status_codes.get(status) or None, limit=500
        )

    def _load_pricebook_rows(self, query: str, target_type: str) -> list[dict[str, Any]]:
        type_codes = {value: key for key, value in TYPE_LABELS.items() if key.isupper()}
        rows = self._management.published_prices(
            target_type=type_codes.get(target_type) or None, query=query or None
        )["records"]
        for row in rows:
            row["target_type_display"] = TYPE_LABELS.get(row.get("target_type"), row.get("target_type"))
            row["unit_display"] = UNIT_LABELS.get(str(row.get("unit", "")).casefold(), row.get("unit"))
        return rows

    def _load_supplier_rows(self, query: str, _filter: str) -> list[dict[str, Any]]:
        rows = self._management.supplier_prices(query=query or None)["records"]
        for row in rows:
            row["status_display"] = STATUS_LABELS.get(row.get("status"), row.get("status"))
            row["material_display"] = display_value("material_code", row.get("material_code"))
        return rows

    def _maintenance_context(self):
        if self._supplier_maintenance is None or self._maintenance_actor is None:
            raise RuntimeError("供应商维护服务尚未登录")
        return self._supplier_maintenance, self._maintenance_actor

    def _load_maintained_suppliers(self, query: str, _filter: str) -> list[dict[str, Any]]:
        service, actor = self._maintenance_context()
        rows = [item.model_dump(mode="json") for item in service.list_suppliers(actor, query=query)]
        for row in rows:
            row["status_display"] = STATUS_LABELS.get(row.get("status"), row.get("status"))
        return rows

    def _approval_context(self):
        if self._price_approval is None or self._maintenance_actor is None:
            raise RuntimeError("当前用户没有价格审核权限")
        return self._price_approval, self._maintenance_actor

    def _load_price_approvals(self, query: str, status: str) -> list[dict[str, Any]]:
        service, actor = self._approval_context()
        status_codes = {
            "待审核": "PENDING_REVIEW",
            "已批准并发布": "APPROVED",
            "已驳回": "REJECTED",
        }
        rows = service.list_items(actor, status=status_codes.get(status), query=query or None)
        for row in rows:
            row["target_name"] = (
                row.get("material_code")
                or row.get("process_code")
                or row.get("surface_code")
                or "待确认"
            )
            row["target_name"] = display_value("target_name", row["target_name"])
            row["tax_display"] = "含税价" if row.get("tax_included") else "未税价"
        return rows

    def _approve_supplier_price(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        if row.get("review_status") != "PENDING_REVIEW":
            messagebox.showwarning("不能批准", "只有待审核的有效价格可以批准", parent=self)
            return
        values = self._maintenance_form(
            "批准并发布正式价格",
            [("review_comment", "审核意见（可选）")],
        )
        if values is None or not messagebox.askyesno(
            "确认发布",
            "批准后将生成不可修改的新正式价格快照，并立即切换所有客户端版本。是否继续？",
            parent=self,
        ):
            return
        service, actor = self._approval_context()
        try:
            review = service.approve(
                actor,
                supplier_id=row["supplier_id"],
                price_record_id=row["price_record_id"],
                review_comment=values["review_comment"],
            )
        except (KeyError, ValueError) as exc:
            messagebox.showerror("发布失败", str(exc), parent=self)
            return
        messagebox.showinfo(
            "发布完成",
            f"已生成正式价格版本：{review.published_price_version_id}\n原始供应商报价保持不变。",
            parent=self,
        )

    def _reject_supplier_price(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        if row.get("review_status") != "PENDING_REVIEW":
            messagebox.showwarning("不能驳回", "只有待审核的有效价格可以驳回", parent=self)
            return
        values = self._maintenance_form(
            "驳回供应商报价",
            [("review_comment", "驳回原因（必填）")],
        )
        if values is None:
            return
        service, actor = self._approval_context()
        try:
            service.reject(
                actor,
                supplier_id=row["supplier_id"],
                price_record_id=row["price_record_id"],
                review_comment=values["review_comment"],
            )
        except (KeyError, ValueError) as exc:
            messagebox.showerror("驳回失败", str(exc), parent=self)
            return
        messagebox.showinfo("审核完成", "已驳回；正式价格表没有修改。", parent=self)

    def _maintenance_form(
        self,
        title: str,
        fields: list[tuple[str, str]],
        initial: dict[str, Any] | None = None,
    ) -> dict[str, str] | None:
        window = tk.Toplevel(self)
        window.title(title)
        window.resizable(False, False)
        window.transient(self)
        window.grab_set()
        variables: dict[str, tk.StringVar] = {}
        result: dict[str, str] | None = None
        for index, (key, label) in enumerate(fields):
            tk.Label(window, text=label).grid(row=index, column=0, sticky=tk.W, padx=12, pady=6)
            variable = tk.StringVar(value=str((initial or {}).get(key) or ""))
            variables[key] = variable
            tk.Entry(window, textvariable=variable, width=38).grid(
                row=index, column=1, padx=12, pady=6
            )

        def confirm() -> None:
            nonlocal result
            result = {key: value.get().strip() for key, value in variables.items()}
            window.destroy()

        buttons = tk.Frame(window)
        buttons.grid(row=len(fields), column=0, columnspan=2, sticky=tk.E, padx=12, pady=12)
        tk.Button(buttons, text="取消", command=window.destroy).pack(side=tk.RIGHT)
        tk.Button(buttons, text="确定", command=confirm).pack(side=tk.RIGHT, padx=8)
        self.wait_window(window)
        return result

    def _add_supplier(self, _row: dict[str, Any] | None) -> None:
        values = self._maintenance_form(
            "新增供应商",
            [("supplier_id", "供应商编号"), ("supplier_name", "供应商名称"),
             ("contact_person", "联系人"), ("phone", "电话"), ("notes", "备注")],
        )
        if values is None:
            return
        service, actor = self._maintenance_context()
        try:
            service.create_supplier(actor, **values)
        except ValueError as exc:
            messagebox.showerror("新增失败", str(exc), parent=self)

    def _edit_supplier(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        values = self._maintenance_form(
            "编辑供应商",
            [("supplier_name", "供应商名称"), ("contact_person", "联系人"),
             ("phone", "电话"), ("quality_rating", "质量等级 A/B/C"), ("notes", "备注")],
            row,
        )
        if values is None:
            return
        service, actor = self._maintenance_context()
        try:
            service.update_supplier(actor, row["supplier_id"], values)
        except ValueError as exc:
            messagebox.showerror("保存失败", str(exc), parent=self)

    def _deactivate_supplier(self, row: dict[str, Any] | None) -> None:
        if row is None or not messagebox.askyesno(
            "确认停用", "停用后不能新增供应商报价，是否继续？", parent=self
        ):
            return
        service, actor = self._maintenance_context()
        service.set_supplier_status(actor, row["supplier_id"], SupplierStatus.INACTIVE)

    def _delete_supplier(self, row: dict[str, Any] | None) -> None:
        if row is None or not messagebox.askyesno(
            "确认删除", "仅无历史报价的供应商可删除，是否继续？", parent=self
        ):
            return
        service, actor = self._maintenance_context()
        try:
            service.delete_supplier(actor, row["supplier_id"])
        except ValueError as exc:
            messagebox.showerror("不能删除", str(exc), parent=self)

    def _add_supplier_price(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        values = self._maintenance_form(
            "新增供应商原始报价",
            [("target_type", "价格类型 MATERIAL/PROFILE/PROCESS/SURFACE"),
             ("target_code", "材料/工序/表面处理代码"), ("material_spec", "规格"),
             ("unit_price", "未税单价（未知请留空）"), ("unit", "单位"),
             ("effective_from", "生效日期 YYYY-MM-DD"), ("quote_number", "供应商报价单号")],
        )
        if values is None:
            return
        service, actor = self._maintenance_context()
        try:
            target_type = TargetType(values["target_type"].upper())
            code_fields = {
                TargetType.MATERIAL: {"material_code": values["target_code"]},
                TargetType.PROFILE: {"material_code": values["target_code"]},
                TargetType.PROCESS: {"process_code": values["target_code"]},
                TargetType.SURFACE: {"surface_code": values["target_code"]},
            }
            service.create_price_record(
                actor,
                supplier_id=row["supplier_id"],
                target_type=target_type,
                unit_price=float(values["unit_price"]) if values["unit_price"] else None,
                unit=values["unit"],
                material_spec=values["material_spec"] or None,
                effective_from=values["effective_from"] or None,
                quote_number=values["quote_number"] or None,
                **code_fields[target_type],
            )
        except (KeyError, ValueError) as exc:
            messagebox.showerror("新增报价失败", str(exc), parent=self)

    def _import_supplier_prices(self, _row: dict[str, Any] | None) -> None:
        path = filedialog.askopenfilename(
            title="导入供应商报价", filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not path:
            return
        service, actor = self._maintenance_context()
        try:
            result = service.import_price_excel(actor, path)
        except ValueError as exc:
            messagebox.showerror("导入失败", str(exc), parent=self)
            return
        messagebox.showinfo(
            "导入完成",
            f"成功：{result['导入成功']} 行\n失败：{result['导入失败']} 行",
            parent=self,
        )

    def _view_supplier_prices(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        service, actor = self._maintenance_context()
        records = [
            item.model_dump(mode="json")
            for item in service.list_price_records(actor, supplier_id=row["supplier_id"])
        ]
        StructuredDetailWindow(
            self,
            f"供应商报价记录 - {row['supplier_name']}",
            [("报价记录", [
                ("price_record_id", "记录编号", 180), ("target_type", "类型", 90),
                ("material_code", "材料", 110), ("material_spec", "规格", 100),
                ("process_code", "工序", 100), ("surface_code", "表面处理", 120),
                ("unit_price", "未税单价", 90), ("unit", "单位", 70),
                ("status", "状态", 120), ("effective_from", "生效日期", 100),
                ("quote_number", "报价单号", 120), ("created_at", "录入时间", 180),
            ], records)],
        )

    def _show_json_row(self, row: dict[str, Any]) -> None:
        StructuredDetailWindow(self, "资料明细", record_detail_sections(row))

    def _show_history_detail(self, row: dict[str, Any]) -> None:
        detail = self._history.get_detail(row["quote_id"])
        if detail:
            StructuredDetailWindow(self, f"报价明细 - {row['drawing_number']}", quote_detail_sections(detail))
        else:
            self._show_json_row(row)

    def _delete_history_quote(self, row: dict[str, Any] | None) -> None:
        if row is None:
            return
        if not messagebox.askyesno(
            "确认删除报价",
            f"将删除报价“{row['quote_id']}”及其明细和审核记录，是否继续？",
            parent=self,
        ):
            return
        if self._history.delete_quote(row["quote_id"]):
            messagebox.showinfo("删除完成", "报价记录已删除", parent=self)
        else:
            messagebox.showwarning("删除失败", "报价记录不存在或已被删除", parent=self)

    def _reexport_history(self, row: dict[str, Any]) -> None:
        path = filedialog.asksaveasfilename(
            title="重新导出报价",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"history_{row['drawing_number']}_{row.get('quote_version', 1)}.xlsx",
        )
        if path:
            export_history_quote(self._history, row["quote_id"], path)
            messagebox.showinfo("完成", f"已匯出：{path}")

    def _open_manual_review(self, row: dict[str, Any]) -> None:
        detail = self._history.get_detail(row["quote_id"])
        if not detail:
            return
        window = tk.Toplevel(self)
        window.title(f"人工審核 - {row['drawing_number']}")
        field_options = {
            "材料": "material", "厚度": "thickness", "尺寸": "dimensions",
            "表面处理": "surface_treatment", "加工方式": "process", "人工单价": "manual_price",
        }
        entries: dict[str, tk.Widget] = {}
        for index, (label, key) in enumerate([
            ("欄位", "field"), ("明細 line_id（人工價必填）", "line_id"),
            ("新值", "new_value"), ("原因", "reason"), ("操作者", "operator"),
        ]):
            tk.Label(window, text=label).grid(row=index, column=0, sticky=tk.W, padx=10, pady=6)
            if key == "field":
                widget = ttk.Combobox(window, values=list(field_options), state="readonly", width=38)
                widget.set("材料")
            else:
                widget = tk.Entry(window, width=42)
            widget.grid(row=index, column=1, padx=10, pady=6)
            entries[key] = widget
        line_ids = ", ".join(str(item["line_id"]) for item in detail["items"])
        tk.Label(window, text=f"可用 line_id：{line_ids}", fg="#7f8c8d").grid(
            row=5, column=0, columnspan=2, padx=10, pady=4
        )

        def submit() -> None:
            try:
                self._history.apply_manual_review(
                    row["quote_id"],
                    field_name=field_options[str(entries["field"].get())],
                    line_id=str(entries["line_id"].get()).strip() or None,
                    new_value=str(entries["new_value"].get()),
                    reason=str(entries["reason"].get()),
                    operator=str(entries["operator"].get()),
                )
                window.destroy()
                messagebox.showinfo("完成", "人工审核已保存；人工价仅对此报价生效")
            except Exception as exc:
                messagebox.showerror("審核失敗", str(exc))

        tk.Button(window, text="保存審核", command=submit).grid(
            row=6, column=0, columnspan=2, pady=12
        )

    # ------------------------------------------------------------------
    # Load demo part
    # ------------------------------------------------------------------

    def _select_quote_file(self) -> None:
        path = filedialog.askopenfilename(
            title="选择需要报价的图纸",
            filetypes=[
                ("机械图纸", "*.dxf *.DXF *.dwg *.DWG *.slddrw *.SLDDRW *.sldprt *.SLDPRT"),
                ("所有文件", "*.*"),
            ],
        )
        if not path:
            return
        self._current_input_file = Path(path)
        self._current_demo_name = None
        page = self._content
        if isinstance(page, NewQuotePage):
            page.clear()
            page.set_selected_file(path)
            page.update_status("图纸已选择，点击“开始解析”", "", 0)

    def _run_selected_quote(self) -> None:
        if self._current_input_file is None:
            messagebox.showwarning("提示", "请先选择 DWG、DXF 或 SolidWorks 图纸")
            return
        page = self._content
        if not isinstance(page, NewQuotePage):
            return
        page.update_status("正在解析并计算报价…", "orange", 0)
        self.update_idletasks()

        import threading

        selected = self._current_input_file
        use_ai = page.use_ai

        def report_progress(message: str, fraction: float) -> None:
            self.after(
                0,
                lambda m=message, f=fraction: (
                    page.update_status(m, "orange", max(0, min(f * 100, 99)))
                    if isinstance(self._content, NewQuotePage)
                    else None
                ),
            )

        def worker() -> None:
            try:
                service = self._create_quotation_service(use_ai)
                result = service.quote_single_file(
                    selected,
                    use_ai=use_ai,
                    progress_callback=report_progress,
                )
                self.after(0, lambda: self._finish_single_quote(result))
            except Exception as exc:
                message = str(exc) or type(exc).__name__
                self.after(0, lambda m=message: self._finish_single_quote_error(m))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_single_quote_error(self, detail: str) -> None:
        page = self._content
        if isinstance(page, NewQuotePage):
            page.update_status("报价失败", "red", 0)
        messagebox.showerror("报价失败", f"报价后台任务异常：{detail}", parent=self)

    def _finish_single_quote(self, result: Any) -> None:
        page = self._content
        if not isinstance(page, NewQuotePage):
            return
        if result.quote is None:
            detail = "；".join(result.errors) or "没有生成可用报价"
            page.update_status("报价失败", "red", 0)
            messagebox.showerror("报价失败", detail)
            return

        self._current_quote = result.quote
        self._current_feature_summary = result.feature_summary
        result.quote.quoted_by = self._quote_operator()
        self._history.save_quote(result, quoted_by=self._quote_operator())
        quote = result.quote
        tax = result.tax or TaxResult.calculate(quote.items, Decimal("0.13"))
        vm = QuoteViewModel(quote=quote, tax=tax)
        route_names = {"SHEET_METAL": "钣金件", "MACHINING": "机加工件"}
        summary = result.feature_summary
        dimensions = str(summary.get("bounding_box", "—")).replace("x", "×")
        page.update_basic_info([
            ("图号", quote.part_number or "—"),
            ("源文件", "、".join(result.source_files)),
            ("材料", quote.material or "待确认"),
            ("规格尺寸", dimensions),
            ("零件类型", route_names.get(summary.get("quotation_route"), "机械零件")),
            ("价格版本", quote.price_version or "—"),
        ])
        page.update_feature_summary(quote_feature_display_fields(summary))
        page.update_table(vm)
        page.update_trace(None)
        page.update_skill_debug(
            (result.ai_suggestions or {}).get("skill_debug_trace")
        )
        page.update_summary(vm)
        from quotation.ui.viewmodels import STATUS_DISPLAY
        status_text = STATUS_DISPLAY.get(result.status, vm.status_text)
        color = "green" if result.status == "COMPLETE" else "orange"
        page.update_status(status_text, color, quote.cost_completion)
        if result.warnings:
            messagebox.showwarning("需要注意", "\n".join(result.warnings[:5]))

    def _recalculate_current_quote(self) -> None:
        if self._current_input_file is not None:
            self._run_selected_quote()
        elif self._current_demo_name:
            self._load_demo(self._current_demo_name)
        else:
            messagebox.showwarning("提示", "请先选择图纸或载入示例")

    def _load_demo(self, part_name: str) -> None:
        """Run the pipeline and populate the UI."""
        page = self._content
        if not isinstance(page, NewQuotePage):
            return

        self._current_input_file = None
        self._current_demo_name = part_name
        page.set_selected_file(f"内置示例：{part_name}")

        page.update_status("解析中...", "orange", 0)
        self.update_idletasks()

        quote, feature_summary, error = run_quotation_pipeline(part_name)

        if error:
            messagebox.showerror("解析錯誤", error)
            page.update_status("解析失敗", "red", 0)
            return

        if quote is None:
            messagebox.showerror("错误", "无法生成报价")
            page.update_status("系統錯誤", "red", 0)
            return

        self._current_quote = quote
        self._current_feature_summary = feature_summary

        # Build ViewModel
        tax = TaxResult.calculate(quote.items, Decimal("0.13"))
        vm = QuoteViewModel(quote=quote, tax=tax)

        # Update basic info
        material_raw = feature_summary.get("material_raw", quote.material or "—")
        part_type = feature_summary.get("part_type", "—")
        basic_fields = [
            ("圖號", quote.part_number or "—"),
            ("料號", quote.part_number or "—"),
            ("材料", material_raw),
            ("規格尺寸", feature_summary.get("bounding_box", "—")),
            ("表面處理", self._extract_surface(quote)),
            ("零件類型", part_type),
            ("規則版本", quote.rule_version or "—"),
            ("價格版本", quote.price_version or "—"),
        ]
        page.update_basic_info(basic_fields)

        # Update feature summary
        fs = feature_summary
        feature_fields = [
            ("Bounding Box", fs.get("bounding_box", "—")),
            ("孔數", str(fs.get("mfg_holes", 0))),
            ("螺紋數", str(fs.get("mfg_threads", 0))),
            ("Frame 數", str(fs.get("frames", 0))),
            ("Assembly 數", str(fs.get("assemblies", 0))),
            ("Accessory 數", str(fs.get("accessories", 0))),
            ("重量", fs.get("weight", "—")),
            ("重量來源", "CAD bounding box 估算"),
            ("Confidence", f"{quote.overall_confidence:.0%}"),
        ]
        page.update_feature_summary(feature_fields)

        # Update table
        page.update_table(vm)

        # Update trace
        page.update_trace(None)

        # Update summary
        page.update_summary(vm)

        # Update status
        page.update_status(
            vm.status_text,
            vm.status_color,
            quote.cost_completion,
        )

    @staticmethod
    def _extract_surface(quote: Quote) -> str:
        """Extract surface treatment info from quote items."""
        for item in quote.items:
            if item.category == "surface":
                return item.name
        return "無"

    # ------------------------------------------------------------------
    # Export Excel
    # ------------------------------------------------------------------

    def _export_excel(self) -> None:
        """Export quotation to Excel using openpyxl."""
        if self._current_quote is None:
            messagebox.showwarning("警告", "请先载入报价数据")
            return

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            quote = self._current_quote
            fs = self._current_feature_summary

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"报价_{quote.part_number or '导出'}_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            )
            if not filepath:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "报价单"

            # Styles
            title_font = Font(name=FONT_FAMILY[0], size=16, bold=True)
            header_font = Font(name=FONT_FAMILY[0], size=11, bold=True)
            normal_font = Font(name=FONT_FAMILY[0], size=10)
            header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
            header_font_white = Font(name=FONT_FAMILY[0], size=10, bold=True, color="ffffff")
            warn_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            row = 1
            # Title
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            ws.cell(row=row, column=1, value="机械加工件智能报价系统 — 报价单").font = title_font
            row += 2

            # Basic info
            info_data = [
                ("圖號", quote.part_number),
                ("材料", quote.material),
                ("規格尺寸", fs.get("bounding_box", "—")),
                ("價格版本", quote.price_version),
                ("生成時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]
            for label, value in info_data:
                ws.cell(row=row, column=1, value=label).font = header_font
                ws.cell(row=row, column=2, value=str(value or "—")).font = normal_font
                row += 1
            row += 1

            # Feature summary
            ws.cell(row=row, column=1, value="特征摘要").font = header_font
            row += 1
            for label, value in quote_feature_display_fields(self._current_feature_summary):
                ws.cell(row=row, column=1, value=label).font = normal_font
                ws.cell(row=row, column=2, value=str(value)).font = normal_font
                row += 1
            row += 1

            # Quote items table
            headers = [
                "序号", "费用类别", "报价项目", "价格来源", "数量", "单位", "单价", "未税金额",
                "智能辅助参考总额", "智能辅助估价说明", "可信度", "状态",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row += 1

            tax = TaxResult.calculate(quote.items, Decimal("0.13"))
            for i, item in enumerate(quote.items, 1):
                is_u = item.source == PriceSource.U
                is_ai = item.source == PriceSource.AI
                values = [
                    i,
                    display_value("category", item.category),
                    item.name,
                    display_value("source", item.source.value),
                    "—" if is_u else item.quantity,
                    "—" if is_u else display_value("unit", item.unit),
                    "—" if is_u else item.unit_price,
                    "待確認" if is_u else item.amount,
                    item.ai_estimated_amount if item.ai_estimated_amount is not None else "—",
                    (
                        f"AI估算已计入本次报价合计，需人工确认；{item.ai_estimate_reason}"
                        if item.ai_estimated_amount is not None else "—"
                    ),
                    display_value("confidence", item.confidence.value),
                    "待确认" if is_u else "AI估算已计入，待人工确认" if is_ai else "已确认",
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = normal_font
                    cell.border = thin_border
                    if is_u or is_ai:
                        cell.fill = warn_fill
                row += 1
            row += 1

            # Tax summary
            tax_rows = [
                ("未稅小計", float(tax.subtotal_excluding_tax)),
                ("稅率", f"{float(tax.tax_rate) * 100:.0f}%"),
                ("稅額", float(tax.tax_amount)),
                ("含稅總價", float(tax.total_including_tax)),
                ("报价完整度", f"{quote.cost_completion:.1f}%"),
                ("價格版本", quote.price_version or "—"),
            ]
            for label, value in tax_rows:
                ws.cell(row=row, column=1, value=label).font = header_font
                cell = ws.cell(row=row, column=2, value=value)
                cell.font = Font(name=FONT_FAMILY[0], size=10, bold=True)
                row += 1

            # Column widths
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["H"].width = 20
            ws.column_dimensions["I"].width = 48
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 35
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 10
            ws.column_dimensions["G"].width = 12
            ws.column_dimensions["H"].width = 14
            ws.column_dimensions["I"].width = 20
            ws.column_dimensions["J"].width = 48
            ws.column_dimensions["K"].width = 10
            ws.column_dimensions["L"].width = 10

            wb.save(filepath)
            messagebox.showinfo("导出成功", f"报价已导出至：\n{filepath}")

        except Exception as e:
            messagebox.showerror("匯出失敗", f"Excel 匯出錯誤：{e}")

    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------

    def _on_close(self) -> None:
        self.destroy()

    # ------------------------------------------------------------------
    # Batch quotation callbacks
    # ------------------------------------------------------------------

    def _batch_scan_dir(self, directory: str, recursive: bool) -> list:
        from quotation.application.file_scanner import FileScanner
        scanner = FileScanner()
        return scanner.scan_directory(Path(directory), recursive=recursive)

    def _batch_scan_files(self, files: list[str]) -> list:
        from quotation.application.file_scanner import FileScanner
        scanner = FileScanner()
        return scanner.scan_selected_files([Path(file_path) for file_path in files])

    def _batch_run(self, bundles: list, use_ai: bool) -> list:
        svc = self._create_quotation_service(use_ai)
        results = svc.quote_batch(bundles, use_ai=use_ai)
        for result in results:
            if result.quote is not None:
                result.quote.quoted_by = self._quote_operator()
            self._history.save_quote(result, quoted_by=self._quote_operator())
        return results

    def _quote_operator(self) -> str:
        return self._session.display_name if self._session is not None else "免登录用户"

    def _create_quotation_service(self, use_ai: bool):
        """Create one service with the optional sidecar key, without exposing it."""
        from quotation.application.quotation_service import QuotationApplicationService
        from quotation.application.external_skill_router import build_external_skill_router
        from quotation.application.settings_service import UserSettingsService
        from quotation.infrastructure.ai.deepseek_client import DeepSeekClient
        from quotation.infrastructure.secrets.secret_locator import SecretLocator

        ai_client = None
        key = SecretLocator.get_deepseek_key()
        if key:
            ai_client = DeepSeekClient(api_key=key, timeout_seconds=20.0)

        return QuotationApplicationService(
            ai_client=ai_client,
            external_skill_router=build_external_skill_router(
                UserSettingsService().load(),
                ai_client=ai_client,
                debug_enabled=bool(
                    self._session and "skill.debug.view" in self._session.permissions
                ),
            ),
        )

    def _batch_export_selected(self, results: list, path: str) -> None:
        from quotation.application.batch_excel import export_batch_excel
        export_batch_excel(results, Path(path))

    def _batch_export_all(self, results: list, path: str) -> None:
        from quotation.application.batch_excel import export_batch_excel
        export_batch_excel(results, Path(path))

    def _batch_open_dir(self, path: str) -> None:
        import os
        full = Path(path)
        if full.exists():
            os.startfile(str(full.resolve()))
        else:
            messagebox.showinfo("提示", f"目錄不存在：{path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    """Launch directly in guest mode; privileged users can log in from the sidebar."""

    DemoApp().mainloop()


if __name__ == "__main__":
    main()
