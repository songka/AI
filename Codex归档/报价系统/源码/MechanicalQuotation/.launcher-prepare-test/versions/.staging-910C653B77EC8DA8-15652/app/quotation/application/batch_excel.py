"""Batch Excel export for quotation results.

Produces multi-sheet workbooks with Summary, Quote Details,
Review Required, Source Files, and Trace information.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from quotation.application.quotation_service import JobStatus, QuoteJobResult
from quotation.domain.quote import PriceSource
from quotation.ui.localization import display_value

# Font family for Excel
_FONT = "Microsoft YaHei UI"


def export_batch_excel(
    results: list[QuoteJobResult],
    output_path: str | Path,
    scan_directory: str = "",
) -> Path:
    """Export batch quotation results to a multi-sheet Excel workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Styles
    title_font = Font(name=_FONT, size=14, bold=True)
    header_font = Font(name=_FONT, size=10, bold=True, color="ffffff")
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    normal_font = Font(name=_FONT, size=9)
    warn_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _write_headers(ws, headers: list[str], row: int) -> int:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        return row + 1

    # ==================================================================
    # Sheet 1: Summary
    # ==================================================================
    ws = wb.active
    ws.title = "报价汇总"
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row=row, column=1, value="机械加工件智能报价系统 — 批量报价汇总").font = title_font
    row += 2

    ws.cell(row=row, column=1, value=f"扫描目录：{scan_directory}").font = normal_font
    row += 1
    ws.cell(
        row=row, column=1, value=f"生成时间：{datetime.now():%Y-%m-%d %H:%M:%S}"
    ).font = normal_font
    row += 1
    ws.cell(row=row, column=1, value=f"文件总数：{len(results)}").font = normal_font
    row += 2

    headers = [
        "图号",
        "文件名",
        "原始路径",
        "配对文件",
        "解析状态",
        "报价状态",
        "报价完整度",
        "待确认项数",
        "未税小计",
        "税率",
        "税额",
        "含税总价",
        "规则版本",
        "价格版本",
        "智能辅助",
        "生成时间",
    ]
    row = _write_headers(ws, headers, row)

    for jr in results:
        q = jr.quote
        matched = (
            ", ".join(f.file_name for f in jr.bundle.files if f != jr.bundle.geometry_source) or "-"
        )
        geom = jr.bundle.geometry_source
        geom_name = geom.file_name if geom else "-"
        values = [
            jr.drawing_number,
            geom_name,
            str(geom.full_path) if geom else "-",
            matched,
            display_value("status", jr.status),
            display_value("status", "COMPLETE" if jr.is_complete else jr.status),
            f"{jr.cost_completion:.1f}%",
            jr.unknown_item_count,
            float(jr.subtotal_excluding_tax),
            f"{float(jr.tax.tax_rate) * 100:.0f}%" if jr.tax else "13%",
            float(jr.tax.tax_amount) if jr.tax else 0,
            float(jr.total_including_tax),
            q.rule_version if q else "-",
            q.price_version if q else "-",
            "是" if jr.ai_used else "否",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = normal_font
            cell.border = thin_border
            if jr.status in (JobStatus.COMPLETE,):
                cell.fill = green_fill
            elif jr.status in (JobStatus.PARSE_FAILED, JobStatus.QUOTE_FAILED):
                cell.fill = red_fill
        row += 1

    # Column widths
    for col, w in enumerate([16, 28, 50, 28, 14, 16, 12, 10, 14, 8, 14, 14, 12, 12, 8, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 2: Quote Details
    # ==================================================================
    ws2 = wb.create_sheet("报价明细")
    row = 1
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws2.cell(row=row, column=1, value="报价明细").font = title_font
    row += 2

    detail_headers = [
        "图号",
        "报价项目",
        "来源",
        "数量",
        "单位",
        "单价",
        "未税金额",
        "状态",
        "智能辅助参考单价",
        "智能辅助参考总额",
        "智能辅助估价说明",
        "定价依据",
    ]
    row = _write_headers(ws2, detail_headers, row)

    for jr in results:
        if jr.quote is None:
            continue
        for item in jr.quote.items:
            is_u = item.source == PriceSource.U
            is_ai = item.source == PriceSource.AI
            values = [
                jr.drawing_number,
                item.name,
                display_value("source", item.source.value),
                "-" if is_u else item.quantity,
                "-" if is_u else display_value("unit", item.unit),
                "-" if is_u else item.unit_price,
                "待确认" if is_u else item.amount,
                "待确认" if is_u else "AI估算已计入，待人工确认" if is_ai else "已确认",
                item.ai_estimated_unit_price if item.ai_estimated_unit_price is not None else "-",
                item.ai_estimated_amount if item.ai_estimated_amount is not None else "-",
                (
                    f"AI估算已计入本次报价合计，需人工确认；{item.ai_estimate_reason}"
                    if item.ai_estimated_amount is not None
                    else "-"
                ),
                item.evidence or item.note or "-",
            ]
            for col, v in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=v)
                cell.font = normal_font
                cell.border = thin_border
                if is_u or is_ai:
                    cell.fill = warn_fill
            row += 1

    for col, w in enumerate([16, 30, 12, 10, 10, 12, 14, 10, 18, 18, 48, 70], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 3: Review Required
    # ==================================================================
    ws3 = wb.create_sheet("待人工审核")
    row = 1
    ws3.cell(row=row, column=1, value="待审核项目").font = title_font
    row += 2

    review_headers = ["图号", "缺失信息", "未知成本项", "解析警告", "智能辅助建议", "建议处理"]
    row = _write_headers(ws3, review_headers, row)

    for jr in results:
        if jr.status in (JobStatus.COMPLETE,):
            continue
        unknown_items = []
        if jr.quote:
            unknown_items = [
                i.name
                for i in jr.quote.items
                if i.source in {PriceSource.U, PriceSource.AI}
            ]
        ai_str = ""
        if jr.ai_suggestions:
            ai_str = "、".join(
                display_value("field_name", field)
                for field in jr.ai_suggestions.get("missing_fields", [])
            )
        values = [
            jr.drawing_number,
            ", ".join(jr.warnings) if jr.warnings else "-",
            ", ".join(unknown_items) if unknown_items else "-",
            ", ".join(jr.warnings) if jr.warnings else "-",
            ai_str if ai_str else "-",
            "人工确认材料、厚度、表面处理及加工方式" if unknown_items else "-",
        ]
        for col, v in enumerate(values, 1):
            cell = ws3.cell(row=row, column=col, value=v)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = warn_fill
        row += 1

    for col, w in enumerate([16, 25, 30, 25, 30, 30], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 4: Source Files
    # ==================================================================
    ws4 = wb.create_sheet("源文件")
    row = 1
    ws4.cell(row=row, column=1, value="原始文件清单").font = title_font
    row += 2

    src_headers = ["原始文件名", "文件类型", "完整路径", "配对关系", "几何来源"]
    row = _write_headers(ws4, src_headers, row)

    for jr in results:
        geom = jr.bundle.geometry_source
        for f in jr.bundle.files:
            values = [
                f.file_name,
                f.extension.upper().lstrip("."),
                str(f.full_path),
                display_value(
                    "status",
                    "MATCHED" if jr.bundle.match_status == MatchStatus.MATCHED else "UNMATCHED",
                ),
                "是" if f == geom else "否",
            ]
            for col, v in enumerate(values, 1):
                ws4.cell(row=row, column=col, value=v).font = normal_font
                ws4.cell(row=row, column=col).border = thin_border
            row += 1

    for col, w in enumerate([30, 10, 55, 12, 10], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 5: Trace
    # ==================================================================
    ws5 = wb.create_sheet("价格来源追踪")
    row = 1
    ws5.cell(row=row, column=1, value="价格来源追踪").font = title_font
    row += 2

    trace_headers = [
        "图号",
        "项目",
        "报价价格来源",
        "定价依据",
        "价格版本",
        "公司价格编号",
        "原始报价记录编号",
        "原始供应商编号",
        "计价口径",
        "是否使用回退规则",
    ]
    row = _write_headers(ws5, trace_headers, row)

    for jr in results:
        if jr.quote is None:
            continue
        for item in jr.quote.items:
            values = [
                jr.drawing_number,
                item.name,
                display_value("quote_price_source", item.quote_price_source),
                display_value("resolution_source", item.resolution_source),
                item.price_version_id or "-",
                item.company_price_id or "-",
                item.origin_price_record_id or "-",
                item.origin_supplier_id or "-",
                item.price_basis or "-",
                "是" if item.fallback_warning else "否",
            ]
            for col, v in enumerate(values, 1):
                ws5.cell(row=row, column=col, value=v).font = normal_font
                ws5.cell(row=row, column=col).border = thin_border
            row += 1

    for col, w in enumerate([16, 28, 16, 24, 16, 22, 22, 18, 14, 12], 1):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ==================================================================
    # Sheet 6: DWG Conversion Trace
    # ==================================================================
    ws6 = wb.create_sheet("DWG转换记录")
    row = 1
    ws6.cell(row=row, column=1, value="DWG转换追踪").font = title_font
    row += 2
    conversion_headers = [
        "图号",
        "原始DWG",
        "转换状态",
        "转换器",
        "配置来源",
        "缓存命中",
        "转换耗时（毫秒）",
        "临时DXF",
        "原始文件未修改",
        "中文状态或错误",
    ]
    row = _write_headers(ws6, conversion_headers, row)
    for jr in results:
        trace = jr.dwg_conversion
        if not trace:
            continue
        values = [
            jr.drawing_number,
            trace.get("source_file", "-"),
            display_value("status", trace.get("status", "-")),
            trace.get("adapter", "-"),
            {
                "user_settings": "用户设置",
                "local_appdata": "本机用户目录",
                "windows_common_path": "系统安装目录",
                "environment": "环境变量",
                "PATH": "系统搜索路径",
                "none": "未配置",
            }.get(trace.get("configuration_source"), trace.get("configuration_source", "-")),
            "是" if trace.get("cache_hit") else "否",
            trace.get("duration_ms", 0),
            trace.get("converted_file") or "-",
            "是" if trace.get("original_preserved") else "否",
            trace.get("error") or "转换成功",
        ]
        for col, value in enumerate(values, 1):
            cell = ws6.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
        row += 1
    for col, width in enumerate([16, 55, 18, 22, 18, 12, 16, 55, 18, 45], 1):
        ws6.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(str(output_path))
    return output_path


def export_single_excel(
    result: QuoteJobResult,
    output_path: str | Path,
) -> Path:
    """Export a single quotation to Excel (same format as batch but single-item)."""
    return export_batch_excel([result], output_path)


# Re-export for convenience
MatchStatus = __import__("quotation.application.file_scanner", fromlist=["MatchStatus"]).MatchStatus
