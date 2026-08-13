# -*- coding: utf-8 -*-
"""Phase 4.6.5: Company Price Publication from reviewed Excel v1.2."""

import json, hashlib, copy, uuid
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

from quotation.application.price_publication import (
    build_supplier_provenance_index,
    origin_supplier_id_for,
)

REVIEWED = Path("data/price-review-r01-v1.2-reviewed-complete.xlsx")
IMPORT_PKG = Path("rules/imports/r01-v1.0/pricing-rules-excel-r01-v1.0.json")
OUT_BOOK = Path("data/company-pricebook-r01-v1.0-draft.json")
OUT_SNAPSHOT = Path("data/company-pricebook-r01-v1.0-snapshot.json")

now = datetime.now(timezone.utc).isoformat()
blocking = []
warnings = []
company_prices = []

with open(IMPORT_PKG, encoding="utf-8") as import_file:
    import_package = json.load(import_file)
supplier_provenance = build_supplier_provenance_index(
    import_package.get("pricing_source_records", [])
)

def log_block(msg):
    blocking.append(msg); print(f"  BLOCK: {msg}")
def log_warn(msg):
    warnings.append(msg); print(f"  WARN: {msg}")

# Load reviewed workbook
wb = openpyxl.load_workbook(REVIEWED)

# === 1. Parse Company Price Candidates ===
ws1 = wb["Company Price Candidates"]
# Find admin columns (they start after the supplier data columns)
header = [c.value for c in ws1[1]]
# Find admin column indices
publish_col = None; origin_col = None; price_col = None; basis_col = None
eff_from_col = None; approver_col = None
for i, h in enumerate(header):
    if h == "Publish?": publish_col = i
    elif h == "Selected Origin Record ID": origin_col = i
    elif h == "Company Price": price_col = i
    elif h == "Price Basis": basis_col = i
    elif h == "Effective From": eff_from_col = i
    elif h == "Approver": approver_col = i

if None in (publish_col, origin_col, price_col, basis_col, eff_from_col, approver_col):
    log_block("Missing admin columns in Company Price Candidates")
else:
    for row in ws1.iter_rows(min_row=2, values_only=True):
        mat = row[0]; spec = row[1]; unit = row[2]
        if not mat: continue
        publish = str(row[publish_col]).strip().upper() if row[publish_col] else ""
        if publish != "TRUE": continue

        origin_id = str(row[origin_col]).strip() if row[origin_col] else ""
        company_price = row[price_col]
        basis = str(row[basis_col]).strip() if row[basis_col] else ""
        eff_from = row[eff_from_col]
        approver = str(row[approver_col]).strip() if row[approver_col] else ""

        # Validate
        if not origin_id: log_block(f"Material {mat}: no origin record ID"); continue
        if not company_price or float(company_price) <= 0: log_block(f"Material {mat}: invalid price {company_price}"); continue
        if not basis: log_block(f"Material {mat}: no price basis"); continue
        if not eff_from: log_block(f"Material {mat}: no effective date"); continue
        if not approver: log_block(f"Material {mat}: no approver"); continue

        company_prices.append({
            "company_price_id": f"CP-{uuid.uuid4().hex[:12]}",
            "target_type": "MATERIAL",
            "canonical_code": mat,
            "specification": spec if spec else None,
            "unit_price": float(company_price),
            "unit": unit,
            "currency": "CNY",
            "price_basis": basis,
            "effective_from": str(eff_from)[:10] if eff_from else None,
            "effective_to": None,
            "origin_type": "SUPPLIER_PRICE_RECORD",
            "origin_supplier_id": origin_supplier_id_for(origin_id, supplier_provenance),
            "origin_price_record_id": origin_id,
            "selection_policy": "MANUAL_ADMIN_SELECTION",
            "approved_by": approver,
            "approved_at": now,
            "price_version_id": "R01-COMPANY-PRICE-V1.0",
        })

print(f"Material C prices selected: {len(company_prices)}")

# === 2. Parse Process Rate Candidates ===
ws2 = wb["Process Rate Candidates"]
h2 = [c.value for c in ws2[1]]
pc = {h: i for i, h in enumerate(h2) if h}
for row in ws2.iter_rows(min_row=2, values_only=True):
    proc = row[0]; publish = str(row[pc.get("Publish?",6)]).strip().upper() if row[pc.get("Publish?",6)] else ""
    if publish != "TRUE": continue
    cp = row[pc.get("Company Price",7)]
    basis = str(row[pc.get("Price Basis",8)]).strip() if row[pc.get("Price Basis",8)] else ""
    eff = row[pc.get("Effective From",9)]
    appr = str(row[pc.get("Approver",12)]).strip() if row[pc.get("Approver",12)] else ""
    if not cp or float(cp) <= 0: log_block(f"Process {proc}: invalid price"); continue
    company_prices.append({
        "company_price_id": f"CP-{uuid.uuid4().hex[:12]}",
        "target_type": "PROCESS",
        "canonical_code": proc,
        "unit_price": float(cp),
        "unit": "hour",
        "currency": "CNY",
        "price_basis": basis,
        "effective_from": str(eff)[:10] if eff else None,
        "approved_by": appr,
        "price_version_id": "R01-COMPANY-PRICE-V1.0",
    })

# === 3. Parse Surface Rate Candidates ===
ws3 = wb["Surface Rate Candidates"]
h3 = [c.value for c in ws3[1]]
sc = {h: i for i, h in enumerate(h3) if h}
for row in ws3.iter_rows(min_row=2, values_only=True):
    surf = row[0]; publish = str(row[sc.get("Publish?",6)]).strip().upper() if row[sc.get("Publish?",6)] else ""
    if publish != "TRUE": continue
    cp = row[sc.get("Company Price",7)]
    basis = str(row[sc.get("Price Basis",8)]).strip() if row[sc.get("Price Basis",8)] else ""
    eff = row[sc.get("Effective From",9)]
    appr = str(row[sc.get("Approver",12)]).strip() if row[sc.get("Approver",12)] else ""
    company_prices.append({
        "company_price_id": f"CP-{uuid.uuid4().hex[:12]}",
        "target_type": "SURFACE",
        "canonical_code": surf,
        "unit_price": float(cp),
        "unit": "kg",
        "currency": "CNY",
        "price_basis": basis,
        "effective_from": str(eff)[:10] if eff else None,
        "approved_by": appr,
        "price_version_id": "R01-COMPANY-PRICE-V1.0",
    })

# === 4. Exceptions check ===
ws5 = wb["Exceptions"]
exc_unresolved = 0
for row in ws5.iter_rows(min_row=2, values_only=True):
    resolved = str(row[8]).strip().upper() if len(row) > 8 and row[8] else "FALSE"
    if resolved != "TRUE":
        exc_unresolved += 1
        log_warn(f"Unresolved exception: {row[1]} {row[4]}")

# === 5. Duplicate check ===
keys = {}
for cp in company_prices:
    k = (cp["target_type"], cp["canonical_code"], cp.get("specification",""), cp.get("unit",""))
    if k in keys:
        log_block(f"Duplicate: {k}")
    keys[k] = cp

# === 6. Count summary ===
mat_count = sum(1 for c in company_prices if c["target_type"] == "MATERIAL")
proc_count = sum(1 for c in company_prices if c["target_type"] == "PROCESS")
surf_count = sum(1 for c in company_prices if c["target_type"] == "SURFACE")

print(f"\n=== Publication Summary ===")
print(f"Material: {mat_count} | Process: {proc_count} | Surface: {surf_count}")
print(f"Total C prices: {len(company_prices)}")
print(f"Blocking errors: {len(blocking)}")
print(f"Warnings: {len(warnings)}")

# === 7. Generate Price Book ===
snapshot = hashlib.sha256(json.dumps(company_prices, sort_keys=True, ensure_ascii=False).encode()).hexdigest()

book = {
    "price_version_id": "R01-COMPANY-PRICE-V1.0",
    "version": "1.0.0",
    "status": "PUBLISHED" if len(blocking) == 0 else "DRAFT_BLOCKED",
    "effective_from": "2026-08-01",
    "created_by": "SYSTEM",
    "approved_by": "songka",
    "created_at": now,
    "approved_at": now,
    "source_package_sha256": hashlib.sha256(IMPORT_PKG.read_bytes()).hexdigest() if IMPORT_PKG.exists() else "",
    "snapshot_sha256": snapshot,
    "record_count": len(company_prices),
    "material_count": mat_count,
    "process_count": proc_count,
    "surface_count": surf_count,
    "blocking_errors": len(blocking),
    "warnings": len(warnings),
    "blocking_error_list": blocking,
    "warning_list": warnings,
    "unresolved_exceptions": exc_unresolved,
    "notes": "Published from admin review v1.2. All prices EXCLUDING_TAX. Tax model disabled.",
    "company_prices": company_prices,
}

Path("data").mkdir(exist_ok=True)
with open(OUT_BOOK, "w", encoding="utf-8") as f:
    json.dump(book, f, ensure_ascii=False, indent=2)
with open(OUT_SNAPSHOT, "w", encoding="utf-8") as f:
    json.dump(book, f, ensure_ascii=False, indent=2)

print(f"\nSaved: {OUT_BOOK}")
print(f"Snapshot SHA256: {snapshot}")
print(f"Status: {book['status']}")

# === 8. Auto-continue condition ===
if len(blocking) == 0:
    print("\n✅ No blocking errors. Auto-publishing R01-COMPANY-PRICE-V1.0.")
else:
    print(f"\n❌ {len(blocking)} blocking errors. Fix required.")

# Output JSON summary for test consumption
summary = {
    "status": book["status"],
    "material_count": mat_count,
    "process_count": proc_count,
    "surface_count": surf_count,
    "total": len(company_prices),
    "blocking": len(blocking),
    "warnings": len(warnings),
    "snapshot_sha256": snapshot,
}
with open("data/phase465_publication_summary.json", "w", encoding="utf-8") as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)
