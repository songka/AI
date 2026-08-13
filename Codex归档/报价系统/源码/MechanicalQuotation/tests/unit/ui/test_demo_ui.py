"""Smoke tests for the quotation demo UI module."""

from __future__ import annotations

import json
import tempfile
from decimal import Decimal
from pathlib import Path

import pytest

from quotation.domain.quote import PriceSource, QuoteConfidence, QuoteItem
from quotation.ui.viewmodels import (
    QuoteItemViewModel,
    QuoteViewModel,
    TaxResult,
)


# ============================================================================
# Test 1: UI module imports
# ============================================================================

class TestUIModuleImports:
    def test_viewmodels_imports(self):
        """UI module can be imported."""
        from quotation.ui import viewmodels
        assert viewmodels is not None

    def test_widgets_imports(self):
        """Widgets module can be imported (tkinter required)."""
        import tkinter
        from quotation.ui import widgets
        assert widgets is not None

    def test_demo_app_imports(self):
        """Demo app module can be imported."""
        import tkinter
        from quotation.ui import demo_app
        assert demo_app is not None

    def test_dwg_status_labels_are_chinese(self):
        from quotation.ui.viewmodels import STATUS_DISPLAY
        assert STATUS_DISPLAY["DWG_CONVERTING"] == "正在轉換DWG圖紙"
        assert STATUS_DISPLAY["DWG_CONVERSION_FAILED"] == "DWG轉換失敗"

    def test_structured_details_use_tabs_and_chinese_labels(self):
        from quotation.ui.widgets import quote_detail_sections, record_detail_sections

        record = record_detail_sections({"target_type": "MATERIAL", "unit": "kg"})
        assert record[0][0] == "基本信息"
        assert {row["field"] for row in record[0][2]} == {"价格类型", "单位"}
        detail = quote_detail_sections({
            "quote": {"quote_id": "Q-1", "quotation_status": "COMPLETE"},
            "items": [{
                "line_id": "L-1", "category": "material", "name": "钢材",
                "source_display": "公司核准价格", "quantity": 1, "unit": "kg",
                "unit_price": 10, "amount": 10, "confidence": "high",
                "status": "已确认", "resolution_display": "已发布公司价格表",
            }],
            "overrides": {}, "reviews": [],
        })
        assert [section[0] for section in detail] == ["报价摘要", "费用明细", "人工调整", "审核记录"]
        assert detail[1][2][0]["unit"] == "千克"
        assert detail[1][2][0]["confidence"] == "高"

    def test_feature_card_separates_itemized_total_from_model_reference(self):
        from quotation.ui.demo_app import quote_feature_display_fields

        fields = dict(
            quote_feature_display_fields(
                {
                    "bounding_box": "60x70 mm",
                    "itemized_subtotal": "128.50 元",
                    "feature_calibration_reference": "160.00 元（仅供审核，不计入正式合计）",
                }
            )
        )

        assert fields["分项未税合计"] == "128.50 元"
        assert fields["整件模型参考价（不计入）"].startswith("160.00 元")


# ============================================================================
# Test 2 & 3: Tax calculation
# ============================================================================

class TestTaxCalculation:
    def test_tax_17_percent_correct(self):
        """TaxCalculator gives 17% rate."""
        items = [
            QuoteItem(line_id="M1", category="material", name="Steel", amount=1000,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
        ]
        tax = TaxResult.calculate(items, Decimal("0.17"))
        assert tax.tax_rate == Decimal("0.17")

    def test_total_including_tax_equals_subtotal_times_1_17(self):
        """Tax-inclusive total = subtotal × 1.17."""
        items = [
            QuoteItem(line_id="M1", category="material", name="Steel", amount=1000,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
            QuoteItem(line_id="P1", category="process", name="CNC", amount=500,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
        ]
        tax = TaxResult.calculate(items, Decimal("0.17"))
        expected_subtotal = Decimal("1500.00")
        assert tax.subtotal_excluding_tax == expected_subtotal
        expected_total = Decimal("1755.00")  # 1500 * 1.17
        assert tax.total_including_tax == expected_total

    def test_unknown_items_excluded_from_tax(self):
        """Unknown (source=U) items should not contribute to tax base."""
        items = [
            QuoteItem(line_id="M1", category="material", name="Steel", amount=1000,
                      source=PriceSource.C, confidence=QuoteConfidence.HIGH),
            QuoteItem(line_id="U1", category="process", name="Magic", amount=0,
                      source=PriceSource.U, confidence=QuoteConfidence.UNCERTAIN),
        ]
        tax = TaxResult.calculate(items, Decimal("0.17"))
        assert tax.subtotal_excluding_tax == Decimal("1000.00")
        assert tax.total_including_tax == Decimal("1170.00")

    def test_ai_estimate_is_included_in_tax_and_marked_for_review(self):
        item = QuoteItem(
            line_id="AI1",
            category="process",
            name="焊接加工",
            quantity=1,
            unit="项",
            unit_price=300,
            amount=300,
            source=PriceSource.AI,
            confidence=QuoteConfidence.UNCERTAIN,
            ai_estimated_amount=300,
        )
        tax = TaxResult.calculate([item], Decimal("0.13"))
        vm = QuoteItemViewModel(item=item)

        assert tax.subtotal_excluding_tax == Decimal("300.00")
        assert tax.total_including_tax == Decimal("339.00")
        assert vm.status_label == "AI估算已计入，需人工确认"
        assert "unknown" in vm.row_tags


# ============================================================================
# Test 4: Unknown amount display
# ============================================================================

class TestUnknownDisplay:
    def test_unknown_amount_displays_as_pending(self):
        """Unknown items show '—' not '¥0.00'."""
        item = QuoteItem(
            line_id="U1", category="process", name="TAP", amount=0,
            source=PriceSource.U, confidence=QuoteConfidence.UNCERTAIN,
        )
        vm = QuoteItemViewModel(item=item, index=1)
        assert vm.is_unknown
        assert vm.display_amount == "—"
        assert vm.display_unit_price == "—"
        assert vm.status_label == "待确认"
        assert ("定价状态", "未找到可用价格") in vm.trace_fields
        assert any(label == "未定价原因" for label, _value in vm.trace_fields)
        assert "unknown" in vm.row_tags

    def test_known_zero_amount_displays_normally(self):
        """Known source with zero amount displays ¥0.00."""
        item = QuoteItem(
            line_id="Z1", category="purchased", name="FreePart", amount=0,
            source=PriceSource.M, confidence=QuoteConfidence.MEDIUM,
        )
        vm = QuoteItemViewModel(item=item, index=1)
        assert not vm.is_unknown
        assert vm.display_amount == "¥0.00"
        assert vm.status_label == "已确认"


# ============================================================================
# Test 5 & 6: J003 / W001 pipeline → ViewModel
# ============================================================================

class TestDemoPipeline:
    def test_j003_viewmodel_generates(self):
        """Full J003 pipeline → QuoteViewModel with tax."""
        from quotation.ui.demo_app import run_quotation_pipeline

        quote, feature_summary, error = run_quotation_pipeline("J003")
        assert error is None, f"Pipeline error: {error}"
        assert quote is not None
        assert len(quote.items) > 0
        assert quote.part_number == "UC1000005854"
        # TAP still comes from a DRAFT legacy rule, so it must remain review-required.
        assert quote.quotation_status == "INCOMPLETE"
        # Material weight is no longer inferred from the drawing-sheet extent;
        # material and draft TAP pricing therefore remain explicitly unknown.
        assert quote.cost_completion == 50.0
        assert any(item.fallback_warning for item in quote.items)

        # Build ViewModel
        tax = TaxResult.calculate(quote.items, Decimal("0.17"))
        vm = QuoteViewModel(quote=quote, tax=tax)
        assert vm.status_color == "orange"
        assert "待確認" in vm.status_text
        assert len(vm.items_vm) > 0

    def test_w001_viewmodel_generates(self):
        """Full W001 pipeline → QuoteViewModel with tax."""
        from quotation.ui.demo_app import run_quotation_pipeline

        quote, feature_summary, error = run_quotation_pipeline("W001")
        assert error is None, f"Pipeline error: {error}"
        assert quote is not None
        assert len(quote.items) > 0
        assert quote.part_number == "UC2020083221"
        assert quote.quotation_status == "INCOMPLETE"
        # No evidence-backed CNC operation exists. Two of seven remaining
        # items require review under the stricter formal-price policy.
        assert quote.cost_completion == pytest.approx(71.4, rel=0.01)
        assert not any("CNC" in item.name for item in quote.items)

        # Build ViewModel
        tax = TaxResult.calculate(quote.items, Decimal("0.17"))
        vm = QuoteViewModel(quote=quote, tax=tax)
        assert vm.status_color == "orange"
        assert "待確認" in vm.status_text
        assert len(vm.unknown_items) >= 1
        assert len(vm.items_vm) > 0


# ============================================================================
# Test 7: Excel export
# ============================================================================

class TestExcelExport:
    def test_excel_export_succeeds(self):
        """Export creates valid .xlsx file."""
        import openpyxl

        from quotation.ui.demo_app import run_quotation_pipeline

        quote, _, error = run_quotation_pipeline("J003")
        assert error is None
        assert quote is not None

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmppath = Path(f.name)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "報價單"

            # Write basic data
            ws.cell(row=1, column=1, value="機械加工件智能報價系統 — 報價單")
            ws.cell(row=3, column=1, value="圖號")
            ws.cell(row=3, column=2, value=quote.part_number or "")

            # Write items
            row = 8
            for i, item in enumerate(quote.items, 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=item.name)
                ws.cell(row=row, column=3, value=item.source.value)
                if item.source == PriceSource.U:
                    ws.cell(row=row, column=7, value="待確認")
                else:
                    ws.cell(row=row, column=7, value=item.amount)
                row += 1

            # Tax
            tax = TaxResult.calculate(quote.items, Decimal("0.17"))
            ws.cell(row=row + 1, column=1, value="未稅小計")
            ws.cell(row=row + 1, column=2, value=float(tax.subtotal_excluding_tax))
            ws.cell(row=row + 2, column=1, value="含稅總價")
            ws.cell(row=row + 2, column=2, value=float(tax.total_including_tax))

            wb.save(str(tmppath))
            assert tmppath.exists()
            assert tmppath.stat().st_size > 0

            # Verify it's a valid xlsx
            wb2 = openpyxl.load_workbook(str(tmppath))
            assert "報價單" in wb2.sheetnames
        finally:
            tmppath.unlink(missing_ok=True)


# ============================================================================
# UI Startup Smoke Test
# ============================================================================

class TestUIStartup:
    def test_auth_form_is_visible_with_hidden_parent(self):
        """Windows must show the setup form even though the auth root is hidden."""
        import tkinter as tk

        try:
            root = tk.Tk()
            root.withdraw()
            from quotation.ui.auth_dialog import _FormDialog

            dialog = _FormDialog(root, "建立首位管理员", [("username", "用户名", False)])
            root.update()
            assert dialog.state() == "normal"
            assert dialog.winfo_viewable() == 1
            dialog.destroy()
            root.destroy()
        except tk.TclError:
            pytest.skip("Tk runtime unavailable")

    def test_demo_app_starts_and_destroys(self):
        """DemoApp can be created, renders, and destroyed without error."""
        import tkinter as tk

        # Skip in headless environments
        try:
            root = tk.Tk()
            root.destroy()
        except tk.TclError:
            pytest.skip("No display available")

        from quotation.ui.demo_app import DemoApp
        app = DemoApp()
        app.update_idletasks()

        # Verify content area exists
        assert app._content_area is not None
        # Verify default page is the new-quotation page.
        assert app._content is not None
        # NewQuotePage should have toolbar
        assert hasattr(app._content, '_toolbar')
        assert app._content.use_ai is True
        app._content._use_ai.set(True)
        assert app._content.use_ai is True

        app.destroy()

    def test_management_navigation_pages_render(self):
        """History, published pricebook, and supplier pages are functional tables."""
        import tkinter as tk

        try:
            root = tk.Tk()
            root.destroy()
        except tk.TclError:
            pytest.skip("No display available")

        from quotation.ui.demo_app import DemoApp
        from quotation.ui.widgets import ManagementPage, SystemSettingsPage

        app = DemoApp()
        try:
            for page_name in ("报价记录", "价格管理", "供应商管理"):
                app._switch_page(page_name)
                app.update_idletasks()
                assert isinstance(app._content, ManagementPage)
                assert hasattr(app._content, "_tree")
            app._switch_page("系统设置")
            app.update_idletasks()
            assert isinstance(app._content, SystemSettingsPage)
            assert app._content._status_labels["转换器"].cget("text") in ("可用", "不可用")
        finally:
            app.destroy()

    def test_authenticated_engineer_sees_supplier_maintenance_actions(
        self, tmp_path, monkeypatch
    ):
        import tkinter as tk

        try:
            root = tk.Tk()
            root.destroy()
        except tk.TclError:
            pytest.skip("No display available")

        from quotation.application.auth_service import AuthService, SessionManager
        from quotation.application.settings_service import UserSettingsService
        from quotation.domain.user import UserRole
        from quotation.infrastructure.auth.encrypted_user_store import EncryptedUserStore
        from quotation.ui.demo_app import DemoApp

        auth = AuthService(
            EncryptedUserStore(tmp_path / "users.json", "test-user-store-key-2026")
        )
        admin = auth.create_initial_admin("admin001", "AdminPass123!", "管理员")
        engineer = auth.create_user(
            admin, "engineer01", "Engineer123!", "工程师", UserRole.ENGINEER
        )
        session = SessionManager().create_session(
            engineer, auth.get_user_permissions(engineer)
        )
        monkeypatch.setattr(
            UserSettingsService,
            "load",
            lambda _self: {
                "smb_root": str(tmp_path / "smb"),
                "smb_cache_dir": str(tmp_path / "cache"),
            },
        )
        try:
            app = DemoApp(session=session, auth_service=auth)
        except tk.TclError:
            pytest.skip("Tk theme runtime became unavailable")
        try:
            app._switch_page("供应商管理")
            app.update_idletasks()

            def button_texts(widget):
                result = []
                for child in widget.winfo_children():
                    if isinstance(child, tk.Button):
                        result.append(child.cget("text"))
                    result.extend(button_texts(child))
                return result

            labels = button_texts(app._content)
            assert "新增供应商" in labels
            assert "新增报价" in labels
            assert "导入报价Excel" in labels
        finally:
            app.destroy()

    def test_chinese_source_labels(self):
        """Source labels use Chinese display names, not internal codes."""
        from quotation.ui.viewmodels import SOURCE_LABELS, SOURCE_SHORT, STATUS_DISPLAY

        # Source labels should NOT contain raw codes
        assert "公司核准價格" in SOURCE_LABELS["C"]
        assert "歷史成交價格" in SOURCE_LABELS["H"]
        assert "系統估算價格" in SOURCE_LABELS["E"]
        assert "價格待確認" in SOURCE_LABELS["U"]
        assert "人工確認價格" in SOURCE_LABELS["M"]

        # Status should use Chinese
        assert STATUS_DISPLAY["COMPLETE"] == "報價完整"
        assert STATUS_DISPLAY["INCOMPLETE"] == "部分價格待確認"
        assert STATUS_DISPLAY["REVIEW_REQUIRED"] == "需要人工審核"
