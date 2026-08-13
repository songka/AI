"""Read-only administration queries and history re-export helpers."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from quotation.application.history_service import QuotationHistory
from quotation.infrastructure.smb.client import cached_public_path
from quotation.ui.localization import (
    display_origin_supplier,
    display_price_code,
    display_value,
)


class ManagementQueryService:
    """Expose published and supplier data without granting mutation access."""

    def __init__(
        self,
        pointer_path: str | Path | None = None,
        import_package_path: str | Path | None = None,
    ):
        local_pointer = Path("data/current-version-pointer.json")
        local_import = Path("rules/imports/r01-v1.0/pricing-rules-excel-r01-v1.0.json")
        self.pointer_path = (
            Path(pointer_path)
            if pointer_path
            else cached_public_path(
                "prices/published/current-version-pointer.json", local_pointer
            )
        )
        self.import_package_path = (
            Path(import_package_path)
            if import_package_path
            else cached_public_path(
                "prices/published/pricing-source-records-r01-v1.0.json", local_import
            )
        )

    def published_prices(
        self, target_type: str | None = None, query: str | None = None
    ) -> dict[str, Any]:
        pointer = self._read_json(self.pointer_path)
        snapshot_path = self.pointer_path.parent / pointer["snapshot_path"]
        snapshot = self._read_json(snapshot_path)
        if snapshot.get("status") != "PUBLISHED":
            raise ValueError("当前价格版本尚未发布")
        records = list(snapshot.get("company_prices", []))
        supplier_names = self._supplier_names()
        for record in records:
            record["canonical_code_display"] = display_price_code(
                record.get("canonical_code")
            )
            supplier_id = record.get("origin_supplier_id")
            record["origin_supplier_name"] = (
                supplier_names.get(str(supplier_id))
                if supplier_id
                else display_origin_supplier(None)
            ) or "供应商名称未维护"
        if target_type:
            records = [r for r in records if r.get("target_type") == target_type.upper()]
        if query:
            needle = query.casefold()
            records = [
                r for r in records
                if needle in str(r.get("canonical_code", "")).casefold()
                or needle in str(r.get("canonical_code_display", "")).casefold()
                or needle in str(r.get("specification", "")).casefold()
                or needle in str(r.get("origin_supplier_name", "")).casefold()
            ]
        return {
            "price_version_id": snapshot["price_version_id"],
            "snapshot_sha256": snapshot.get("snapshot_sha256"),
            "read_only": True,
            "total": len(records),
            "records": records,
        }

    def _supplier_names(self) -> dict[str, str]:
        try:
            package = self._read_json(self.import_package_path)
        except (FileNotFoundError, OSError, ValueError, json.JSONDecodeError):
            return {}
        return {
            str(record.get("supplier_id")): str(record.get("supplier_name"))
            for record in package.get("supplier_master", [])
            if record.get("supplier_id") and record.get("supplier_name")
        }

    def supplier_prices(
        self, supplier_id: str | None = None, query: str | None = None
    ) -> dict[str, Any]:
        package = self._read_json(self.import_package_path)
        records = list(package.get("pricing_source_records", []))
        for record in records:
            record["material_display"] = display_price_code(
                record.get("material_code")
            )
        if supplier_id:
            records = [r for r in records if r.get("supplier_id") == supplier_id]
        if query:
            needle = query.casefold()
            records = [
                r for r in records
                if any(
                    needle in str(r.get(field, "")).casefold()
                    for field in (
                        "record_id", "supplier_name", "material_code",
                        "material_display", "material_spec",
                    )
                )
            ]
        return {
            "read_only": True,
            "total": len(records),
            "suppliers": package.get("supplier_master", []),
            "records": records,
        }

    @staticmethod
    def _read_json(path: Path) -> dict[str, Any]:
        with path.open(encoding="utf-8") as source:
            return json.load(source)


def export_history_quote(
    history: QuotationHistory, quote_id: str, output_path: str | Path
) -> Path:
    """Re-export a persisted quote, including manual-review audit information."""

    detail = history.get_detail(quote_id)
    if detail is None:
        raise KeyError(quote_id)
    quote = detail["quote"]
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "报价摘要"
    summary.append(["字段", "值"])
    summary_fields = [
        ("报价编号", "quote_id"), ("图号", "drawing_number"),
        ("报价状态", "quotation_status"), ("报价完整度", "cost_completion"),
        ("未税小计", "subtotal_excl_tax"), ("税率", "tax_rate"),
        ("税额", "tax_amount"), ("含税总价", "total_incl_tax"),
        ("规则版本", "rule_version"), ("价格版本", "price_version"),
        ("报价版本", "quote_version"), ("更新时间", "updated_at"),
    ]
    for label, key in summary_fields:
        summary.append([label, display_value(key, quote.get(key))])

    items = wb.create_sheet("报价明细")
    item_fields = [
        ("费用行编号", "line_id"), ("费用类别", "category"), ("报价项目", "name"),
        ("价格来源", "source_display"), ("数量", "quantity"), ("单位", "unit"),
        ("单价", "unit_price"), ("未税金额", "amount"),
        ("定价依据", "resolution_display"), ("可信度", "confidence"), ("状态", "status"),
        ("智能辅助参考单价", "ai_estimated_unit_price"),
        ("智能辅助参考总额", "ai_estimated_amount"),
        ("智能辅助估价单位", "ai_estimated_unit"),
        ("智能辅助估价依据", "ai_estimate_reason"),
        ("智能辅助估价可信度", "ai_estimate_confidence"),
    ]
    items.append([label for label, _key in item_fields])
    for item in detail["items"]:
        items.append([display_value(key, item.get(key)) for _label, key in item_fields])

    overrides = wb.create_sheet("人工调整")
    overrides.append(["调整字段", "调整值", "更新时间"])
    for key, override in detail["overrides"].items():
        overrides.append([
            display_value("field_name", key),
            display_value("value", override.get("value")),
            display_value("updated_at", override.get("updated_at")),
        ])

    reviews = wb.create_sheet("人工审核记录")
    review_fields = [
        ("审核编号", "review_id"), ("调整字段", "field_name"),
        ("费用行编号", "line_id"), ("调整前", "old_value"), ("调整后", "new_value"),
        ("调整原因", "reason"), ("操作人", "operator"),
        ("调整前版本", "quote_version_before"), ("调整后版本", "quote_version_after"),
        ("时间", "created_at"),
    ]
    reviews.append([label for label, _key in review_fields])
    for review in detail["reviews"]:
        reviews.append([display_value(key, review.get(key)) for _label, key in review_fields])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(name="Microsoft YaHei UI", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A5276")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 3, 42)
            sheet.column_dimensions[column[0].column_letter].width = width
    wb.save(output)
    return output
