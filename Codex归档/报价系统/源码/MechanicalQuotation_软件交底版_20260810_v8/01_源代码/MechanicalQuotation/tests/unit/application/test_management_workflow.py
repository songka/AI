from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import openpyxl

from quotation.application.history_service import QuotationHistory
from quotation.application.management_service import (
    ManagementQueryService,
    export_history_quote,
)
from quotation.domain.quote import PriceSource, Quote, QuoteConfidence, QuoteItem


def _save_review_quote(history: QuotationHistory) -> str:
    item = QuoteItem(
        line_id="LINE-1",
        category="process",
        name="TAP",
        quantity=2,
        unit="hour",
        unit_price=0,
        amount=0,
        source=PriceSource.U,
        confidence=QuoteConfidence.UNCERTAIN,
        resolution_source="LEGACY_YAML_DRAFT",
        fallback_warning=True,
        ai_estimated_unit_price=80,
        ai_estimated_amount=160,
        ai_estimated_unit="小时",
        ai_estimate_reason="智能辅助参考，等待人工确认",
        ai_estimate_confidence=0.55,
    )
    quote = Quote(
        id="Q-1",
        drawing_id="D-1",
        part_number="J003",
        items=[item],
        quotation_status="INCOMPLETE",
        cost_completion=0,
        price_version="R01-COMPANY-PRICE-V1.1",
        rule_version="1.0",
    )
    result = SimpleNamespace(
        job_id="JOB-J003-1",
        drawing_number="J003",
        bundle=SimpleNamespace(
            geometry_source=SimpleNamespace(file_name="J003.dxf", full_path="J003.dxf")
        ),
        status="INCOMPLETE",
        cost_completion=0.0,
        unknown_item_count=1,
        subtotal_excluding_tax=Decimal("0"),
        tax=SimpleNamespace(tax_amount=Decimal("0")),
        total_including_tax=Decimal("0"),
        quote=quote,
        ai_used=False,
    )
    history.save_quote(result)
    return result.job_id


def test_manual_price_is_quote_scoped_versioned_and_audited(tmp_path):
    history = QuotationHistory(tmp_path / "history.db")
    quote_id = _save_review_quote(history)
    pricebook = Path("data/company-pricebook-r01-v1.1-snapshot.json")
    pricebook_before = pricebook.read_bytes()

    detail = history.apply_manual_review(
        quote_id,
        field_name="manual_price",
        line_id="LINE-1",
        new_value="12.5",
        reason="主管確認本次攻牙工時價",
        operator="reviewer-a",
    )

    item = detail["items"][0]
    assert item["source"] == "M"
    assert item["source_display"] == "人工確認價格"
    assert item["unit_price"] == 12.5
    assert item["amount"] == 25.0
    assert detail["quote"]["quote_version"] == 2
    assert detail["quote"]["quotation_status"] == "COMPLETE"
    assert detail["reviews"][0]["old_value"] == "0.0"
    assert detail["reviews"][0]["new_value"] == "12.5"
    assert detail["reviews"][0]["quote_version_before"] == 1
    assert detail["reviews"][0]["quote_version_after"] == 2
    assert pricebook.read_bytes() == pricebook_before


def test_quote_history_records_operator_pc_identity_and_can_delete(tmp_path):
    history = QuotationHistory(tmp_path / "history.db")
    quote_id = _save_review_quote(history)
    result = SimpleNamespace(
        job_id="JOB-AUDIT-1",
        drawing_number="AUDIT-1",
        bundle=SimpleNamespace(
            geometry_source=SimpleNamespace(file_name="audit.dwg", full_path="audit.dwg")
        ),
        status="COMPLETE",
        cost_completion=100.0,
        unknown_item_count=0,
        subtotal_excluding_tax=Decimal("10"),
        tax=SimpleNamespace(tax_rate=Decimal("0.13"), tax_amount=Decimal("1.3")),
        total_including_tax=Decimal("11.3"),
        quote=Quote(id="Q-AUDIT", drawing_id="D-AUDIT", items=[]),
        ai_used=False,
    )
    history.save_quote(
        result,
        quoted_by="报价工程师",
        pc_identity={
            "pc_username": "windows-user",
            "pc_name": "PC-001",
            "pc_ip": "10.97.0.88",
        },
    )

    saved = history.get_quote("JOB-AUDIT-1")
    assert saved["quoted_by"] == "报价工程师"
    assert saved["pc_username"] == "windows-user"
    assert saved["pc_name"] == "PC-001"
    assert saved["pc_ip"] == "10.97.0.88"
    assert history.delete_quote(quote_id) is True
    assert history.get_detail(quote_id) is None
    assert history.delete_quote(quote_id) is False


def test_feature_override_and_history_reexport(tmp_path):
    history = QuotationHistory(tmp_path / "history.db")
    quote_id = _save_review_quote(history)
    history.apply_manual_review(
        quote_id,
        field_name="thickness",
        new_value="1.5 mm",
        reason="依 PDF 標註補充",
        operator="reviewer-b",
    )

    output = export_history_quote(history, quote_id, tmp_path / "reviewed.xlsx")
    wb = openpyxl.load_workbook(output, read_only=True, data_only=True)

    assert set(wb.sheetnames) == {"报价摘要", "报价明细", "人工调整", "人工审核记录"}
    summary_rows = list(wb["报价摘要"].iter_rows(values_only=True))
    assert ("报价状态", "部分价格待确认") in summary_rows
    assert not any(cell == "quotation_status" for row in summary_rows for cell in row)
    detail_headers = next(wb["报价明细"].iter_rows(values_only=True))
    assert "费用类别" in detail_headers
    assert "价格来源" in detail_headers
    assert "智能辅助参考总额" in detail_headers
    assert history.get_detail(quote_id)["items"][0]["ai_estimated_amount"] == 160
    assert history.get_detail(quote_id)["overrides"]["thickness"]["value"] == "1.5 mm"
    wb.close()


def test_management_queries_are_read_only_and_filterable():
    service = ManagementQueryService()

    published = service.published_prices(target_type="SURFACE", query="RAL9003")
    suppliers = service.supplier_prices(supplier_id="SUP-TONGRUI", query="S50C")

    assert published["read_only"] is True
    assert published["price_version_id"] == "R01-COMPANY-PRICE-V1.1"
    assert published["records"][0]["unit_price"] == 25.0
    assert suppliers["read_only"] is True
    assert all(r["supplier_id"] == "SUP-TONGRUI" for r in suppliers["records"])


def test_price_management_uses_chinese_names_instead_of_codes():
    service = ManagementQueryService()

    aluminum = service.published_prices(target_type="MATERIAL", query="铝")

    assert aluminum["records"]
    assert any("铝" in row["canonical_code_display"] for row in aluminum["records"])
    assert all(not row["origin_supplier_name"].startswith("SUP-") for row in aluminum["records"])
    supplier_names = {row["origin_supplier_name"] for row in aluminum["records"]}
    assert supplier_names.intersection({"良伟", "稳迪", "捷密达", "公司内部核准价"})


def test_pending_supplier_records_are_query_only_not_formal_prices():
    service = ManagementQueryService()
    supplier_records = service.supplier_prices()["records"]
    published_records = service.published_prices()["records"]

    assert any(r.get("status") != "PUBLISHED" for r in supplier_records)
    assert all(r.get("origin_type") != "PENDING_SUPPLIER" for r in published_records)
