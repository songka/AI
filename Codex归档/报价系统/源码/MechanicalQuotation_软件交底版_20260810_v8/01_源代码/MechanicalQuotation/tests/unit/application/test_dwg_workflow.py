from __future__ import annotations

from pathlib import Path

import ezdxf
import openpyxl

from quotation.application.batch_excel import export_batch_excel
from quotation.application.file_scanner import DrawingFile, JobBundle, MatchStatus
from quotation.application.quotation_service import JobStatus, QuotationApplicationService
from quotation.infrastructure.dwg.converter import (
    ConversionStatus,
    ConverterHealth,
    DwgConversionResult,
    DwgConversionService,
    DwgConverterAdapter,
)


class PipelineConverter(DwgConverterAdapter):
    def __init__(self, fail_names: set[str] | None = None):
        self.fail_names = fail_names or set()

    @property
    def identity(self):
        return "pipeline-fake-v1"

    def health(self):
        return ConverterHealth(True, True, adapter="測試轉換器", configuration_source="test")

    def convert(self, source_path, output_path, cancellation_check=None):
        if source_path.name in self.fail_names:
            return DwgConversionResult(
                str(source_path),
                ConversionStatus.FAILED,
                adapter="測試轉換器",
                configuration_source="test",
                error="轉換失敗",
            )
        doc = ezdxf.new()
        msp = doc.modelspace()
        for start, end in [
            ((0, 0), (100, 0)),
            ((100, 0), (100, 50)),
            ((100, 50), (0, 50)),
            ((0, 50), (0, 0)),
        ]:
            msp.add_line(start, end)
        msp.add_text("S50C", height=5).set_placement((0, 55))
        doc.saveas(output_path)
        return DwgConversionResult(
            str(source_path),
            ConversionStatus.SUCCESS,
            str(output_path),
            adapter="測試轉換器",
            configuration_source="test",
        )


class FailingIfCalledSolidWorksService:
    def convert(self, *args, **kwargs):
        raise AssertionError("选择 DWG 时不得调用 SolidWorks 转换")


def _bundle(path: Path) -> JobBundle:
    drawing_file = DrawingFile.from_path(path)
    assert drawing_file is not None
    return JobBundle(drawing_file.drawing_number, [drawing_file], MatchStatus.UNMATCHED)


def test_dwg_flows_through_existing_dxf_pipeline(tmp_path):
    dwg = tmp_path / "中文 DWG.dwg"
    dwg.write_bytes(b"AC1027-pipeline")
    service = QuotationApplicationService(
        dwg_conversion_service=DwgConversionService(
            PipelineConverter(), tmp_path / "conversion-cache"
        )
    )
    result = service.quote_single_file(dwg)
    assert result.status in (JobStatus.COMPLETE, JobStatus.REVIEW_REQUIRED)
    assert result.quote is not None
    assert result.dwg_conversion["status"] == ConversionStatus.SUCCESS
    assert result.dwg_conversion["original_preserved"] is True
    assert result.dwg_conversion["converted_file_deleted"] is True
    assert result.dwg_conversion["converted_file_retained"] is False
    assert result.dwg_conversion["cleanup_status"] == "DELETED"
    assert dwg.is_file()
    assert not Path(result.dwg_conversion["converted_file"]).exists()


def test_selected_dwg_does_not_use_same_name_solidworks_file(tmp_path):
    dwg = tmp_path / "SELECTED.dwg"
    solidworks = tmp_path / "SELECTED.sldprt"
    dwg.write_bytes(b"AC1027-selected")
    solidworks.write_bytes(b"solidworks-sidecar")
    service = QuotationApplicationService(
        dwg_conversion_service=DwgConversionService(
            PipelineConverter(), tmp_path / "conversion-cache"
        ),
        solidworks_conversion_service=FailingIfCalledSolidWorksService(),
    )

    result = service.quote_single_file(dwg)

    assert result.status in (JobStatus.COMPLETE, JobStatus.REVIEW_REQUIRED)
    assert result.bundle.geometry_source is not None
    assert result.bundle.geometry_source.full_path == dwg.resolve()
    assert result.dwg_conversion["status"] == ConversionStatus.SUCCESS


def test_one_dwg_failure_does_not_abort_batch(tmp_path):
    bad = tmp_path / "BAD.dwg"
    good = tmp_path / "GOOD.dwg"
    bad.write_bytes(b"AC1027-bad")
    good.write_bytes(b"AC1027-good")
    service = QuotationApplicationService(
        dwg_conversion_service=DwgConversionService(
            PipelineConverter({bad.name}), tmp_path / "conversion-cache"
        )
    )
    results = service.quote_batch([_bundle(bad), _bundle(good)])
    assert len(results) == 2
    assert results[0].status == JobStatus.DWG_CONVERSION_FAILED
    assert results[1].quote is not None


def test_excel_contains_dwg_conversion_trace(tmp_path):
    dwg = tmp_path / "TRACE.dwg"
    dwg.write_bytes(b"AC1027-trace")
    service = QuotationApplicationService(
        dwg_conversion_service=DwgConversionService(
            PipelineConverter(), tmp_path / "conversion-cache"
        )
    )
    result = service.quote_single_file(dwg)
    output = export_batch_excel([result], tmp_path / "trace.xlsx")
    workbook = openpyxl.load_workbook(output, read_only=True)
    assert "DWG转换记录" in workbook.sheetnames
    rows = list(workbook["DWG转换记录"].iter_rows(values_only=True))
    flat = [value for row in rows for value in row]
    assert "DWG转换追踪" in flat
    assert "成功" in flat
    assert "測試轉換器" in flat
