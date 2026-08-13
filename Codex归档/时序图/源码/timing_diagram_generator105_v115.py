# -*- coding: utf-8 -*-
"""
时序图自动生成工具

新版思路：
1. 用户在软件里按“动作”填写流程。
2. 软件根据重复次数和触发规则展开成 Excel 明细。
3. 绘图和导出都使用同一份展开后的时序数据。
"""

from __future__ import annotations

import datetime as _dt
import math
import traceback
from collections import defaultdict, deque
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import tkinter as tk
import tkinter.font as tkfont
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox, ttk

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

    HAS_MATPLOTLIB = True
except ModuleNotFoundError:
    plt = None
    FigureCanvasTkAgg = None
    HAS_MATPLOTLIB = False


DETAIL_HEADERS = [
    "动作主体",
    "步骤",
    "轮次",
    "动作",
    "启动步",
    "前步骤",
    "时间",
    "次数",
    "来源动作",
    "依赖说明",
    "开始时间",
    "结束时间",
]

FLOW_HEADERS = [
    "动作编号",
    "动作主体",
    "动作",
    "时间",
    "重复次数",
    "第一轮等待动作编号",
    "后续轮等待动作编号",
    "触发方式",
    "触发次数/偏移",
    "备注",
]

TRIGGER_MODES = ("同次完成", "固定次数完成", "上一次完成")
WAIT_NONE_LABEL = "不等待其它动作主体"
TRIGGER_VALUE_LABELS = (
    "第一次完成",
    "第二次完成",
    "第三次完成",
    "第四次完成",
    "第五次完成",
    "第六次完成",
    "第七次完成",
    "第八次完成",
    "第九次完成",
    "第十次完成",
)
TRIGGER_VALUE_BY_LABEL = {label: index + 1 for index, label in enumerate(TRIGGER_VALUE_LABELS)}
COUNT_LABELS = ("一次", "两次", "三次", "四次", "五次", "六次", "七次", "八次", "九次", "十次")
COUNT_BY_LABEL = {label: index + 1 for index, label in enumerate(COUNT_LABELS)}
CYCLE_LABELS = ("1轮", "2轮", "3轮", "4轮", "5轮", "6轮", "7轮", "8轮", "9轮", "10轮")
CYCLE_BY_LABEL = {label: index + 1 for index, label in enumerate(CYCLE_LABELS)}
VISIBLE_DEP_KINDS = {"first_wait", "later_wait"}

HEADER_ALIASES = {
    "station": ("动作主体", "加工位", "工位"),
    "module": ("模块", "模組"),
    "step": ("步骤", "步驟", "step", "Step"),
    "action": ("动作", "動作"),
    "start": ("启动步", "啟動步"),
    "prev": ("前步骤", "前步驟", "前置步骤", "前置步驟"),
    "time": ("时间", "時間"),
}


@dataclass
class FlowAction:
    action_id: int
    station: str = ""
    module: str = ""
    action: str = ""
    duration: float = 0.1
    repeat: int = 1
    depends_on: str = ""
    later_depends_on: str = ""
    trigger_mode: str = "同次完成"
    trigger_value: int = 1
    note: str = ""


def normalize_number_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    for unit in ("秒", "sec", "Sec", "SEC", "s", "S", "次", "轮", "遍", "个"):
        if text.endswith(unit):
            text = text[: -len(unit)].strip()
            break
    return text


def to_int(value, default: Optional[int] = None) -> Optional[int]:
    text = normalize_number_text(value)
    if text == "":
        return default
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return default


def to_float(value, default: Optional[float] = None) -> Optional[float]:
    text = normalize_number_text(value)
    if text == "":
        return default
    try:
        number = float(text)
    except (TypeError, ValueError):
        return default
    if math.isnan(number):
        return default
    return number


def trigger_value_to_int(value, default: int = 1) -> int:
    if value in TRIGGER_VALUE_BY_LABEL:
        return TRIGGER_VALUE_BY_LABEL[value]
    return to_int(value, default) or default


def trigger_value_to_label(value: int) -> str:
    if 1 <= int(value or 1) <= len(TRIGGER_VALUE_LABELS):
        return TRIGGER_VALUE_LABELS[int(value or 1) - 1]
    return TRIGGER_VALUE_LABELS[0]


def count_to_int(value, default: int = 1) -> int:
    if value in COUNT_BY_LABEL:
        return COUNT_BY_LABEL[value]
    return to_int(value, default) or default


def count_to_label(value: int) -> str:
    if 1 <= int(value or 1) <= len(COUNT_LABELS):
        return COUNT_LABELS[int(value or 1) - 1]
    return COUNT_LABELS[0]


def cycle_to_int(value, default: int = 1) -> int:
    if value in CYCLE_BY_LABEL:
        return CYCLE_BY_LABEL[value]
    return to_int(value, default) or default


def split_ids(text) -> List[int]:
    if text is None:
        return []
    items = []
    for raw in str(text).replace("，", ",").replace(";", ",").split(","):
        value = to_int(raw)
        if value is not None:
            items.append(value)
    return items


def join_ids(values: Iterable[int]) -> str:
    return ",".join(str(v) for v in values)


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def header_map(ws) -> Dict[str, int]:
    result: Dict[str, int] = {}
    first_row = [clean_text(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in first_row:
                result[key] = first_row.index(alias) + 1
                break
    return result


def resolve_dependency_occurrence(
    current_occurrence: int,
    dependency_repeat: int,
    trigger_mode: str,
    trigger_value: int,
) -> Optional[int]:
    if dependency_repeat < 1:
        return None
    trigger_value = max(1, int(trigger_value or 1))
    if trigger_mode == "固定次数完成":
        return min(trigger_value, dependency_repeat)
    if trigger_mode == "上一次完成":
        occurrence = current_occurrence - trigger_value
        return occurrence if occurrence >= 1 else None
    return min(current_occurrence, dependency_repeat)


def add_event_dependency(event: dict, dep_event: dict, note: str, kind: str) -> None:
    dep_step = dep_event["step"]
    if dep_step not in event["deps"]:
        event["deps"].append(dep_step)
    if note:
        event["dep_notes"].append(note)
    edge = {"step": dep_step, "kind": kind}
    if edge not in event.setdefault("dep_edges", []):
        event["dep_edges"].append(edge)


def build_events_from_actions(actions: List[FlowAction], cycle_count: int = 1) -> List[dict]:
    if not actions:
        raise ValueError("请先添加至少一个动作。")
    cycle_count = max(1, int(cycle_count or 1))

    ids = [a.action_id for a in actions]
    if len(ids) != len(set(ids)):
        raise ValueError("动作编号不能重复。")

    action_by_id = {a.action_id: a for a in actions}
    previous_action_for_station: Dict[int, FlowAction] = {}
    last_action_by_station: Dict[str, FlowAction] = {}
    station_last_action: Dict[str, FlowAction] = {}
    events: List[dict] = []
    event_by_key: Dict[Tuple[int, int, int], dict] = {}
    step = 1

    for cycle in range(1, cycle_count + 1):
        for action in actions:
            if action.duration <= 0:
                raise ValueError(f"动作 {action.action_id} 的时间必须大于 0。")
            if action.repeat < 1:
                raise ValueError(f"动作 {action.action_id} 的重复次数必须大于 0。")
            station_key = action.station.strip()
            if cycle == 1:
                if station_key and station_key in last_action_by_station:
                    previous_action_for_station[action.action_id] = last_action_by_station[station_key]
                if station_key:
                    last_action_by_station[station_key] = action
                    station_last_action[station_key] = action
            label = action.action or f"动作 {action.action_id}"
            for occurrence in range(1, action.repeat + 1):
                event = {
                    "step": step,
                    "cycle": cycle,
                    "station": action.station,
                    "module": action.module,
                    "action": label,
                    "duration": round(float(action.duration), 3),
                    "occurrence": occurrence,
                    "source_action": action.action_id,
                    "deps": [],
                    "dep_edges": [],
                    "dep_notes": [],
                    "start": 0.0,
                    "end": 0.0,
                }
                events.append(event)
                event_by_key[(action.action_id, cycle, occurrence)] = event
                step += 1

    for cycle in range(1, cycle_count + 1):
        for action in actions:
            dep_ids = split_ids(action.depends_on if cycle == 1 else action.later_depends_on)
            station_prev = previous_action_for_station.get(action.action_id)
            for occurrence in range(1, action.repeat + 1):
                event = event_by_key[(action.action_id, cycle, occurrence)]
                if occurrence > 1:
                    prev_event = event_by_key[(action.action_id, cycle, occurrence - 1)]
                    add_event_dependency(event, prev_event, f"本动作第 {occurrence - 1} 次完成", "self_repeat")
                elif station_prev is not None:
                    prev_event = event_by_key[(station_prev.action_id, cycle, station_prev.repeat)]
                    add_event_dependency(event, prev_event, "同动作主体上一动作完成", "same_subject")
                elif cycle > 1:
                    station_last = station_last_action.get(action.station.strip())
                    if station_last is not None:
                        prev_event = event_by_key[(station_last.action_id, cycle - 1, station_last.repeat)]
                        add_event_dependency(event, prev_event, "同动作主体上一轮最后动作完成", "same_subject_cycle")
                    else:
                        prev_cycle_event = event_by_key[(action.action_id, cycle - 1, action.repeat)]
                        add_event_dependency(event, prev_cycle_event, "本动作上一轮完成", "self_cycle")

                for dep_id in dep_ids:
                    if dep_id not in action_by_id:
                        raise ValueError(f"动作 {action.action_id} 依赖了不存在的动作 {dep_id}。")
                    dep_action = action_by_id[dep_id]
                    dep_occurrence = resolve_dependency_occurrence(
                        occurrence,
                        dep_action.repeat,
                        action.trigger_mode,
                        action.trigger_value,
                    )
                    if dep_occurrence is None:
                        continue
                    dep_cycle = cycle - 1 if cycle > 1 and action.later_depends_on else cycle
                    dep_event = event_by_key[(dep_id, dep_cycle, dep_occurrence)]
                    cycle_note = "上一轮" if dep_cycle != cycle else "本轮"
                    edge_kind = "later_wait" if cycle > 1 and action.later_depends_on else "first_wait"
                    add_event_dependency(
                        event,
                        dep_event,
                        f"{cycle_note}动作 {dep_id} 第 {dep_occurrence} 次完成",
                        edge_kind,
                    )

    schedule_events(events)
    return events


def schedule_events(events: List[dict]) -> None:
    step_map = {event["step"]: event for event in events}
    graph: Dict[int, List[int]] = defaultdict(list)
    indegree: Dict[int, int] = {}

    for event in events:
        valid_deps = []
        for dep in event.get("deps", []):
            if dep in step_map and dep != event["step"]:
                valid_deps.append(dep)
        event["deps"] = sorted(set(valid_deps))
        valid_dep_set = set(event["deps"])
        event["dep_edges"] = [
            edge
            for edge in event.get("dep_edges", [])
            if edge.get("step") in valid_dep_set and edge.get("step") != event["step"]
        ]
        indegree[event["step"]] = len(event["deps"])
        for dep in event["deps"]:
            graph[dep].append(event["step"])

    queue = deque(sorted(step for step, count in indegree.items() if count == 0))
    processed = 0

    while queue:
        step = queue.popleft()
        event = step_map[step]
        event["start"] = round(max((step_map[d]["end"] for d in event["deps"]), default=0.0), 3)
        event["end"] = round(event["start"] + float(event["duration"]), 3)
        processed += 1

        for next_step in sorted(graph[step]):
            indegree[next_step] -= 1
            if indegree[next_step] == 0:
                queue.append(next_step)

    if processed != len(events):
        raise ValueError("等待条件存在循环，请检查前后关系。")


def load_actions_from_sheet(ws) -> List[FlowAction]:
    headers = [clean_text(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    positions = {name: headers.index(name) + 1 for name in headers if name}
    if "动作编号" not in positions:
        raise ValueError("流程输入表缺少“动作编号”列。")

    actions: List[FlowAction] = []
    for row in range(2, ws.max_row + 1):
        action_id = to_int(ws.cell(row=row, column=positions["动作编号"]).value)
        if action_id is None:
            continue
        action = FlowAction(
            action_id=action_id,
            station=clean_text(ws.cell(row=row, column=positions.get("动作主体", positions.get("加工位", 1))).value),
            module=clean_text(ws.cell(row=row, column=positions.get("模块", 1)).value),
            action=clean_text(ws.cell(row=row, column=positions.get("动作", 1)).value),
            duration=to_float(ws.cell(row=row, column=positions.get("时间", 1)).value, 0.1) or 0.1,
            repeat=count_to_int(ws.cell(row=row, column=positions.get("重复次数", 1)).value, 1),
            depends_on=clean_text(
                ws.cell(
                    row=row,
                    column=positions.get("第一轮等待动作编号", positions.get("依赖动作编号", 1)),
                ).value
            ),
            later_depends_on=clean_text(ws.cell(row=row, column=positions.get("后续轮等待动作编号", 1)).value),
            trigger_mode=clean_text(ws.cell(row=row, column=positions.get("触发方式", 1)).value)
            or "同次完成",
            trigger_value=trigger_value_to_int(ws.cell(row=row, column=positions.get("触发次数/偏移", 1)).value, 1),
            note=clean_text(ws.cell(row=row, column=positions.get("备注", 1)).value),
        )
        if action.trigger_mode not in TRIGGER_MODES:
            action.trigger_mode = "同次完成"
        actions.append(action)
    return actions


def load_events_from_detail_sheet(ws) -> List[dict]:
    positions = header_map(ws)
    required = ("step", "action", "time")
    missing = [name for name in required if name not in positions]
    if missing:
        raise ValueError("Excel 缺少必要列：步骤、动作、时间。")

    events: List[dict] = []
    for row in range(2, ws.max_row + 1):
        step = to_int(ws.cell(row=row, column=positions["step"]).value)
        if step is None:
            continue
        start_flag = to_int(ws.cell(row=row, column=positions.get("start", 1)).value, 0) or 0
        deps = [] if start_flag == 1 else split_ids(ws.cell(row=row, column=positions.get("prev", 1)).value)
        events.append(
            {
                "step": step,
                "station": clean_text(ws.cell(row=row, column=positions.get("station", 1)).value),
                "module": clean_text(ws.cell(row=row, column=positions.get("module", 1)).value),
                "action": clean_text(ws.cell(row=row, column=positions["action"]).value),
                "duration": to_float(ws.cell(row=row, column=positions["time"]).value, 0.1) or 0.1,
                "occurrence": 1,
                "source_action": step,
                "deps": deps,
                "dep_edges": [],
                "dep_notes": ["启动步"] if start_flag == 1 else [],
                "start": 0.0,
                "end": 0.0,
            }
        )

    if not events:
        raise ValueError("Excel 中没有可用的步骤数据。")
    schedule_events(events)
    return events


def load_workbook_data(path: str, cycle_count: int = 1) -> Tuple[List[FlowAction], List[dict]]:
    wb = load_workbook(path, data_only=True)
    if "流程输入" in wb.sheetnames:
        actions = load_actions_from_sheet(wb["流程输入"])
        return actions, build_events_from_actions(actions, cycle_count)

    first_sheet = wb[wb.sheetnames[0]]
    first_headers = [clean_text(first_sheet.cell(row=1, column=c).value) for c in range(1, first_sheet.max_column + 1)]
    if "动作编号" in first_headers:
        actions = load_actions_from_sheet(first_sheet)
        return actions, build_events_from_actions(actions, cycle_count)

    sheet_name = "时序明细" if "时序明细" in wb.sheetnames else wb.sheetnames[0]
    events = load_events_from_detail_sheet(wb[sheet_name])
    return [], events


def style_header(ws, max_col: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="808080"))
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def autofit(ws, widths: Dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def calculate_cycle_summary(events: List[dict]) -> dict:
    cycles = sorted({event.get("cycle", 1) for event in events})
    cycle_spans = []
    for cycle in cycles:
        cycle_events = [event for event in events if event.get("cycle", 1) == cycle]
        if not cycle_events:
            continue
        start = round(min(event["start"] for event in cycle_events), 3)
        end = round(max(event["end"] for event in cycle_events), 3)
        cycle_spans.append({"cycle": cycle, "start": start, "end": end, "duration": round(end - start, 3)})

    ends = [item["end"] for item in cycle_spans]
    intervals = [round(ends[index] - ends[index - 1], 3) for index in range(1, len(ends))]
    stable_period = None
    stable_start_cycle = None
    stable_note = "轮数不足，暂不能判断稳定周期"
    if intervals:
        last_interval = intervals[-1]
        stable_from = len(intervals) - 1
        while stable_from > 0 and abs(intervals[stable_from - 1] - last_interval) <= 0.001:
            stable_from -= 1
        stable_period = last_interval
        stable_start_cycle = cycle_spans[min(stable_from + 1, len(cycle_spans) - 1)]["cycle"]
        stable_note = (
            f"参考周期 {last_interval:g}s（只有 2 轮，建议增加轮数确认）"
            if len(intervals) == 1
            else f"稳定周期 {last_interval:g}s（第 {stable_start_cycle} 轮完成起）"
        )

    total_start = min((event["start"] for event in events), default=0)
    total_end = max((event["end"] for event in events), default=0)
    return {
        "total_time": round(total_end - total_start, 3),
        "cycle_spans": cycle_spans,
        "intervals": intervals,
        "stable_period": stable_period,
        "stable_start_cycle": stable_start_cycle,
        "stable_note": stable_note,
    }


def cycle_summary_text(events: List[dict]) -> str:
    summary = calculate_cycle_summary(events)
    return f"总时长 {summary['total_time']:g}s，{summary['stable_note']}"


def visible_dependency_edges(event: dict) -> List[dict]:
    return [edge for edge in event.get("dep_edges", []) if edge.get("kind") in VISIBLE_DEP_KINDS]


def describe_dependency_ids(depends_on: str, action_by_id: Dict[int, FlowAction]) -> str:
    labels = []
    for dep_id in split_ids(depends_on):
        action = action_by_id.get(dep_id)
        if action is None:
            labels.append(f"动作 {dep_id}")
            continue
        station = action.station.strip() or "未设置动作主体"
        name = action.action.strip() or f"动作 {dep_id}"
        labels.append(f"{station} / {name}")
    return "；".join(labels)


def write_workbook(path: str, actions: List[FlowAction], events: List[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "时序图"
    write_timeline_sheet(ws, actions, events)
    wb.save(path)


def write_timeline_sheet(ws, actions: List[FlowAction], events: List[dict]) -> None:
    summary = calculate_cycle_summary(events)
    readable_wait_headers = ["第一轮等待说明", "后续轮等待说明"]
    summary_headers = ["总时长", "稳定周期", "稳定开始轮", "周期说明", "各轮起点", "各轮用时"]
    headers = FLOW_HEADERS[:7] + readable_wait_headers + FLOW_HEADERS[7:] + summary_headers
    for idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=idx).value = header

    max_time = max((event["end"] for event in events), default=0)
    time_slots = max(1, int(math.ceil(max_time * 10)))
    for slot in range(time_slots):
        ws.cell(row=1, column=len(headers) + 1 + slot).value = round((slot + 1) / 10, 1)

    row_by_action = {}
    action_by_id = {action.action_id: action for action in actions}
    colors = ["F4B183", "9DC3E6", "A9D18E", "FFD966", "C9C9FF", "F8CBAD", "B4C6E7", "C6E0B4"]
    cycle_starts = "；".join(f"第{item['cycle']}轮 {item['start']:g}s" for item in summary["cycle_spans"])
    cycle_durations = "；".join(f"第{item['cycle']}轮 {item['duration']:g}s" for item in summary["cycle_spans"])

    for row, action in enumerate(actions, 2):
        row_by_action[action.action_id] = row
        values = [
            action.action_id,
            action.station,
            action.action,
            action.duration,
            action.repeat,
            action.depends_on,
            action.later_depends_on,
            describe_dependency_ids(action.depends_on, action_by_id),
            describe_dependency_ids(action.later_depends_on, action_by_id),
            action.trigger_mode if action.depends_on or action.later_depends_on else "",
            action.trigger_value if action.depends_on or action.later_depends_on else "",
            action.note,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col).value = value
        if row == 2:
            summary_values = [
                summary["total_time"],
                summary["stable_period"],
                summary["stable_start_cycle"],
                summary["stable_note"],
                cycle_starts,
                cycle_durations,
            ]
            for offset, value in enumerate(summary_values, len(FLOW_HEADERS) + len(readable_wait_headers) + 1):
                ws.cell(row=row, column=offset).value = value

    for event in events:
        row = row_by_action.get(event.get("source_action"))
        if row is None:
            continue
        fill = PatternFill("solid", fgColor=colors[(event.get("source_action", event["step"]) - 1) % len(colors)])
        start_col = len(headers) + 1 + int(round(event["start"] * 10))
        end_col = len(headers) + int(math.ceil(event["end"] * 10))
        for col in range(start_col, max(start_col, end_col) + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = 0.1
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    style_header(ws, len(headers) + time_slots)
    autofit(
        ws,
        {
            1: 10,
            2: 16,
            3: 36,
            4: 10,
            5: 10,
            6: 18,
            7: 20,
            8: 22,
            9: 22,
            10: 16,
            11: 14,
            12: 24,
            13: 10,
            14: 10,
            15: 12,
            16: 30,
            17: 28,
            18: 28,
        },
    )
    for col in range(len(headers) + 1, len(headers) + time_slots + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4
    ws.freeze_panes = "A2"


class TimingDiagramApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("时序图自动生成工具")
        self.root.geometry("1280x760")
        self.root.minsize(1100, 680)
        self.configure_fonts()
        self.actions: List[FlowAction] = []
        self.events: List[dict] = []
        self.current_file = ""
        self.cycle_var = tk.StringVar(value=CYCLE_LABELS[0])

        if HAS_MATPLOTLIB:
            plt.rcParams["font.family"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS"]
            plt.rcParams["axes.unicode_minus"] = False

        self._build_ui()
        self.load_example()

    def configure_fonts(self) -> None:
        for font_name in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont"):
            try:
                tkfont.nametofont(font_name).configure(family="Microsoft YaHei", size=10)
            except tk.TclError:
                pass

    def current_cycle_count(self) -> int:
        raw_value = self.cycle_var.get().strip()
        cycle_count = to_int(raw_value)
        if cycle_count is None or cycle_count < 1:
            raise ValueError("绘制轮数必须是大于 0 的整数，例如 3。")
        return cycle_count

    def _build_ui(self) -> None:
        toolbar = ttk.Frame(self.root, padding=(8, 8, 8, 4))
        toolbar.pack(side=tk.TOP, fill=tk.X)

        ttk.Button(toolbar, text="导入 Excel", command=self.import_excel).pack(side=tk.LEFT, padx=3)
        ttk.Button(toolbar, text="生成时序图", command=self.generate_diagram).pack(side=tk.LEFT, padx=3)
        ttk.Button(toolbar, text="导出 Excel", command=self.export_excel).pack(side=tk.LEFT, padx=3)
        ttk.Button(toolbar, text="导出图片", command=self.export_image).pack(side=tk.LEFT, padx=3)
        ttk.Button(toolbar, text="导出模板", command=self.export_template).pack(side=tk.LEFT, padx=3)
        ttk.Label(toolbar, text="绘制轮数").pack(side=tk.LEFT, padx=(14, 3))
        ttk.Entry(toolbar, textvariable=self.cycle_var, width=6).pack(side=tk.LEFT)
        self.status_var = tk.StringVar(value="请填写动作，或导入已有 Excel。")
        ttk.Label(toolbar, textvariable=self.status_var).pack(side=tk.LEFT, padx=14)

        paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        left = ttk.Frame(paned, padding=6)
        right = ttk.Frame(paned, padding=6)
        paned.add(left, weight=0)
        paned.add(right, weight=1)

        form = ttk.LabelFrame(left, text="动作设置")
        form.pack(fill=tk.X)
        self._build_form(form)

        list_frame = ttk.LabelFrame(left, text="动作清单")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 0))
        self._build_tree(list_frame)

        preview = ttk.LabelFrame(right, text="时序图预览")
        preview.pack(fill=tk.BOTH, expand=True)
        if HAS_MATPLOTLIB:
            self.fig, self.ax = plt.subplots(figsize=(9, 6))
            self.canvas = FigureCanvasTkAgg(self.fig, master=preview)
            self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        else:
            self.fig = None
            self.ax = None
            canvas_frame = ttk.Frame(preview)
            canvas_frame.pack(fill=tk.BOTH, expand=True)
            self.canvas = tk.Canvas(canvas_frame, bg="white", highlightthickness=0)
            yscroll = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
            xscroll = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
            self.canvas.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
            self.canvas.grid(row=0, column=0, sticky=tk.NSEW)
            yscroll.grid(row=0, column=1, sticky=tk.NS)
            xscroll.grid(row=1, column=0, sticky=tk.EW)
            canvas_frame.rowconfigure(0, weight=1)
            canvas_frame.columnconfigure(0, weight=1)

    def _build_form(self, parent: ttk.Frame) -> None:
        self.vars = {
            "action_id": tk.StringVar(),
            "station": tk.StringVar(),
            "module": tk.StringVar(),
            "duration": tk.StringVar(value="0.1"),
            "repeat": tk.StringVar(value="1"),
            "depends_on": tk.StringVar(),
            "trigger_mode": tk.StringVar(value="同次完成"),
            "trigger_value": tk.StringVar(value=TRIGGER_VALUE_LABELS[0]),
            "note": tk.StringVar(),
        }
        self.dependency_var = tk.StringVar(value=WAIT_NONE_LABEL)
        self.dep_label_to_id: Dict[str, int] = {}
        self.selected_dep_ids: List[int] = []
        self.selected_later_dep_ids: List[int] = []

        labels = [
            ("动作主体", "station"),
            ("时间(s)", "duration"),
            ("重复次数", "repeat"),
            ("等待其它动作主体动作", "depends_on"),
            ("触发方式", "trigger_mode"),
            ("等待到", "trigger_value"),
        ]
        self.form_labels = {}
        self.form_widgets = {}
        for row, (label, key) in enumerate(labels):
            label_widget = ttk.Label(parent, text=label)
            label_widget.grid(row=row, column=0, sticky=tk.W, padx=4, pady=3)
            if key == "trigger_mode":
                widget = ttk.Combobox(parent, textvariable=self.vars[key], values=TRIGGER_MODES, state="readonly", width=18)
            elif key == "trigger_value":
                widget = ttk.Combobox(
                    parent,
                    textvariable=self.vars[key],
                    values=TRIGGER_VALUE_LABELS,
                    state="readonly",
                    width=18,
                )
            elif key == "station":
                self.station_combo = ttk.Combobox(parent, textvariable=self.vars[key], values=[], width=18)
                self.station_combo.bind(
                    "<<ComboboxSelected>>",
                    lambda _event: self.refresh_choice_options(to_int(self.vars["action_id"].get())),
                )
                self.station_combo.bind(
                    "<FocusOut>",
                    lambda _event: self.refresh_choice_options(to_int(self.vars["action_id"].get())),
                )
                widget = self.station_combo
            elif key == "depends_on":
                self.dependency_combo = ttk.Combobox(
                    parent,
                    textvariable=self.dependency_var,
                    values=[WAIT_NONE_LABEL],
                    state="readonly",
                    width=28,
                )
                widget = self.dependency_combo
            else:
                widget = ttk.Entry(parent, textvariable=self.vars[key], width=21)
            widget.grid(row=row, column=1, sticky=tk.EW, padx=4, pady=3)
            self.form_labels[key] = label_widget
            self.form_widgets[key] = widget

        ttk.Label(parent, text="动作").grid(row=0, column=2, sticky=tk.W, padx=(12, 4), pady=3)
        self.action_text = tk.Text(parent, width=34, height=5, wrap=tk.WORD)
        self.action_text.grid(row=1, column=2, rowspan=4, sticky=tk.NSEW, padx=(12, 4), pady=3)
        ttk.Label(parent, text="备注").grid(row=5, column=2, sticky=tk.W, padx=(12, 4), pady=3)
        ttk.Entry(parent, textvariable=self.vars["note"]).grid(row=6, column=2, sticky=tk.EW, padx=(12, 4), pady=3)

        ttk.Label(parent, text="第一轮等待").grid(row=7, column=0, sticky=tk.W, padx=4, pady=3)
        self.dep_listbox = tk.Listbox(parent, height=3, exportselection=False)
        self.dep_listbox.grid(row=7, column=1, sticky=tk.EW, padx=4, pady=3)
        dep_buttons = ttk.Frame(parent)
        dep_buttons.grid(row=7, column=2, sticky=tk.W, padx=(12, 4), pady=3)
        ttk.Button(dep_buttons, text="加入第一轮", command=lambda: self.add_selected_dependency("first")).pack(side=tk.LEFT, padx=2)
        ttk.Button(dep_buttons, text="移除第一轮", command=lambda: self.remove_selected_dependency("first")).pack(side=tk.LEFT, padx=2)
        ttk.Button(dep_buttons, text="清空第一轮", command=lambda: self.clear_selected_dependencies("first")).pack(side=tk.LEFT, padx=2)

        ttk.Label(parent, text="后续轮等待").grid(row=8, column=0, sticky=tk.W, padx=4, pady=3)
        self.later_dep_listbox = tk.Listbox(parent, height=3, exportselection=False)
        self.later_dep_listbox.grid(row=8, column=1, sticky=tk.EW, padx=4, pady=3)
        later_buttons = ttk.Frame(parent)
        later_buttons.grid(row=8, column=2, sticky=tk.W, padx=(12, 4), pady=3)
        ttk.Button(later_buttons, text="加入后续轮", command=lambda: self.add_selected_dependency("later")).pack(side=tk.LEFT, padx=2)
        ttk.Button(later_buttons, text="移除后续轮", command=lambda: self.remove_selected_dependency("later")).pack(side=tk.LEFT, padx=2)
        ttk.Button(later_buttons, text="清空后续轮", command=lambda: self.clear_selected_dependencies("later")).pack(side=tk.LEFT, padx=2)

        buttons = ttk.Frame(parent)
        buttons.grid(row=9, column=0, columnspan=3, sticky=tk.EW, pady=(8, 4))
        ttk.Button(buttons, text="添加", command=self.add_action).pack(side=tk.LEFT, padx=2)
        ttk.Button(buttons, text="更新选中", command=self.update_action).pack(side=tk.LEFT, padx=2)
        ttk.Button(buttons, text="删除选中", command=self.delete_action).pack(side=tk.LEFT, padx=2)
        ttk.Button(buttons, text="上移", command=lambda: self.move_action(-1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(buttons, text="下移", command=lambda: self.move_action(1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(buttons, text="示例", command=self.load_example).pack(side=tk.LEFT, padx=2)
        ttk.Button(buttons, text="清空", command=self.clear_actions).pack(side=tk.LEFT, padx=2)

        self.update_trigger_visibility()
        parent.columnconfigure(1, weight=1)
        parent.columnconfigure(2, weight=1)

    def _build_tree(self, parent: ttk.Frame) -> None:
        columns = ("action", "duration", "repeat", "first_deps", "later_deps", "trigger")
        self.tree = ttk.Treeview(parent, columns=columns, show="tree headings", height=12)
        headings = {
            "action": "动作",
            "duration": "时间",
            "repeat": "次数",
            "first_deps": "第一轮等待",
            "later_deps": "后续轮等待",
            "trigger": "触发",
        }
        self.tree.heading("#0", text="动作主体")
        self.tree.column("#0", width=140, anchor=tk.W, stretch=False)
        widths = {"action": 260, "duration": 60, "repeat": 70, "first_deps": 180, "later_deps": 180, "trigger": 112}
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], anchor=tk.W, stretch=col == "action")

        yscroll = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self.tree.yview)
        xscroll = ttk.Scrollbar(parent, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.grid(row=0, column=0, sticky=tk.NSEW)
        yscroll.grid(row=0, column=1, sticky=tk.NS)
        xscroll.grid(row=1, column=0, sticky=tk.EW)
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.bind("<Double-1>", self.on_tree_double_click)

    def next_action_id(self) -> int:
        used = {action.action_id for action in self.actions}
        value = 1
        while value in used:
            value += 1
        return value

    def action_display_label(self, action: FlowAction) -> str:
        parts = [
            action.station.strip() or "未设置动作主体",
            (action.action.strip().splitlines()[0] if action.action.strip() else "未命名动作"),
        ]
        return " / ".join(parts)

    def refresh_choice_options(self, exclude_action_id: Optional[int] = None) -> None:
        stations = sorted({action.station for action in self.actions if action.station})
        if hasattr(self, "station_combo"):
            self.station_combo.configure(values=stations)

        current_station = self.vars["station"].get().strip() if hasattr(self, "vars") else ""
        labels = [WAIT_NONE_LABEL]
        self.dep_label_to_id = {}
        for action in self.actions:
            if exclude_action_id is not None and action.action_id == exclude_action_id:
                continue
            if current_station and action.station.strip() == current_station:
                continue
            label = self.action_display_label(action)
            base_label = label
            suffix_names = ["二", "三", "四", "五", "六", "七", "八", "九", "十"]
            suffix_index = 0
            while label in self.dep_label_to_id:
                suffix = suffix_names[suffix_index] if suffix_index < len(suffix_names) else "多"
                label = f"{base_label}（同名{suffix}）"
                suffix_index += 1
            self.dep_label_to_id[label] = action.action_id
            labels.append(label)
        if hasattr(self, "dependency_combo"):
            self.dependency_combo.configure(values=labels)
            if self.dependency_var.get() not in labels:
                self.dependency_var.set(WAIT_NONE_LABEL)

    def dependency_label_to_ids(self, label: str) -> str:
        action_id = self.dep_label_to_id.get(label)
        return "" if action_id is None else str(action_id)

    def dependency_ids_to_labels(self, depends_on: str) -> List[str]:
        labels = []
        for target in split_ids(depends_on):
            for action in self.actions:
                if action.action_id == target:
                    labels.append(self.action_display_label(action))
                    break
        return labels

    def dependency_ids_to_label(self, depends_on: str) -> str:
        ids = split_ids(depends_on)
        if not ids:
            return WAIT_NONE_LABEL
        labels = self.dependency_ids_to_labels(depends_on)
        return "；".join(labels) if labels else WAIT_NONE_LABEL

    def refresh_dependency_listboxes(self) -> None:
        if hasattr(self, "dep_listbox"):
            self.dep_listbox.delete(0, tk.END)
            for label in self.dependency_ids_to_labels(join_ids(self.selected_dep_ids)):
                self.dep_listbox.insert(tk.END, label)
        if hasattr(self, "later_dep_listbox"):
            self.later_dep_listbox.delete(0, tk.END)
            for label in self.dependency_ids_to_labels(join_ids(self.selected_later_dep_ids)):
                self.later_dep_listbox.insert(tk.END, label)
        self.update_trigger_visibility()

    def update_trigger_visibility(self) -> None:
        if not hasattr(self, "form_widgets"):
            return
        has_wait = bool(self.selected_dep_ids or self.selected_later_dep_ids)
        for key in ("trigger_mode", "trigger_value"):
            label = self.form_labels.get(key)
            widget = self.form_widgets.get(key)
            if not label or not widget:
                continue
            if has_wait:
                label.grid()
                widget.grid()
            else:
                label.grid_remove()
                widget.grid_remove()

    def add_selected_dependency(self, target: str) -> None:
        action_id = self.dep_label_to_id.get(self.dependency_var.get())
        if action_id is None:
            return
        dep_ids = self.selected_later_dep_ids if target == "later" else self.selected_dep_ids
        if action_id not in dep_ids:
            dep_ids.append(action_id)
        self.refresh_dependency_listboxes()

    def remove_selected_dependency(self, target: str) -> None:
        listbox = self.later_dep_listbox if target == "later" else self.dep_listbox
        dep_ids = self.selected_later_dep_ids if target == "later" else self.selected_dep_ids
        selection = listbox.curselection()
        if not selection:
            return
        index = selection[0]
        if 0 <= index < len(dep_ids):
            del dep_ids[index]
        self.refresh_dependency_listboxes()

    def clear_selected_dependencies(self, target: str) -> None:
        if target == "later":
            self.selected_later_dep_ids = []
        else:
            self.selected_dep_ids = []
        self.refresh_dependency_listboxes()

    def form_to_action(self) -> FlowAction:
        action_id = to_int(self.vars["action_id"].get(), self.next_action_id())
        duration = to_float(self.vars["duration"].get(), None)
        repeat = count_to_int(self.vars["repeat"].get(), None)
        trigger_value = trigger_value_to_int(self.vars["trigger_value"].get(), 1)
        station = self.vars["station"].get().strip()
        action_text = self.action_text.get("1.0", tk.END).strip()
        if action_id is None:
            raise ValueError("动作编号必须是数字。")
        if duration is None or duration <= 0:
            raise ValueError("时间必须是大于 0 的数字，例如 0.5 或 0.5秒。")
        if repeat is None or repeat < 1:
            raise ValueError("重复次数必须是大于 0 的整数，例如 3。")
        if trigger_value is None or trigger_value < 1:
            raise ValueError("触发次数/偏移必须大于 0。")
        if not station:
            raise ValueError("请填写动作主体。")
        if not action_text:
            raise ValueError("请填写动作。")
        self.refresh_choice_options(action_id)
        return FlowAction(
            action_id=action_id,
            station=station,
            module="",
            action=action_text,
            duration=round(duration, 3),
            repeat=repeat,
            depends_on=join_ids(self.selected_dep_ids),
            later_depends_on=join_ids(self.selected_later_dep_ids),
            trigger_mode=self.vars["trigger_mode"].get().strip() or "同次完成",
            trigger_value=trigger_value,
            note=self.vars["note"].get().strip(),
        )

    def action_to_form(self, action: FlowAction) -> None:
        self.vars["action_id"].set(str(action.action_id))
        self.vars["station"].set(action.station)
        self.vars["duration"].set(str(action.duration))
        self.vars["repeat"].set(str(action.repeat))
        self.vars["depends_on"].set(action.depends_on)
        self.selected_dep_ids = split_ids(action.depends_on)
        self.selected_later_dep_ids = split_ids(action.later_depends_on)
        self.refresh_choice_options(exclude_action_id=action.action_id)
        self.dependency_var.set(self.dependency_ids_to_label(action.depends_on))
        self.refresh_dependency_listboxes()
        self.vars["trigger_mode"].set(action.trigger_mode)
        self.vars["trigger_value"].set(trigger_value_to_label(action.trigger_value))
        self.vars["note"].set(action.note)
        self.action_text.delete("1.0", tk.END)
        self.action_text.insert("1.0", action.action)

    def clear_form(self) -> None:
        for key in ("station", "depends_on", "note"):
            self.vars[key].set("")
        self.selected_dep_ids = []
        self.selected_later_dep_ids = []
        self.refresh_choice_options()
        self.dependency_var.set(WAIT_NONE_LABEL)
        self.refresh_dependency_listboxes()
        self.vars["action_id"].set(str(self.next_action_id()))
        self.vars["duration"].set("0.1")
        self.vars["repeat"].set("1")
        self.vars["trigger_mode"].set("同次完成")
        self.vars["trigger_value"].set(TRIGGER_VALUE_LABELS[0])
        self.action_text.delete("1.0", tk.END)

    def refresh_tree(self) -> None:
        self.refresh_choice_options()
        for item in self.tree.get_children():
            self.tree.delete(item)
        station_nodes = {}
        for action in self.actions:
            station = action.station.strip() or "未设置动作主体"
            if station not in station_nodes:
                node_id = f"station::{station}"
                suffix = 2
                while self.tree.exists(node_id):
                    node_id = f"station::{station}::{suffix}"
                    suffix += 1
                station_nodes[station] = node_id
                self.tree.insert("", tk.END, iid=node_id, text=station, open=True, values=("", "", "", "", "", ""))
            first_dep_label = self.dependency_ids_to_label(action.depends_on)
            later_dep_label = self.dependency_ids_to_label(action.later_depends_on)
            trigger = "" if first_dep_label == WAIT_NONE_LABEL and later_dep_label == WAIT_NONE_LABEL else f"{action.trigger_mode}:{trigger_value_to_label(action.trigger_value)}"
            self.tree.insert(
                station_nodes[station],
                tk.END,
                iid=str(action.action_id),
                text="",
                values=(
                    action.action,
                    action.duration,
                    action.repeat,
                    first_dep_label,
                    later_dep_label,
                    trigger,
                ),
            )
        self.clear_form()

    def invalidate_diagram_preview(self) -> None:
        self.events = []
        self.clear_preview()
        self.status_var.set("动作清单已变化，请重新生成时序图。")

    def selected_action_index(self) -> Optional[int]:
        selection = self.tree.selection()
        if not selection:
            return None
        action_id = to_int(selection[0])
        if action_id is None:
            return None
        for idx, action in enumerate(self.actions):
            if action.action_id == action_id:
                return idx
        return None

    def add_action(self) -> None:
        try:
            action = self.form_to_action()
            if any(existing.action_id == action.action_id for existing in self.actions):
                raise ValueError("动作编号已存在，请换一个编号或使用“更新选中”。")
            self.actions.append(action)
            self.invalidate_diagram_preview()
            self.refresh_tree()
        except Exception as exc:
            messagebox.showerror("输入错误", str(exc))

    def update_action(self) -> None:
        idx = self.selected_action_index()
        if idx is None:
            messagebox.showinfo("提示", "请先选择一个动作。")
            return
        try:
            action = self.form_to_action()
            for pos, existing in enumerate(self.actions):
                if pos != idx and existing.action_id == action.action_id:
                    raise ValueError("动作编号已存在。")
            self.actions[idx] = action
            self.invalidate_diagram_preview()
            self.refresh_tree()
        except Exception as exc:
            messagebox.showerror("输入错误", str(exc))

    def delete_action(self) -> None:
        idx = self.selected_action_index()
        if idx is None:
            return
        del self.actions[idx]
        self.invalidate_diagram_preview()
        self.refresh_tree()

    def move_action(self, direction: int) -> None:
        idx = self.selected_action_index()
        if idx is None:
            return
        new_idx = idx + direction
        if new_idx < 0 or new_idx >= len(self.actions):
            return
        self.actions[idx], self.actions[new_idx] = self.actions[new_idx], self.actions[idx]
        self.invalidate_diagram_preview()
        self.refresh_tree()
        self.tree.selection_set(str(self.actions[new_idx].action_id))

    def clear_actions(self) -> None:
        if self.actions and not messagebox.askyesno("确认", "确定清空当前动作清单吗？"):
            return
        self.actions = []
        self.events = []
        self.refresh_tree()
        self.clear_preview()
        self.status_var.set("已清空。")

    def load_example(self) -> None:
        self.actions = [
            FlowAction(action_id=1, station="人工", module="上料", action="人工放料", duration=0.3, repeat=3, note="连续放 3 次"),
            FlowAction(
                action_id=2,
                station="动作主体1",
                module="滑台",
                action="滑台移动到贴附位",
                duration=1.0,
                repeat=1,
                depends_on="1",
                later_depends_on="4",
                trigger_mode="固定次数完成",
                trigger_value=3,
                note="第一轮等人工第 3 次，后续轮等检测回位",
            ),
            FlowAction(action_id=3, station="动作主体1", module="定位", action="侧定位模组定位", duration=0.5, repeat=1, note="同动作主体默认接在滑台后"),
            FlowAction(action_id=4, station="动作主体1", module="检测", action="检测完成并回到出料位", duration=0.8, repeat=1, note="同动作主体默认接在定位后"),
        ]
        self.events = []
        self.refresh_tree()
        self.status_var.set("已载入示例：动作 2 等动作 1 第 3 次完成后开始。")

    def on_tree_select(self, _event=None) -> None:
        idx = self.selected_action_index()
        if idx is not None:
            self.action_to_form(self.actions[idx])

    def on_tree_double_click(self, event) -> None:
        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or column_id not in ("#1", "#2", "#3"):
            return
        action_id = to_int(item_id)
        action = next((item for item in self.actions if item.action_id == action_id), None)
        if action is None:
            return
        bbox = self.tree.bbox(item_id, column_id)
        if not bbox:
            return
        x, y, width, height = bbox
        if column_id == "#1":
            field = "action"
            old_value = action.action
        elif column_id == "#2":
            field = "duration"
            old_value = str(action.duration)
        else:
            field = "repeat"
            old_value = str(action.repeat)
        editor = ttk.Entry(self.tree)
        editor.insert(0, old_value)
        editor.select_range(0, tk.END)
        editor.place(x=x, y=y, width=width, height=height)
        editor.focus_set()

        def commit(_event=None) -> None:
            if not editor.winfo_exists():
                return
            value = editor.get().strip()
            try:
                if field == "duration":
                    number = to_float(value)
                    if number is None or number <= 0:
                        raise ValueError("时间必须是大于 0 的数字，例如 0.5 或 0.5秒。")
                    action.duration = round(number, 3)
                elif field == "repeat":
                    number = count_to_int(value, None)
                    if number is None or number < 1:
                        raise ValueError("重复次数必须是大于 0 的整数，例如 3。")
                    action.repeat = number
                else:
                    if not value:
                        raise ValueError("请填写动作。")
                    action.action = value
                self.invalidate_diagram_preview()
                editor.destroy()
                self.refresh_tree()
                self.tree.selection_set(str(action.action_id))
            except Exception as exc:
                messagebox.showerror("修改失败", str(exc))
                editor.focus_set()

        def cancel(_event=None) -> None:
            if editor.winfo_exists():
                editor.destroy()

        editor.bind("<Return>", commit)
        editor.bind("<FocusOut>", commit)
        editor.bind("<Escape>", cancel)

    def import_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if not path:
            return
        try:
            actions, events = load_workbook_data(path, self.current_cycle_count())
            self.actions = actions
            self.events = events
            self.current_file = path
            self.refresh_tree()
            self.draw_events(events)
            source = "动作清单" if actions else "时序明细"
            cycle_note = f"，已按 {self.current_cycle_count()} 轮生成预览" if actions else ""
            self.status_var.set(f"已导入 {source}{cycle_note}：{path}")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("导入失败", str(exc))

    def generate_diagram(self) -> None:
        try:
            cycle_count = self.current_cycle_count()
            if self.actions:
                self.events = build_events_from_actions(self.actions, cycle_count)
            if not self.events:
                raise ValueError("请先填写动作，或导入 Excel。")
            self.draw_events(self.events)
            self.status_var.set(
                f"已生成 {cycle_count} 轮时序图，共 {len(self.events)} 个展开步骤。{cycle_summary_text(self.events)}。"
            )
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("生成失败", str(exc))

    def export_excel(self) -> None:
        try:
            cycle_count = self.current_cycle_count()
            if self.actions:
                self.events = build_events_from_actions(self.actions, cycle_count)
            if not self.events:
                raise ValueError("没有可导出的时序数据。")
            default_name = f"时序图数据_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel 文件", "*.xlsx")],
            )
            if not path:
                return
            actions = self.actions or self.events_to_actions(self.events)
            write_workbook(path, actions, self.events)
            self.current_file = path
            self.status_var.set(f"已导出 Excel：{path}")
            messagebox.showinfo("成功", f"Excel 已导出：\n{path}")
        except Exception as exc:
            traceback.print_exc()
            messagebox.showerror("导出失败", str(exc))

    def export_template(self) -> None:
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="时序图填写模板.xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not path:
            return
        try:
            cycle_count = self.current_cycle_count()
            events = build_events_from_actions(self.actions or [], cycle_count)
        except Exception:
            actions = [
                FlowAction(action_id=1, station="人工", module="上料", action="人工放料", duration=0.3, repeat=3, note="连续放 3 次"),
                FlowAction(
                    action_id=2,
                    station="动作主体1",
                    module="滑台",
                    action="滑台移动到贴附位",
                    duration=1.0,
                    repeat=1,
                    depends_on="1",
                    trigger_mode="固定次数完成",
                    trigger_value=3,
                    note="等待动作 1 第 3 次完成",
                ),
            ]
            events = build_events_from_actions(actions, cycle_count)
            write_workbook(path, actions, events)
        else:
            write_workbook(path, self.actions, events)
        messagebox.showinfo("成功", f"模板已导出：\n{path}")

    def export_image(self) -> None:
        if not self.events:
            self.generate_diagram()
            if not self.events:
                return
        if not HAS_MATPLOTLIB:
            path = filedialog.asksaveasfilename(
                defaultextension=".ps",
                filetypes=[("PostScript 图片", "*.ps")],
            )
            if not path:
                return
            self.canvas.postscript(file=path, colormode="color")
            self.status_var.set(f"已导出图片：{path}")
            messagebox.showinfo("成功", f"当前环境没有 Matplotlib，已导出 PostScript 图片：\n{path}")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG 图片", "*.png"), ("JPEG 图片", "*.jpg")],
        )
        if not path:
            return
        self.fig.savefig(path, dpi=300, bbox_inches="tight")
        self.status_var.set(f"已导出图片：{path}")
        messagebox.showinfo("成功", f"图片已导出：\n{path}")

    def events_to_actions(self, events: List[dict]) -> List[FlowAction]:
        actions = []
        for event in events:
            actions.append(
                FlowAction(
                    action_id=event.get("source_action", event["step"]),
                    station=event.get("station", ""),
                    module=event.get("module", ""),
                    action=event.get("action", ""),
                    duration=event.get("duration", 0.1),
                    repeat=1,
                    depends_on=join_ids(event.get("deps", [])),
                    trigger_mode="同次完成",
                    trigger_value=1,
                )
            )
        return actions

    def draw_events(self, events: List[dict]) -> None:
        if not HAS_MATPLOTLIB:
            self.draw_events_on_canvas(events)
            return

        self.ax.clear()
        if not events:
            self.canvas.draw()
            return

        group_keys = []
        labels = []
        for event in events:
            key = event.get("source_action", event["step"])
            if key not in group_keys:
                group_keys.append(key)
                name = event.get("action", "")
                labels.append(f"{key}. {name}".strip())
        y_map = {key: idx for idx, key in enumerate(group_keys)}
        event_by_step = {event["step"]: event for event in events}
        colors = ["#E89B3A", "#4F9BD9", "#70AD47", "#D96AA7", "#8064A2", "#00A6A6", "#A5A5A5"]

        for event in events:
            y = y_map[event.get("source_action", event["step"])]
            color = colors[(event.get("source_action", event["step"]) - 1) % len(colors)]
            self.ax.barh(
                y,
                event["duration"],
                left=event["start"],
                height=0.42,
                color=color,
                edgecolor="#222222",
            )
            text = f"{event.get('cycle', 1)}-{event.get('occurrence', 1)}"
            self.ax.text(
                event["start"] + event["duration"] / 2,
                y,
                text,
                ha="center",
                va="center",
                fontsize=9,
                color="black",
            )

        for event in events:
            y2 = y_map[event.get("source_action", event["step"])]
            for edge in visible_dependency_edges(event):
                dep = event_by_step.get(edge.get("step"))
                if not dep:
                    continue
                y1 = y_map[dep.get("source_action", dep["step"])]
                self.ax.annotate(
                    "",
                    xy=(event["start"], y2),
                    xytext=(dep["end"], y1),
                    arrowprops=dict(arrowstyle="->", color="#2E7D32", lw=1.4, connectionstyle="arc3,rad=0.12"),
                )

        max_time = max(event["end"] for event in events)
        self.ax.set_xlim(0, max(max_time * 1.05, 1))
        self.ax.set_ylim(len(group_keys) - 0.4, -0.6)
        self.ax.set_yticks(range(len(labels)))
        self.ax.set_yticklabels(labels)
        self.ax.set_xlabel("时间（秒）")
        self.ax.set_title("时序图")
        self.ax.grid(axis="x", linestyle="--", alpha=0.35)
        self.fig.tight_layout()
        self.canvas.draw()

    def clear_preview(self) -> None:
        if HAS_MATPLOTLIB:
            self.ax.clear()
            self.canvas.draw()
        else:
            self.canvas.delete("all")

    def draw_events_on_canvas(self, events: List[dict]) -> None:
        self.canvas.delete("all")
        if not events:
            return

        self.canvas.update_idletasks()
        width = max(self.canvas.winfo_width(), 900)
        height = max(self.canvas.winfo_height(), 520)
        left = 260
        right = 40
        top = 44
        row_h = 54

        group_keys = []
        labels = []
        for event in events:
            key = event.get("source_action", event["step"])
            if key not in group_keys:
                group_keys.append(key)
                name = event.get("action", "")
                labels.append(f"{key}. {name}".strip())
        y_map = {key: idx for idx, key in enumerate(group_keys)}
        max_time = max(event["end"] for event in events)
        scale = (width - left - right) / max(max_time, 1)
        colors = ["#E89B3A", "#4F9BD9", "#70AD47", "#D96AA7", "#8064A2", "#00A6A6", "#A5A5A5"]

        needed_height = top + row_h * len(group_keys) + 60
        self.canvas.configure(scrollregion=(0, 0, width, max(height, needed_height)))
        self.canvas.create_text(width / 2, 18, text="时序图", font=("Microsoft YaHei", 14, "bold"))

        for tick in range(0, int(math.ceil(max_time)) + 1):
            x = left + tick * scale
            self.canvas.create_line(x, top - 12, x, needed_height - 34, fill="#DDDDDD")
            self.canvas.create_text(x, needed_height - 22, text=f"{tick}s", font=("Microsoft YaHei", 9))

        event_by_step = {event["step"]: event for event in events}
        for idx, label in enumerate(labels):
            y = top + idx * row_h + row_h / 2
            self.canvas.create_text(10, y, text=label, anchor=tk.W, width=left - 20, font=("Microsoft YaHei", 9))
            self.canvas.create_line(left, y + row_h / 2 - 4, width - right, y + row_h / 2 - 4, fill="#EEEEEE")

        for event in events:
            y = top + y_map[event.get("source_action", event["step"])] * row_h + row_h / 2
            x1 = left + event["start"] * scale
            x2 = left + event["end"] * scale
            color = colors[(event.get("source_action", event["step"]) - 1) % len(colors)]
            self.canvas.create_rectangle(x1, y - 14, x2, y + 14, fill=color, outline="#222222")
            self.canvas.create_text((x1 + x2) / 2, y, text=str(event.get("occurrence", 1)), font=("Microsoft YaHei", 9))

        for event in events:
            y2 = top + y_map[event.get("source_action", event["step"])] * row_h + row_h / 2
            x2 = left + event["start"] * scale
            for edge in visible_dependency_edges(event):
                dep = event_by_step.get(edge.get("step"))
                if not dep:
                    continue
                y1 = top + y_map[dep.get("source_action", dep["step"])] * row_h + row_h / 2
                x1 = left + dep["end"] * scale
                self.canvas.create_line(x1, y1, x2, y2, fill="#2E7D32", arrow=tk.LAST, smooth=True, width=1.6)


def main() -> None:
    root = tk.Tk()
    app = TimingDiagramApp(root)
    root.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()


if __name__ == "__main__":
    main()













