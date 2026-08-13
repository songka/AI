"""Mechanical Quotation System — Demo UI Application.

Launch: .venv/Scripts/python -m quotation.ui.demo_app
"""

from __future__ import annotations

import json
import sys
import traceback
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from tkinter import filedialog, messagebox
import tkinter as tk
from typing import Any

# Ensure the project root is on the path
_PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import ezdxf

from quotation.domain.quote import PriceSource, Quote, QuoteItem
from quotation.infrastructure.dxf.reader import DxfReader
from quotation.infrastructure.feature.geometric import GeometricExtractor
from quotation.infrastructure.feature.manufacturing import ManufacturingExtractor
from quotation.infrastructure.feature.quotation_mapper import QuotationMapper
from quotation.infrastructure.rules.pricing_resolver import PricingResolver
from quotation.infrastructure.rules.quote_builder import QuoteBuilder
from quotation.ui.viewmodels import QuoteViewModel, TaxResult
from quotation.ui.widgets import (
    CONTENT_BG, FONT_FAMILY, HEADER_BG, HEADER_FG,
    BatchQuotePage, NavPanel, NewQuotePage, PlaceholderPage,
)

# ---------------------------------------------------------------------------
# Demo part definitions (mirrors cli/main.py DEMO_PARTS)
# ---------------------------------------------------------------------------

DEMO_PARTS = {
    "J003": {
        "part_number": "UC1000005854",
        "part_name": "J003",
        "material": "S50C",
        "historical_price": 1425.0,
        "size": (928, 796),
        "circles": [(200, 398, 3), (350, 398, 3), (500, 398, 3), (650, 398, 3)],
        "texts": [
            ("S50C", 10, 810, 8),
            ("6-M6", 200, 400, 5),
            ("表面鍍鉻", 10, 820, 5),
        ],
    },
    "W001": {
        "part_number": "UC2020083221",
        "part_name": "W001",
        "material": "鋁型材",
        "size": (1300, 1300),
        "circles": [],
        "texts": [
            ("鋁型材 40x40", 10, 1320, 6),
            ("防護圍欄", 10, 1340, 6),
            ("門組件", 10, 1360, 5),
            ("白色透明亞克力", 10, 1380, 4),
            ("合頁", 10, 1400, 4),
            ("磁吸", 10, 1420, 4),
            ("把手", 10, 1440, 4),
            ("角碼", 10, 1460, 4),
            ("加強筋焊接", 10, 1480, 4),
        ],
    },
}


# ---------------------------------------------------------------------------
# Pipeline runner (reuses existing infrastructure — no formula duplication)
# ---------------------------------------------------------------------------

def run_quotation_pipeline(part_name: str) -> tuple[Quote | None, dict[str, Any], str | None]:
    """Run the full 6-layer quotation pipeline for a demo part.

    Returns:
        (Quote, feature_summary_dict, error_message_or_None)
    """
    part = DEMO_PARTS.get(part_name)
    if not part:
        return None, {}, f"Unknown part: {part_name}"

    cwd = Path.cwd()

    try:
        # 1. Generate DXF
        doc = ezdxf.new()
        doc.header["$INSUNITS"] = 4
        msp = doc.modelspace()
        w, h = part["size"]
        msp.add_line((0, 0), (w, 0))
        msp.add_line((w, 0), (w, h))
        msp.add_line((w, h), (0, h))
        msp.add_line((0, h), (0, 0))
        for cx, cy, r in part["circles"]:
            msp.add_circle((cx, cy), radius=r)
        for content, x, y, height in part["texts"]:
            msp.add_text(content, height=height).set_placement((x, y))
        dxf_path = cwd / f"demo_{part_name}.dxf"
        doc.saveas(str(dxf_path))

        # 2. CAD Import -> Drawing
        reader = DxfReader()
        import_result = reader.read(dxf_path)
        drawing = import_result.drawing

        # 3. Feature Extraction
        geo_ext = GeometricExtractor()
        geo = geo_ext.extract(drawing.raw_entities)

        mfg_ext = ManufacturingExtractor()
        mfg = mfg_ext.extract(geo)

        # 4. Quotation Mapping
        mapper = QuotationMapper()
        qf = mapper.map(mfg, geo)

        # 5. Pricing
        resolver = PricingResolver()
        items: list[QuoteItem] = []
        for mq in qf.machining:
            items.extend(resolver.resolve_machining(mq))
        for fq in qf.frames:
            items.extend(resolver.resolve_frame(fq))
        for aq in qf.assemblies:
            items.extend(resolver.resolve_assembly(aq))

        # 6. Quote Builder
        builder = QuoteBuilder()
        feat_conf = mfg.material.confidence if mfg.material else None
        quote = builder.build(
            quote_id=f"Q-DEMO-{part_name}",
            drawing_id=f"DEMO-{part_name}",
            part_number=part["part_number"],
            part_name=part["part_name"],
            material=part["material"],
            items=items,
            feature_confidence=feat_conf,
            price_version=resolver.price_version,
            rule_version="1.0",
        )

        # Cleanup temp DXF
        dxf_path.unlink(missing_ok=True)

        # Feature summary
        bbox = geo.bounding_box
        weight_kg = None
        for item in quote.items:
            if item.category == "material" and item.evidence:
                import re
                m = re.search(r"weight_kg=([\d.]+)", item.evidence)
                if m:
                    weight_kg = float(m.group(1))
                    break

        feature_summary = {
            "bounding_box": f"{bbox.length:.0f}×{bbox.width:.0f} mm" if bbox else "—",
            "hole_candidates": geo.candidate_count,
            "mfg_holes": mfg.total_holes,
            "mfg_threads": mfg.total_threads,
            "frames": len(mfg.frames),
            "assemblies": len(mfg.structure_assemblies),
            "accessories": len(mfg.structure_accessories),
            "welds": len(mfg.welds),
            "weight": f"{weight_kg:.1f} kg" if weight_kg else "—",
            "part_type": "加工件" if len(mfg.frames) == 0 else "結構件",
            "material_raw": part["material"],
        }

        return quote, feature_summary, None

    except FileNotFoundError as e:
        return None, {}, f"規則文件缺失：{e}"
    except Exception as e:
        return None, {}, f"解析失敗：{e}\n{traceback.format_exc()}"


# ---------------------------------------------------------------------------
# Main Demo Application
# ---------------------------------------------------------------------------

class DemoApp(tk.Tk):
    """Main Tkinter application window."""

    def __init__(self):
        super().__init__()
        self.title("機械加工件智能報價系統 — Mechanical Quotation System")
        self.geometry("1280x720")
        self.minsize(1024, 600)
        self.configure(bg=CONTENT_BG)

        # Center on screen
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - 1280) // 2
        y = (sh - 720) // 2
        self.geometry(f"+{x}+{y}")

        # Font defaults
        self.option_add("*Font", (FONT_FAMILY[0], 10))

        # State
        self._current_quote: Quote | None = None
        self._current_feature_summary: dict[str, Any] = {}
        self._content: tk.Frame | None = None

        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ------------------------------------------------------------------
    # UI Setup
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        # Main container
        self._main = tk.Frame(self, bg=CONTENT_BG)
        self._main.pack(fill=tk.BOTH, expand=True)

        # Left nav
        self._nav = NavPanel(self._main, on_nav_change=self._switch_page)
        self._nav.pack(side=tk.LEFT, fill=tk.Y)

        # Right content area
        self._content_area = tk.Frame(self._main, bg=CONTENT_BG)
        self._content_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Show default page
        self._switch_page("新建報價")

    def _switch_page(self, name: str) -> None:
        """Destroy current content and build the requested page."""
        if self._content is not None:
            self._content.destroy()

        if name == "新建報價":
            self._content = NewQuotePage(
                self._content_area,
                on_load_j003=lambda: self._load_demo("J003"),
                on_load_w001=lambda: self._load_demo("W001"),
                on_export=self._export_excel,
            )
        elif name == "批量報價":
            self._content = BatchQuotePage(
                self._content_area,
                on_scan_dir=self._batch_scan_dir,
                on_scan_files=self._batch_scan_files,
                on_run_batch=self._batch_run,
                on_export_selected=self._batch_export_selected,
                on_export_all=self._batch_export_all,
                on_open_dir=self._batch_open_dir,
            )
        else:
            self._content = PlaceholderPage(self._content_area, name)

        self._content.pack(fill=tk.BOTH, expand=True)

    # ------------------------------------------------------------------
    # Load demo part
    # ------------------------------------------------------------------

    def _load_demo(self, part_name: str) -> None:
        """Run the pipeline and populate the UI."""
        page = self._content
        if not isinstance(page, NewQuotePage):
            return

        page.update_status("解析中...", "orange", 0)
        self.update_idletasks()

        quote, feature_summary, error = run_quotation_pipeline(part_name)

        if error:
            messagebox.showerror("解析錯誤", error)
            page.update_status("解析失敗", "red", 0)
            return

        if quote is None:
            messagebox.showerror("錯誤", "無法生成報價")
            page.update_status("系統錯誤", "red", 0)
            return

        self._current_quote = quote
        self._current_feature_summary = feature_summary

        # Build ViewModel
        tax = TaxResult.calculate(quote.items, Decimal("0.17"))
        vm = QuoteViewModel(quote=quote, tax=tax)

        # Update basic info
        material_raw = feature_summary.get("material_raw", quote.material or "—")
        part_type = feature_summary.get("part_type", "—")
        basic_fields = [
            ("圖號", quote.part_number or "—"),
            ("料號", quote.part_number or "—"),
            ("材料", material_raw),
            ("規格尺寸", feature_summary.get("bounding_box", "—")),
            ("表面處理", self._extract_surface(quote)),
            ("零件類型", part_type),
            ("規則版本", quote.rule_version or "—"),
            ("價格版本", quote.price_version or "—"),
        ]
        page.update_basic_info(basic_fields)

        # Update feature summary
        fs = feature_summary
        feature_fields = [
            ("Bounding Box", fs.get("bounding_box", "—")),
            ("孔數", str(fs.get("mfg_holes", 0))),
            ("螺紋數", str(fs.get("mfg_threads", 0))),
            ("Frame 數", str(fs.get("frames", 0))),
            ("Assembly 數", str(fs.get("assemblies", 0))),
            ("Accessory 數", str(fs.get("accessories", 0))),
            ("重量", fs.get("weight", "—")),
            ("重量來源", "CAD bounding box 估算"),
            ("Confidence", f"{quote.overall_confidence:.0%}"),
        ]
        page.update_feature_summary(feature_fields)

        # Update table
        page.update_table(vm)

        # Update trace
        page.update_trace(None)

        # Update summary
        page.update_summary(vm)

        # Update status
        page.update_status(
            vm.status_text,
            vm.status_color,
            quote.cost_completion,
        )

    @staticmethod
    def _extract_surface(quote: Quote) -> str:
        """Extract surface treatment info from quote items."""
        for item in quote.items:
            if item.category == "surface":
                return item.name
        return "無"

    # ------------------------------------------------------------------
    # Export Excel
    # ------------------------------------------------------------------

    def _export_excel(self) -> None:
        """Export quotation to Excel using openpyxl."""
        if self._current_quote is None:
            messagebox.showwarning("警告", "請先載入報價資料")
            return

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            quote = self._current_quote
            fs = self._current_feature_summary

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"quote_{quote.part_number or 'export'}_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            )
            if not filepath:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "報價單"

            # Styles
            title_font = Font(name=FONT_FAMILY[0], size=16, bold=True)
            header_font = Font(name=FONT_FAMILY[0], size=11, bold=True)
            normal_font = Font(name=FONT_FAMILY[0], size=10)
            header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
            header_font_white = Font(name=FONT_FAMILY[0], size=10, bold=True, color="ffffff")
            warn_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            row = 1
            # Title
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.cell(row=row, column=1, value="機械加工件智能報價系統 — 報價單").font = title_font
            row += 2

            # Basic info
            info_data = [
                ("圖號", quote.part_number),
                ("材料", quote.material),
                ("規格尺寸", fs.get("bounding_box", "—")),
                ("價格版本", quote.price_version),
                ("生成時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]
            for label, value in info_data:
                ws.cell(row=row, column=1, value=label).font = header_font
                ws.cell(row=row, column=2, value=str(value or "—")).font = normal_font
                row += 1
            row += 1

            # Feature summary
            ws.cell(row=row, column=1, value="Feature 摘要").font = header_font
            row += 1
            for label, value in self._current_feature_summary.items():
                ws.cell(row=row, column=1, value=label).font = normal_font
                ws.cell(row=row, column=2, value=str(value)).font = normal_font
                row += 1
            row += 1

            # Quote items table
            headers = ["序號", "報價項目", "價格來源", "數量", "單位", "單價", "未稅金額", "Confidence", "狀態"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row += 1

            tax = TaxResult.calculate(quote.items, Decimal("0.17"))
            for i, item in enumerate(quote.items, 1):
                is_u = item.source == PriceSource.U
                values = [
                    i,
                    item.name,
                    item.source.value,
                    "—" if is_u else item.quantity,
                    "—" if is_u else item.unit,
                    "—" if is_u else item.unit_price,
                    "待確認" if is_u else item.amount,
                    item.confidence.value,
                    "待確認" if is_u else "已確認",
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.font = normal_font
                    cell.border = thin_border
                    if is_u:
                        cell.fill = warn_fill
                row += 1
            row += 1

            # Tax summary
            tax_rows = [
                ("未稅小計", float(tax.subtotal_excluding_tax)),
                ("稅率", f"{float(tax.tax_rate) * 100:.0f}%"),
                ("稅額", float(tax.tax_amount)),
                ("含稅總價", float(tax.total_including_tax)),
                ("報價完整度", f"{quote.cost_completion:.1f}%"),
                ("價格版本", quote.price_version or "—"),
            ]
            for label, value in tax_rows:
                ws.cell(row=row, column=1, value=label).font = header_font
                cell = ws.cell(row=row, column=2, value=value)
                cell.font = Font(name=FONT_FAMILY[0], size=10, bold=True)
                row += 1

            # Column widths
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 8
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 14
            ws.column_dimensions["H"].width = 12
            ws.column_dimensions["I"].width = 10

            wb.save(filepath)
            messagebox.showinfo("匯出成功", f"報價已匯出至：\n{filepath}")

        except Exception as e:
            messagebox.showerror("匯出失敗", f"Excel 匯出錯誤：{e}")

    # ------------------------------------------------------------------
    # Cleanup
    # ------------------------------------------------------------------

    def _on_close(self) -> None:
        self.destroy()

    # ------------------------------------------------------------------
    # Batch quotation callbacks
    # ------------------------------------------------------------------

    def _batch_scan_dir(self, directory: str, recursive: bool) -> list:
        from quotation.application.file_scanner import FileScanner
        scanner = FileScanner()
        return scanner.scan_directory(Path(directory), recursive=recursive)

    def _batch_scan_files(self, files: list[str]) -> list:
        from quotation.application.file_scanner import FileScanner
        scanner = FileScanner()
        bundles = []
        for f in files:
            bundle = scanner.scan_single_file(Path(f))
            if bundle.files:
                bundles.append(bundle)
        return bundles

    def _batch_run(self, bundles: list, use_ai: bool) -> list:
        from quotation.application.quotation_service import QuotationApplicationService
        from quotation.infrastructure.ai.deepseek_client import DeepSeekClient
        from quotation.infrastructure.secrets.secret_locator import SecretLocator

        ai_client = None
        if use_ai:
            key = SecretLocator.get_deepseek_key()
            if key:
                ai_client = DeepSeekClient(api_key=key)

        svc = QuotationApplicationService(ai_client=ai_client)
        return svc.quote_batch(bundles, use_ai=use_ai)

    def _batch_export_selected(self, results: list, path: str) -> None:
        from quotation.application.batch_excel import export_batch_excel
        export_batch_excel(results, Path(path))

    def _batch_export_all(self, results: list, path: str) -> None:
        from quotation.application.batch_excel import export_batch_excel
        export_batch_excel(results, Path(path))

    def _batch_open_dir(self, path: str) -> None:
        import os
        full = Path(path)
        if full.exists():
            os.startfile(str(full.resolve()))
        else:
            messagebox.showinfo("提示", f"目錄不存在：{path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    """Launch the demo UI application."""
    app = DemoApp()
    app.mainloop()


if __name__ == "__main__":
    main()
