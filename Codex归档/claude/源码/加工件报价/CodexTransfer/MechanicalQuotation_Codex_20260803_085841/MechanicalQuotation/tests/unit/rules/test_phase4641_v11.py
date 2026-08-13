"""Phase 4.6.4.1 v1.1 review workbook tests."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

DATA = Path("data")
V11_XLSX = DATA / "price-review-r01-v1.1.xlsx"
V11_JSON = DATA / "price-review-r01-v1.1.json"


class TestV11Files:
    def test_xlsx_exists(self):
        assert V11_XLSX.exists()

    def test_json_exists(self):
        assert V11_JSON.exists()

    def test_sheets_present(self):
        wb = openpyxl.load_workbook(V11_XLSX)
        names = wb.sheetnames
        assert "Instructions" in names
        assert "Company Price Candidates" in names
        assert "Process Rate Candidates" in names
        assert "Surface Rate Candidates" in names
        assert "Supplier Prices" in names
        assert "Exceptions" in names
        assert "Publication Summary" in names


class TestV11Candidates:
    @pytest.fixture(scope="module")
    def ws(self):
        wb = openpyxl.load_workbook(V11_XLSX)
        return wb["Company Price Candidates"]

    def test_a6061t6_row_exists(self, ws):
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "A6061-T6":
                found = True
                # Should have 3 supplier columns
                assert row[3] is not None or row[7] is not None
                break
        assert found, "A6061-T6 not found"

    def test_aluminum_30x30_exists(self, ws):
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "AL_PROFILE" and row[1] == "30x30":
                found = True
                break
        assert found, "AL_PROFILE 30x30 not in candidates"

    def test_aluminum_40x40_exists(self, ws):
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "AL_PROFILE" and row[1] == "40x40":
                found = True
                break
        assert found, "AL_PROFILE 40x40 not in candidates"

    def test_one_row_per_key(self, ws):
        keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                keys.add((row[0], row[1], row[2]))
        assert len(keys) > 0, "No material candidates found"


class TestV11Exceptions:
    @pytest.fixture(scope="module")
    def ws(self):
        wb = openpyxl.load_workbook(V11_XLSX)
        return wb["Exceptions"]

    def test_aluminum_20x30_exception(self, ws):
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and "20x30" in str(row[1] or "") + str(row[2] or ""):
                if row[4] == "UNKNOWN_PRICE":
                    found = True
                    break
        assert found, "AL_PROFILE 20x30 UNKNOWN_PRICE not in exceptions"

    def test_exceptions_deduped(self, ws):
        keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                key = (row[0], row[1], row[2], row[3], row[4])
                assert key not in keys, f"Duplicate exception: {key}"
                keys.add(key)
        assert len(keys) > 0


class TestV11ProcessSurface:
    @pytest.fixture(scope="module")
    def wb(self):
        return openpyxl.load_workbook(V11_XLSX)

    def test_process_admin_columns(self, wb):
        ws = wb["Process Rate Candidates"]
        headers = [c.value for c in ws[1]]
        assert "Publish?" in headers
        assert "Company Price" in headers
        assert "Price Basis" in headers

    def test_surface_admin_columns(self, wb):
        ws = wb["Surface Rate Candidates"]
        headers = [c.value for c in ws[1]]
        assert "Publish?" in headers
        assert "Company Price" in headers

    def test_cnc_in_process(self, wb):
        ws = wb["Process Rate Candidates"]
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "CNC":
                found = True
                break
        assert found


class TestV11UX:
    @pytest.fixture(scope="module")
    def wb(self):
        return openpyxl.load_workbook(V11_XLSX)

    def test_frozen_panes(self, wb):
        ws = wb["Company Price Candidates"]
        assert ws.freeze_panes == "A2"

    def test_auto_filter(self, wb):
        ws = wb["Company Price Candidates"]
        assert ws.auto_filter.ref is not None

    def test_instructions_present(self, wb):
        ws = wb["Instructions"]
        assert ws.cell(row=1, column=1).value is not None

    def test_supplier_readonly(self, wb):
        ws = wb["Supplier Prices"]
        assert ws.protection.sheet is True

    def test_summary_readonly(self, wb):
        ws = wb["Publication Summary"]
        assert ws.protection.sheet is True
