"""Generate admin price review Excel and JSON from imported R01 package."""

from __future__ import annotations

import json, hashlib
from datetime import datetime, timezone
from pathlib import Path
from collections import defaultdict

PACKAGE = Path("rules/imports/r01-v1.0")
OUT_DIR = Path("data")

def load():
    with open(PACKAGE / "pricing-rules-excel-r01-v1.0.json", encoding="utf-8") as f:
        return json.load(f)

def hash_file(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()

def build_review(data):
    records = data.get("pricing_source_records", [])
    suppliers = data.get("supplier_master", [])
    now = datetime.now(timezone.utc).isoformat()

    # === 1. Company Price Candidates (group by canonical_material + spec + unit) ===
    mat_groups = defaultdict(list)
    for r in records:
        if r.get("target_type") != "MATERIAL":
            continue
        key = (r.get("canonical_material_code","?"), r.get("material_spec") or "",
               r.get("normalized_unit","kg"), r.get("currency","CNY"))
        mat_groups[key].append(r)

    candidates = []
    for (mat, spec, unit, curr), recs in sorted(mat_groups.items()):
        candidates.append({
            "canonical_material_code": mat,
            "material_spec": spec or None,
            "normalized_unit": unit,
            "currency": curr,
            "source_count": len(recs),
            "candidate_sources": [
                {
                    "record_id": r["record_id"],
                    "supplier_id": r.get("supplier_id"),
                    "supplier_name": r.get("supplier_name"),
                    "source_price_book": r.get("source_price_book"),
                    "unit_price": r.get("unit_price"),
                    "tax_inclusion_status": r.get("tax_inclusion_status"),
                    "effective_from": r.get("effective_from"),
                    "source_cell": r.get("source_cell"),
                    "status": r.get("status"),
                    "price_source": r.get("price_source"),
                }
                for r in recs
            ],
            "admin_fields": {
                "publish_as_company_default": False,
                "selected_origin_record_id": None,
                "company_unit_price": None,
                "company_price_basis": None,
                "effective_from": None,
                "effective_to": None,
                "approval_reason": None,
                "approver": None,
                "review_note": None,
            }
        })

    # === 2. Process Rate Candidates ===
    proc_recs = [r for r in records if r.get("target_type") in ("PROCESS",) and r.get("process_code")]
    proc_candidates = []
    for r in proc_recs:
        proc_candidates.append({
            "process_code": r.get("process_code"),
            "source_organization": r.get("source_organization_id"),
            "unit_price": r.get("unit_price"),
            "unit": r.get("normalized_unit"),
            "status": r.get("status"),
            "source_cell": r.get("source_cell"),
            "admin_selection": None,
        })

    # === 3. Surface Rate Candidates ===
    surf_recs = [r for r in records if r.get("target_type") in ("SURFACE",) and r.get("surface_code")]
    surf_candidates = []
    for r in surf_recs:
        surf_candidates.append({
            "surface_code": r.get("surface_code"),
            "source_organization": r.get("source_organization_id"),
            "unit_price": r.get("unit_price"),
            "unit": r.get("normalized_unit"),
            "status": r.get("status"),
            "source_cell": r.get("source_cell"),
            "admin_selection": None,
        })

    # === 4. Supplier Prices (all S records) ===
    supplier_prices = []
    for r in records:
        if r.get("price_source") == "S":
            supplier_prices.append({
                "record_id": r["record_id"],
                "supplier_name": r.get("supplier_name"),
                "material": r.get("canonical_material_code"),
                "spec": r.get("material_spec"),
                "unit_price": r.get("unit_price"),
                "unit": r.get("normalized_unit"),
                "tax": r.get("tax_inclusion_status"),
                "status": r.get("status"),
                "source_cell": r.get("source_cell"),
            })

    # === 5. Exceptions ===
    BLOCKED = {"CONFLICT", "UNIT_CONFLICT", "UNKNOWN_PRICE", "AMBIGUOUS_MATERIAL_SPEC"}
    exceptions = []
    for r in records:
        if r.get("status") in BLOCKED:
            exceptions.append({
                "record_id": r["record_id"],
                "status": r["status"],
                "material": r.get("canonical_material_code") or r.get("process_code") or r.get("surface_code"),
                "supplier": r.get("supplier_name"),
                "issue": "; ".join(r.get("issues", [])) if isinstance(r.get("issues"), list) else str(r.get("issues", "")),
                "source_cell": r.get("source_cell"),
                "publish_allowed": False,
                "admin_resolution": None,
            })
    # Add known exceptions even if not in records
    known_issues = [
        ("SUJ2", "SUP-TONGRUI", "CONFLICT", "Two prices 25 and 15 from same supplier"),
        ("隔熱板", "SUP-LIANGWEI", "UNIT_CONFLICT", "650/m2 vs kg unit"),
        ("鋁型材20x30", None, "UNKNOWN_PRICE", "No supplier price"),
        ("S50C/S45C", "SUP-JMD", "AMBIGUOUS_MATERIAL_SPEC", "Combined spec"),
        ("大理石", None, "UNKNOWN_PRICE", "No supplier price"),
        ("其它", "ORG-WS2-PROCESS", "UNKNOWN_PRICE", "Process rate unknown"),
        ("夾頭", "ORG-WS2-PROCESS", "UNKNOWN_PRICE", "Process rate unknown"),
    ]
    for mat, supp, stat, note in known_issues:
        if not any(e.get("material") == mat and e.get("status") == stat for e in exceptions):
            exceptions.append({
                "material": mat, "supplier": supp, "status": stat,
                "issue": note, "publish_allowed": False, "admin_resolution": None,
            })

    # === 6. Publication Summary ===
    summary = {
        "price_version_id": "R01-COMPANY-PRICE-V1.0-DRAFT",
        "version": "1.0.0-draft",
        "status": "DRAFT",
        "total_candidates": len(candidates),
        "total_process": len(proc_candidates),
        "total_surface": len(surf_candidates),
        "total_supplier_prices": len(supplier_prices),
        "total_exceptions": len(exceptions),
        "publishable_count": 0,
        "blocked_count": len(exceptions),
        "generated_at": now,
        "source_package_sha256": hash_file(PACKAGE / "pricing-rules-excel-r01-v1.0.json"),
        "notes": "All 96 records pending admin review. No effective dates. No auto-selection.",
    }

    review = {
        "generated_at": now,
        "source_package": str(PACKAGE),
        "package_sha256": summary["source_package_sha256"],
        "company_price_candidates": candidates,
        "process_rate_candidates": proc_candidates,
        "surface_rate_candidates": surf_candidates,
        "supplier_prices": supplier_prices,
        "exceptions": exceptions,
        "publication_summary": summary,
    }

    # Save JSON
    OUT_DIR.mkdir(exist_ok=True)
    json_path = OUT_DIR / "price-review-r01-v1.0.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(review, f, ensure_ascii=False, indent=2)

    # Save Company Price Book DRAFT
    book = {
        "price_version_id": "R01-COMPANY-PRICE-V1.0-DRAFT",
        "version": "1.0.0-draft",
        "status": "DRAFT",
        "effective_from": None,
        "created_by": "SYSTEM",
        "approved_by": None,
        "source_package_sha256": summary["source_package_sha256"],
        "record_count": 0,
        "notes": "DRAFT — pending admin review. No C prices published yet.",
        "company_prices": [],
    }
    with open(OUT_DIR / "company-pricebook-r01-v1.0-draft.json", "w", encoding="utf-8") as f:
        json.dump(book, f, ensure_ascii=False, indent=2)

    # Generate Excel
    _generate_xlsx(review)

    return review

def _generate_xlsx(review):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    # Sheet 1: Company Price Candidates
    ws1 = wb.active; ws1.title = "Company Price Candidates"
    ws1.append(["Material", "Spec", "Unit", "Supplier", "Source Book", "Price", "Tax", "Eff Date",
                 "Source Cell", "Status", "Publish?", "Selected Origin", "Company Price", "Basis",
                 "Eff From", "Eff To", "Reason", "Approver", "Note"])
    for c in review["company_price_candidates"]:
        for s in c["candidate_sources"]:
            ws1.append([c["canonical_material_code"], c["material_spec"], c["normalized_unit"],
                        s["supplier_name"], s["source_price_book"], s["unit_price"],
                        s["tax_inclusion_status"], s["effective_from"], s["source_cell"],
                        s["status"], "", "", "", "", "", "", "", "", ""])

    # Sheet 2: Process Rate Candidates
    ws2 = wb.create_sheet("Process Rate Candidates")
    ws2.append(["Process", "Source Org", "Price", "Unit", "Status", "Source Cell", "Admin Selection"])
    for p in review["process_rate_candidates"]:
        ws2.append([p["process_code"], p["source_organization"], p["unit_price"],
                     p["unit"], p["status"], p["source_cell"], ""])

    # Sheet 3: Surface Rate Candidates
    ws3 = wb.create_sheet("Surface Rate Candidates")
    ws3.append(["Surface", "Source Org", "Price", "Unit", "Status", "Source Cell", "Admin Selection"])
    for s in review["surface_rate_candidates"]:
        ws3.append([s["surface_code"], s["source_organization"], s["unit_price"],
                     s["unit"], s["status"], s["source_cell"], ""])

    # Sheet 4: Supplier Prices
    ws4 = wb.create_sheet("Supplier Prices")
    ws4.append(["Record ID", "Supplier", "Material", "Spec", "Price", "Unit", "Tax", "Status", "Cell"])
    for s in review["supplier_prices"]:
        ws4.append([s["record_id"], s["supplier_name"], s["material"], s["spec"],
                     s["unit_price"], s["unit"], s["tax"], s["status"], s["source_cell"]])

    # Sheet 5: Exceptions
    ws5 = wb.create_sheet("Exceptions")
    ws5.append(["Material", "Supplier", "Status", "Issue", "Publish Allowed", "Admin Resolution"])
    for e in review["exceptions"]:
        ws5.append([e["material"], e["supplier"], e["status"], e["issue"],
                     "FALSE", ""])

    # Sheet 6: Publication Summary
    ws6 = wb.create_sheet("Publication Summary")
    for k, v in review["publication_summary"].items():
        ws6.append([k, str(v)])

    wb.save(OUT_DIR / "price-review-r01-v1.0.xlsx")
    print(f"Excel saved: {OUT_DIR / 'price-review-r01-v1.0.xlsx'}")

if __name__ == "__main__":
    data = load()
    review = build_review(data)
    s = review["publication_summary"]
    print(f"Candidates: {s['total_candidates']} | Process: {s['total_process']} | Surface: {s['total_surface']}")
    print(f"Supplier prices: {s['total_supplier_prices']} | Exceptions: {s['total_exceptions']}")
    print(f"Saved: data/price-review-r01-v1.0.json")
    print(f"Saved: data/company-pricebook-r01-v1.0-draft.json")
